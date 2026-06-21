"""Assessment portal endpoints — wraps the sat_scanner package.

Runs Databricks security/compliance scans and UC inventory in the background,
persists results as JSON under ~/.clone-xs/assessment/, and exposes them via
REST endpoints for the /assessment/* frontend portal.
"""

from __future__ import annotations

import asyncio
import json
import re
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from fastapi import APIRouter, BackgroundTasks, Header, HTTPException, Query
from fastapi.responses import HTMLResponse, FileResponse, Response

# ---------------------------------------------------------------------------
# Ensure sat_scanner is importable
# ---------------------------------------------------------------------------
_SAT_PATH = Path.home() / ".clone-xs" / "sat_scanner_path.txt"
_DEFAULT_SAT = Path("/Users/viralkumarjpatel/source/databricks-assesment-tool/assessment")

def _find_sat_path() -> Path | None:
    """Locate the sat_scanner package directory."""
    if _DEFAULT_SAT.exists():
        return _DEFAULT_SAT
    if _SAT_PATH.exists():
        p = Path(_SAT_PATH.read_text().strip())
        if p.exists():
            return p
    return None

_sat_root = _find_sat_path()
if _sat_root and str(_sat_root) not in sys.path:
    sys.path.insert(0, str(_sat_root))

# ---------------------------------------------------------------------------
# Storage paths
# ---------------------------------------------------------------------------
_STORE = Path.home() / ".clone-xs" / "assessment"
_STORE.mkdir(parents=True, exist_ok=True)

# In-memory job tracker: job_id → {"status", "progress", "error", "result_id"}
_JOBS: dict[str, dict] = {}

router = APIRouter()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _scan_dir(scan_id: str) -> Path:
    d = _STORE / scan_id
    d.mkdir(parents=True, exist_ok=True)
    return d


def _list_results() -> list[dict]:
    """Return scan metadata sorted newest-first."""
    items = []
    for p in _STORE.iterdir():
        meta = p / "meta.json"
        if p.is_dir() and meta.exists():
            try:
                items.append(json.loads(meta.read_text()))
            except Exception:
                pass
    return sorted(items, key=lambda x: x.get("scanned_at", ""), reverse=True)


def _latest_result() -> dict | None:
    results = _list_results()
    return results[0] if results else None


def _load_result(scan_id: str) -> dict | None:
    p = _STORE / scan_id / "result.json"
    if not p.exists():
        return None
    try:
        return json.loads(p.read_text())
    except Exception:
        return None


def _html_path(scan_id: str, view: str) -> Path | None:
    """Return the HTML file path for a given view name.

    The exporter generates dynamic filenames like:
      sat-uc-inventory[-{workspace}]-{date}-tree.html
      sat-uc-inventory[-{workspace}]-{date}-star.html   (sunburst)
      sat-uc-inventory[-{workspace}]-{date}-hubspoke.html
      sat[-{workspace}]-{date}.html                     (security report)
    We glob by suffix to find them regardless of workspace name or date.
    """
    scan_dir = _STORE / scan_id
    if not scan_dir.exists():
        return None

    # Inventory views: search for files ending with the known suffix
    suffix_map = {
        "tree":     "-tree.html",
        "sunburst": "-star.html",     # exporter uses "star" not "sunburst"
        "hubspoke": "-hubspoke.html",
        "overview": "-overview.html",
        "topology": "-topology.html",
    }
    if view in suffix_map:
        suffix = suffix_map[view]
        matches = sorted(scan_dir.glob(f"*{suffix}"))
        return matches[0] if matches else None

    if view == "report":
        # Security report: sat[-workspace]-date.html — does NOT end with an
        # inventory suffix.  Prefer an explicitly renamed report.html first.
        explicit = scan_dir / "report.html"
        if explicit.exists():
            return explicit
        inventory_suffixes = set(suffix_map.values())
        candidates = [
            p for p in sorted(scan_dir.glob("sat*.html"))
            if not any(p.name.endswith(s) for s in inventory_suffixes)
        ]
        return candidates[0] if candidates else None

    return None


def _grade(score: float | int) -> str:
    if score >= 90:
        return "A"
    if score >= 75:
        return "B"
    if score >= 60:
        return "C"
    if score >= 45:
        return "D"
    return "F"


def _severity_order(s: str) -> int:
    return {"critical": 0, "high": 1, "medium": 2, "low": 3}.get(s.lower(), 4)


# ---------------------------------------------------------------------------
# Workspace resource collection (raw object lists, separate from UC inventory)
# ---------------------------------------------------------------------------

async def _collect_notebooks_bfs(base: str, hdrs: dict) -> list:
    """2-level BFS workspace listing to collect notebooks and files.

    Root → level-1 dirs → level-2 dirs (capped at 50) so we don't flood
    the API. Returns up to 1000 items total.
    """
    import httpx

    async def list_path(path: str) -> list:
        try:
            async with httpx.AsyncClient(timeout=15) as c:
                r = await c.get(f"{base}/api/2.0/workspace/list", headers=hdrs, params={"path": path})
                return r.json().get("objects", []) if r.status_code == 200 else []
        except Exception:
            return []

    level0 = await list_path("/")
    l1_dirs = [o for o in level0 if o.get("object_type") == "DIRECTORY"]
    l1_items = list(await asyncio.gather(*[list_path(d["path"]) for d in l1_dirs]))

    found: list = []
    subdirs: list = []
    for items in l1_items:
        for obj in items:
            ot = obj.get("object_type")
            if ot in ("NOTEBOOK", "FILE"):
                found.append(obj)
            elif ot == "DIRECTORY":
                subdirs.append(obj)

    for items in await asyncio.gather(*[list_path(d["path"]) for d in subdirs[:50]]):
        for obj in items:
            if obj.get("object_type") in ("NOTEBOOK", "FILE"):
                found.append(obj)

    return found[:1000]


async def _collect_workspace_resources(host: str, token: str) -> dict:
    """Fetch raw workspace resource lists from Databricks REST APIs.

    Called during every scan using the live credentials. Results saved to
    workspace_resources.json alongside result.json / inventory.json.
    All failures are silently swallowed so they never abort the main scan.
    """
    import httpx  # available in venv

    base = host.rstrip("/")
    hdrs = {"Authorization": f"Bearer {token}"}
    resources: dict[str, Any] = {}

    async def get(path: str, params: dict | None = None) -> dict | None:
        try:
            async with httpx.AsyncClient(timeout=20) as c:
                r = await c.get(f"{base}{path}", headers=hdrs, params=params)
                if r.status_code == 200:
                    return r.json()
        except Exception:
            pass
        return None

    async def collect(key: str, path: str, list_key: str, params: dict | None = None):
        resp = await get(path, params)
        if not list_key:
            resources[key] = resp if isinstance(resp, list) else []
        else:
            resources[key] = (resp or {}).get(list_key, [])

    await asyncio.gather(
        collect("jobs",                    "/api/2.1/jobs/list",                        "jobs",                       {"expand_tasks": "false", "limit": 200}),
        collect("job_runs",                "/api/2.1/jobs/runs/list",                   "runs",                       {"limit": 50}),
        collect("clusters",                "/api/2.0/clusters/list",                    "clusters"),
        collect("cluster_policies",        "/api/2.0/policies/clusters/list",           "policies"),
        collect("instance_pools",          "/api/2.0/instance-pools/list",              "instance_pools"),
        collect("warehouses",              "/api/2.0/sql/warehouses",                   "warehouses"),
        collect("serving_endpoints",       "/api/2.0/serving-endpoints",                "endpoints"),
        collect("tokens",                  "/api/2.0/token-management/tokens",          "token_infos"),
        collect("users",                   "/api/2.0/preview/scim/v2/Users",            "Resources",                  {"count": 200}),
        collect("groups",                  "/api/2.0/preview/scim/v2/Groups",           "Resources",                  {"count": 200}),
        collect("service_principals",      "/api/2.0/preview/scim/v2/ServicePrincipals","Resources",                  {"count": 200}),
        collect("pipelines",               "/api/2.0/pipelines",                        "statuses"),
        collect("secret_scopes",           "/api/2.0/secrets/scopes/list",              "scopes"),
        collect("repos",                   "/api/2.0/repos",                            "repos"),
        collect("apps",                    "/api/2.0/apps",                             "apps"),
        collect("genie_spaces",            "/api/2.0/genie/spaces",                     "spaces"),
        collect("dashboards",              "/api/2.0/lakeview/dashboards",              "dashboards"),
        collect("vector_search",           "/api/2.0/vector-search/endpoints",          "endpoints"),
        collect("global_init_scripts",     "/api/2.0/global-init-scripts",              "scripts"),
        collect("experiments",             "/api/2.0/mlflow/experiments/search",        "experiments",                {"max_results": 200, "view_type": "ACTIVE_ONLY"}),
        collect("sql_queries",             "/api/2.0/sql/queries",                      "results",                    {"page_size": 100}),
        collect("sql_alerts",              "/api/2.0/sql/alerts",                       ""),
        collect("dbfs_mounts",             "/api/2.0/dbfs/list",                        "files",                      {"path": "/mnt"}),
        collect("marketplace_listings",    "/api/2.1/marketplace-consumer/listings",    "listings",                   {"page_size": 50}),
        collect("notification_destinations","/api/2.0/notification-destinations",       "notification_destinations",  {"page_size": 100}),
        collect("git_credentials",         "/api/2.0/git-credentials",                  "credentials"),
    )

    resources["notebooks"] = await _collect_notebooks_bfs(base, hdrs)

    return resources


# ---------------------------------------------------------------------------
# Background scan runner
# ---------------------------------------------------------------------------

async def _run_inventory_only(
    job_id: str,
    scan_id: str,
    host: str,
    token: str,
    workspace_name: str,
) -> None:
    """Run UC inventory scan only — no security checks."""
    try:
        _JOBS[job_id]["status"] = "running"
        _JOBS[job_id]["progress"] = "Importing sat_scanner…"

        try:
            from sat_scanner.inventory import run_inventory
            from sat_scanner.exporters import export_inventory_hierarchy_html
        except ImportError as e:
            _JOBS[job_id]["status"] = "error"
            _JOBS[job_id]["error"] = f"sat_scanner package not found: {e}."
            return

        _JOBS[job_id]["progress"] = "Scanning Unity Catalog objects…"
        inv = await run_inventory(
            host=host.rstrip("/"),
            token=token,
            workspace_name=workspace_name,
            quiet=True,
            grants="coarse",
        )

        out_dir = _scan_dir(scan_id)
        inv_dict = inv.to_dict()

        _JOBS[job_id]["progress"] = "Generating visualisations…"
        (out_dir / "inventory.json").write_text(json.dumps(inv_dict, default=str, indent=2))
        export_inventory_hierarchy_html(inv, out_dir)

        # Collect raw workspace resources (jobs, endpoints, tokens, etc.)
        _JOBS[job_id]["progress"] = "Collecting workspace resources…"
        try:
            ws_resources = await _collect_workspace_resources(host, token)
            (out_dir / "workspace_resources.json").write_text(
                json.dumps(ws_resources, default=str, indent=2)
            )
        except Exception:
            pass

        now = datetime.now(timezone.utc).isoformat()
        meta = {
            "scan_id": scan_id,
            "workspace_url": host,
            "workspace_name": workspace_name,
            "scanned_at": inv_dict.get("scanned_at", now),
            "overall_score": None,
            "grade": None,
            "total_checks": 0,
            "passed": 0,
            "failed": 0,
            "warnings": 0,
            "not_applicable": 0,
            "scan_type": "inventory",
            "with_inventory": True,
            "catalog_count": inv_dict.get("catalog_count", 0),
            "schema_count": inv_dict.get("schema_count", 0),
            "table_count": inv_dict.get("table_count", 0),
        }
        (out_dir / "meta.json").write_text(json.dumps(meta, indent=2))

        await _fire_scan_webhooks(scan_id, meta)
        _JOBS[job_id]["status"] = "completed"
        _JOBS[job_id]["progress"] = "Done"
        _JOBS[job_id]["result_id"] = scan_id

    except Exception as exc:
        _JOBS[job_id]["status"] = "error"
        _JOBS[job_id]["error"] = str(exc)


async def _run_scan_task(
    job_id: str,
    scan_id: str,
    host: str,
    token: str,
    workspace_name: str,
    scan_type: str,  # "full" | "security" | "inventory"
) -> None:
    """Run sat_scanner scan (and optionally UC inventory) then persist results."""
    if scan_type == "inventory":
        await _run_inventory_only(job_id, scan_id, host, token, workspace_name)
        return

    with_inventory = scan_type == "full"

    try:
        _JOBS[job_id]["status"] = "running"
        _JOBS[job_id]["progress"] = "Importing sat_scanner…"

        try:
            from sat_scanner.scanner import run_scan
            from sat_scanner.exporters import (
                export_json, export_html, export_inventory_hierarchy_html,
            )
            from sat_scanner.models import SATScanResult
        except ImportError as e:
            _JOBS[job_id]["status"] = "error"
            _JOBS[job_id]["error"] = (
                f"sat_scanner package not found: {e}. "
                f"Expected at {_DEFAULT_SAT}"
            )
            return

        _JOBS[job_id]["progress"] = "Running security scan (345 checks)…"
        result: SATScanResult = await run_scan(
            host=host.rstrip("/"),
            token=token,
            workspace_name=workspace_name,
            quiet=True,
        )

        out_dir = _scan_dir(scan_id)

        _JOBS[job_id]["progress"] = "Saving results…"
        result_dict = result.to_dict()
        (out_dir / "result.json").write_text(json.dumps(result_dict, default=str, indent=2))

        # Generate HTML report — exporter writes sat[-workspace]-{date}.html;
        # _html_path("report") finds it via glob so no rename needed.
        try:
            export_html(result, out_dir)
        except Exception:
            pass

        # Collect raw workspace resources in parallel with optional UC inventory
        _JOBS[job_id]["progress"] = "Collecting workspace resources…"
        ws_task = asyncio.create_task(_collect_workspace_resources(host, token))

        # UC Inventory
        if with_inventory:
            _JOBS[job_id]["progress"] = "Running UC inventory scan…"
            try:
                from sat_scanner.inventory import run_inventory
                inv = await run_inventory(
                    host=host.rstrip("/"),
                    token=token,
                    workspace_name=workspace_name,
                    quiet=True,
                    grants="coarse",
                )
                (out_dir / "inventory.json").write_text(
                    json.dumps(inv.to_dict(), default=str, indent=2)
                )
                export_inventory_hierarchy_html(inv, out_dir)
            except Exception as inv_err:
                _JOBS[job_id]["inventory_error"] = str(inv_err)

        # Save workspace resources (collected in background above)
        try:
            ws_resources = await ws_task
            (out_dir / "workspace_resources.json").write_text(
                json.dumps(ws_resources, default=str, indent=2)
            )
        except Exception:
            pass

        overall = result_dict.get("overall_score", 0)
        meta = {
            "scan_id": scan_id,
            "workspace_url": host,
            "workspace_name": workspace_name,
            "scanned_at": result_dict.get("scanned_at", datetime.now(timezone.utc).isoformat()),
            "overall_score": overall,
            "grade": _grade(overall),
            "total_checks": result_dict.get("total_checks", 0),
            "passed": result_dict.get("passed", 0),
            "failed": result_dict.get("failed", 0),
            "warnings": result_dict.get("warnings", 0),
            "not_applicable": result_dict.get("not_applicable", 0),
            "scan_type": scan_type,
            "with_inventory": with_inventory,
        }
        (out_dir / "meta.json").write_text(json.dumps(meta, indent=2))

        await _fire_scan_webhooks(scan_id, meta)
        _JOBS[job_id]["status"] = "completed"
        _JOBS[job_id]["progress"] = "Done"
        _JOBS[job_id]["result_id"] = scan_id

    except Exception as exc:
        _JOBS[job_id]["status"] = "error"
        _JOBS[job_id]["error"] = str(exc)


# ---------------------------------------------------------------------------
# Feature 5 helper — fire webhooks on scan completion
# ---------------------------------------------------------------------------

async def _fire_scan_webhooks(scan_id: str, meta: dict) -> None:
    """POST scan summary to all configured webhooks (non-blocking, never raises)."""
    try:
        wh_file = Path("config/webhooks.json")
        if not wh_file.exists():
            return
        webhooks = json.loads(wh_file.read_text())
        if not webhooks:
            return
        import httpx
        payload = {
            "event": "scan_complete",
            "scan_id": scan_id,
            "workspace": meta.get("workspace_name") or meta.get("workspace_url", ""),
            "score": meta.get("overall_score"),
            "grade": meta.get("grade"),
            "passed": meta.get("passed"),
            "failed": meta.get("failed"),
            "scan_type": meta.get("scan_type"),
            "scanned_at": meta.get("scanned_at"),
        }
        async with httpx.AsyncClient(timeout=10) as c:
            for wh in webhooks:
                try:
                    await c.post(
                        wh.get("url", ""), json=payload,
                        headers={"Content-Type": "application/json"},
                    )
                except Exception:
                    pass
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Scan scheduler background task (Feature 3)
# ---------------------------------------------------------------------------

_SCHEDULE_PATH = _STORE.parent / "scan_schedule.json"


async def start_scan_scheduler() -> None:
    """Check every 60 s whether a scheduled scan is due and trigger it."""
    from datetime import timedelta
    while True:
        await asyncio.sleep(60)
        try:
            if not _SCHEDULE_PATH.exists():
                continue
            cfg = json.loads(_SCHEDULE_PATH.read_text())
            if not cfg.get("enabled"):
                continue
            host = cfg.get("host", "")
            token = cfg.get("token", "")
            if not host or not token:
                continue
            next_run_str = cfg.get("next_run")
            if next_run_str:
                next_run = datetime.fromisoformat(next_run_str)
                if next_run.tzinfo is None:
                    next_run = next_run.replace(tzinfo=timezone.utc)
                if next_run > datetime.now(timezone.utc):
                    continue
            # Trigger a new scan
            job_id = str(uuid.uuid4())
            scan_id = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S") + "_" + job_id[:8]
            _JOBS[job_id] = {
                "job_id": job_id,
                "scan_id": scan_id,
                "scan_type": cfg.get("scan_type", "full"),
                "status": "queued",
                "progress": "Scheduled scan queued…",
                "submitted_at": datetime.now(timezone.utc).isoformat(),
                "result_id": None,
                "error": None,
            }
            asyncio.create_task(
                _run_scan_task(
                    job_id, scan_id,
                    host, token,
                    cfg.get("workspace_name", ""),
                    cfg.get("scan_type", "full"),
                )
            )
            # Advance next_run
            freq = cfg.get("frequency", "daily")
            delta = timedelta(days=1 if freq == "daily" else 7)
            cfg["next_run"] = (datetime.now(timezone.utc) + delta).isoformat()
            cfg["last_triggered"] = datetime.now(timezone.utc).isoformat()
            _SCHEDULE_PATH.write_text(json.dumps(cfg, indent=2))
        except Exception:
            pass


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@router.post("/run")
async def run_assessment(
    background_tasks: BackgroundTasks,
    x_databricks_host: str | None = Header(None),
    x_databricks_token: str | None = Header(None),
    workspace_name: str = Query(""),
    scan_type: str = Query("full"),  # "full" | "security" | "inventory"
):
    """Trigger an async assessment scan. Returns job_id to poll.

    scan_type:
      - full       — 345 security checks + UC inventory (default)
      - security   — 345 security checks only
      - inventory  — UC inventory only (no security checks)
    """
    host = x_databricks_host or ""
    token = x_databricks_token or ""
    if not host or not token:
        raise HTTPException(status_code=401, detail="Databricks host and token required")
    if scan_type not in ("full", "security", "inventory"):
        raise HTTPException(status_code=400, detail="scan_type must be full, security, or inventory")

    job_id = str(uuid.uuid4())
    scan_id = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S") + "_" + job_id[:8]
    _JOBS[job_id] = {
        "job_id": job_id,
        "scan_id": scan_id,
        "scan_type": scan_type,
        "status": "queued",
        "progress": "Queued…",
        "submitted_at": datetime.now(timezone.utc).isoformat(),
        "result_id": None,
        "error": None,
    }
    background_tasks.add_task(
        _run_scan_task, job_id, scan_id, host, token, workspace_name, scan_type
    )
    return {"job_id": job_id, "scan_id": scan_id, "scan_type": scan_type, "status": "queued"}


@router.get("/status/{job_id}")
async def get_job_status(job_id: str):
    """Poll scan job status."""
    job = _JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return job


@router.get("/results")
async def list_results():
    """List all past scan results (metadata only), newest first."""
    return _list_results()


@router.get("/results/{scan_id}")
async def get_result(scan_id: str):
    """Get full result JSON for a specific scan."""
    data = _load_result(scan_id)
    if not data:
        raise HTTPException(status_code=404, detail="Scan result not found")
    return data


@router.get("/latest")
async def get_latest():
    """Return metadata + findings summary for the most recent scan."""
    meta = _latest_result()
    if not meta:
        return None

    scan_id = meta.get("scan_id", "")
    result = _load_result(scan_id)
    if not result:
        return meta

    # Attach category scores and top findings to the meta
    meta["category_scores"] = result.get("category_scores", {})
    meta["findings_preview"] = [
        f for f in result.get("findings", [])
        if f.get("status") == "FAIL" and f.get("severity", "").lower() == "critical"
    ][:5]
    return meta


@router.get("/findings")
async def get_findings(
    scan_id: str | None = Query(None),
    severity: str | None = Query(None),
    category: str | None = Query(None),
    status: str | None = Query(None),
):
    """Return filtered findings list. Uses latest scan when scan_id is omitted."""
    if scan_id:
        result = _load_result(scan_id)
    else:
        meta = _latest_result()
        result = _load_result(meta["scan_id"]) if meta else None

    if not result:
        return []

    findings: list[dict] = result.get("findings", [])

    if severity:
        sevs = {s.strip().lower() for s in severity.split(",")}
        findings = [f for f in findings if f.get("severity", "").lower() in sevs]
    if category:
        cats = {c.strip().lower() for c in category.split(",")}
        findings = [f for f in findings if f.get("category", "").lower() in cats]
    if status:
        statuses = {s.strip().upper() for s in status.split(",")}
        findings = [f for f in findings if f.get("status", "").upper() in statuses]

    # Sort by severity then status
    findings.sort(key=lambda f: (
        _severity_order(f.get("severity", "")),
        0 if f.get("status") == "FAIL" else 1,
    ))
    return findings


@router.get("/categories")
async def get_categories(scan_id: str | None = Query(None)):
    """Return per-category scores for the latest (or specified) scan."""
    if scan_id:
        result = _load_result(scan_id)
    else:
        meta = _latest_result()
        result = _load_result(meta["scan_id"]) if meta else None

    if not result:
        return []

    scores: dict = result.get("category_scores", {})
    findings: list[dict] = result.get("findings", [])

    # Build per-category count breakdown
    cat_counts: dict[str, dict] = {}
    for f in findings:
        cat = f.get("category", "Unknown")
        if cat not in cat_counts:
            cat_counts[cat] = {"passed": 0, "failed": 0, "warnings": 0, "not_applicable": 0}
        st = f.get("status", "").upper()
        if st == "PASS":
            cat_counts[cat]["passed"] += 1
        elif st == "FAIL":
            cat_counts[cat]["failed"] += 1
        elif st == "WARN":
            cat_counts[cat]["warnings"] += 1
        else:
            cat_counts[cat]["not_applicable"] += 1

    return [
        {
            "category": cat,
            "score": scores.get(cat, 0),
            "grade": _grade(scores.get(cat, 0)),
            **cat_counts.get(cat, {"passed": 0, "failed": 0, "warnings": 0, "not_applicable": 0}),
        }
        for cat in sorted(scores.keys())
    ]


@router.get("/recommendations")
async def get_recommendations(scan_id: str | None = Query(None)):
    """Return prioritised recommendations (FAIL findings, deduplicated by title)."""
    if scan_id:
        result = _load_result(scan_id)
    else:
        meta = _latest_result()
        result = _load_result(meta["scan_id"]) if meta else None

    if not result:
        return []

    findings: list[dict] = result.get("findings", [])

    # Group FAIL/WARN findings by title and deduplicate
    seen: dict[str, dict] = {}
    for f in findings:
        if f.get("status") not in ("FAIL", "WARN"):
            continue
        key = f.get("title", f.get("check_id", ""))
        if key not in seen:
            seen[key] = {
                "title": f.get("title", ""),
                "category": f.get("category", ""),
                "severity": f.get("severity", ""),
                "status": f.get("status", ""),
                "recommendation": f.get("recommendation", ""),
                "effort": f.get("effort", ""),
                "benefits": f.get("benefits", ""),
                "reference_url": f.get("reference_url", ""),
                "count": 1,
                "priority": _severity_order(f.get("severity", "")),
            }
        else:
            seen[key]["count"] += 1

    recs = sorted(seen.values(), key=lambda r: (r["priority"], r["status"] != "FAIL"))
    for i, r in enumerate(recs, 1):
        r["rank"] = i
    return recs


@router.get("/inventory")
async def get_inventory(scan_id: str | None = Query(None)):
    """Return UC inventory summary JSON."""
    sid = scan_id
    if not sid:
        meta = _latest_result()
        sid = meta["scan_id"] if meta else None
    if not sid:
        return None

    p = _STORE / sid / "inventory.json"
    if not p.exists():
        return None
    try:
        return json.loads(p.read_text())
    except Exception:
        return None


@router.get("/inventory/export")
async def export_inventory(
    fmt: str = Query("json"),
    scan_id: str | None = Query(None),
):
    """Download UC inventory in the requested format: json | csv_tables | csv_columns | excel."""
    sid = scan_id or ((_latest_result() or {}).get("scan_id"))
    if not sid:
        raise HTTPException(status_code=404, detail="No scan found")

    inv_path = _STORE / sid / "inventory.json"
    if not inv_path.exists():
        raise HTTPException(
            status_code=404,
            detail="Inventory not available for this scan — run a scan with UC Inventory enabled",
        )

    inv = json.loads(inv_path.read_text())

    if fmt == "json":
        return FileResponse(
            str(inv_path),
            media_type="application/json",
            filename=f"uc_inventory_{sid}.json",
        )

    elif fmt == "csv_tables":
        import csv
        import io as _io
        buf = _io.StringIO()
        w = csv.writer(buf)
        w.writerow([
            "catalog", "catalog_type", "schema", "table", "full_name",
            "table_type", "data_format", "owner", "comment",
            "storage_location", "created_at", "updated_at", "column_count", "grant_count",
        ])
        for cat in inv.get("catalogs", []):
            for sch in cat.get("schemas", []):
                for tbl in sch.get("tables", []):
                    w.writerow([
                        cat["name"], cat.get("catalog_type", ""), sch["name"],
                        tbl["name"], tbl.get("full_name", ""), tbl.get("table_type", ""),
                        tbl.get("data_source_format", ""), tbl.get("owner", ""),
                        tbl.get("comment", ""), tbl.get("storage_location", ""),
                        tbl.get("created_at", ""), tbl.get("updated_at", ""),
                        len(tbl.get("columns", [])), len(tbl.get("grants", [])),
                    ])
        return Response(
            content=buf.getvalue().encode(),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="uc_tables_{sid}.csv"'},
        )

    elif fmt == "csv_columns":
        import csv
        import io as _io
        buf = _io.StringIO()
        w = csv.writer(buf)
        w.writerow([
            "catalog", "schema", "table", "full_table_name",
            "column", "position", "type", "nullable", "comment", "masked",
        ])
        for cat in inv.get("catalogs", []):
            for sch in cat.get("schemas", []):
                for tbl in sch.get("tables", []):
                    for col in tbl.get("columns", []):
                        w.writerow([
                            cat["name"], sch["name"], tbl["name"],
                            tbl.get("full_name", ""), col["name"],
                            col.get("position", ""), col.get("type_text", ""),
                            col.get("nullable", ""), col.get("comment", ""),
                            "yes" if col.get("mask") else "no",
                        ])
        return Response(
            content=buf.getvalue().encode(),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="uc_columns_{sid}.csv"'},
        )

    elif fmt == "excel":
        try:
            import io as _io
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise HTTPException(status_code=503, detail="openpyxl not installed — cannot generate Excel export")

        HEADER_FILL = PatternFill("solid", fgColor="E8453C")
        HEADER_FONT = Font(color="FFFFFF", bold=True)

        def _add_sheet(wb, name, headers, rows):
            ws = wb.create_sheet(name)
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            for row in rows:
                ws.append(row)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        _add_sheet(wb, "Catalogs",
            ["Name", "Type", "Owner", "Comment", "Storage Root", "Isolation Mode"],
            [[c["name"], c.get("catalog_type", ""), c.get("owner", ""), c.get("comment", ""),
              c.get("storage_root", ""), c.get("isolation_mode", "")]
             for c in inv.get("catalogs", [])])

        _add_sheet(wb, "Schemas",
            ["Full Name", "Catalog", "Schema", "Owner", "Comment"],
            [[s.get("full_name", ""), cat["name"], s["name"], s.get("owner", ""), s.get("comment", "")]
             for cat in inv.get("catalogs", []) for s in cat.get("schemas", [])])

        _add_sheet(wb, "Tables",
            ["Full Name", "Catalog", "Schema", "Table", "Type", "Format", "Owner", "Comment",
             "Storage Location", "Columns", "Grants", "Created", "Updated"],
            [[t.get("full_name", ""), cat["name"], sch["name"], t["name"],
              t.get("table_type", ""), t.get("data_source_format", ""),
              t.get("owner", ""), t.get("comment", ""), t.get("storage_location", ""),
              len(t.get("columns", [])), len(t.get("grants", [])),
              t.get("created_at", ""), t.get("updated_at", "")]
             for cat in inv.get("catalogs", []) for sch in cat.get("schemas", [])
             for t in sch.get("tables", [])])

        _add_sheet(wb, "Columns",
            ["Table", "Catalog", "Schema", "Column", "Position", "Type", "Nullable", "Comment", "Masked"],
            [[t.get("full_name", ""), cat["name"], sch["name"], col["name"],
              col.get("position", ""), col.get("type_text", ""), col.get("nullable", ""),
              col.get("comment", ""), "yes" if col.get("mask") else "no"]
             for cat in inv.get("catalogs", []) for sch in cat.get("schemas", [])
             for t in sch.get("tables", []) for col in t.get("columns", [])])

        grant_rows: list[list] = []
        for g in inv.get("metastore_grants", []):
            grant_rows.append(["METASTORE", "—", g.get("full_name", ""), g.get("principal", ""),
                                ", ".join(g.get("privileges", [])), g.get("inherited_from", "")])
        for cat in inv.get("catalogs", []):
            for g in cat.get("grants", []):
                grant_rows.append(["CATALOG", cat["name"], g.get("full_name", ""), g.get("principal", ""),
                                    ", ".join(g.get("privileges", [])), g.get("inherited_from", "")])
            for sch in cat.get("schemas", []):
                for g in sch.get("grants", []):
                    grant_rows.append(["SCHEMA", cat["name"], g.get("full_name", ""), g.get("principal", ""),
                                        ", ".join(g.get("privileges", [])), g.get("inherited_from", "")])
                for t in sch.get("tables", []):
                    for g in t.get("grants", []):
                        grant_rows.append(["TABLE", cat["name"], g.get("full_name", ""), g.get("principal", ""),
                                            ", ".join(g.get("privileges", [])), g.get("inherited_from", "")])
        _add_sheet(wb, "Grants",
            ["Level", "Catalog", "Object", "Principal", "Privileges", "Inherited From"],
            grant_rows)

        _add_sheet(wb, "External Locations",
            ["Name", "URL", "Credential", "Read Only", "Owner", "Comment"],
            [[e.get("name", ""), e.get("url", ""), e.get("credential_name", ""),
              str(e.get("read_only", "")), e.get("owner", ""), e.get("comment", "")]
             for e in inv.get("external_locations", [])])

        buf = _io.BytesIO()
        wb.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="uc_inventory_{sid}.xlsx"'},
        )

    elif fmt == "html":
        import tempfile
        import zipfile
        # Collect UC inventory HTML views only (not the security findings report)
        html_files: list[Path] = []
        for view in ("overview", "tree", "sunburst", "hubspoke", "topology"):
            p = _html_path(sid, view)
            if p and p.exists():
                html_files.append(p)
        if not html_files:
            raise HTTPException(
                status_code=404,
                detail="No inventory HTML found — run a scan with UC Inventory enabled",
            )
        if len(html_files) == 1:
            return FileResponse(str(html_files[0]), media_type="text/html",
                                filename=f"uc_inventory_{sid}.html")
        # Multiple files — bundle as ZIP (temp file, then FileResponse for reliable streaming)
        tmp = tempfile.NamedTemporaryFile(suffix=".zip", delete=False)
        tmp.close()
        with zipfile.ZipFile(tmp.name, "w", zipfile.ZIP_DEFLATED) as zf:
            for p in html_files:
                zf.write(str(p), arcname=p.name)
        return FileResponse(
            tmp.name,
            media_type="application/zip",
            filename=f"uc_inventory_{sid}.zip",
        )

    raise HTTPException(status_code=400, detail=f"Unknown format: {fmt}")


@router.get("/workspace-resources")
async def get_workspace_resources(
    scan_id: str | None = Query(None),
    resource_type: str | None = Query(None),
):
    """Return raw workspace resource lists (jobs, clusters, tokens, etc.).

    Populated during every scan. Pass ?resource_type=jobs to get just one list.
    Returns 404 if the scan predates this feature — re-run to populate.
    """
    sid = scan_id
    if not sid:
        meta = _latest_result()
        sid = meta["scan_id"] if meta else None
    if not sid:
        raise HTTPException(status_code=404, detail="No scan results found")

    path = _STORE / sid / "workspace_resources.json"
    if not path.exists():
        raise HTTPException(
            status_code=404,
            detail="No workspace resource data for this scan. Re-run the scan to populate.",
        )
    try:
        data = json.loads(path.read_text())
    except Exception:
        raise HTTPException(status_code=500, detail="Failed to read workspace resources")

    if resource_type:
        if resource_type not in data:
            raise HTTPException(status_code=404, detail=f"Resource type '{resource_type}' not found")
        return data[resource_type]
    return data


@router.post("/collect-resources")
async def collect_resources(
    x_databricks_host: str | None = Header(None),
    x_databricks_token: str | None = Header(None),
    scan_id: str | None = Query(None),
):
    """Collect raw workspace resources for an existing scan (no full re-scan needed).

    Supply Databricks credentials via X-Databricks-Host and X-Databricks-Token headers.
    Saves workspace_resources.json to the scan directory (default: latest scan).
    """
    host = x_databricks_host or ""
    token = x_databricks_token or ""
    if not host or not token:
        raise HTTPException(status_code=401, detail="Databricks host and token required")

    sid = scan_id
    if not sid:
        meta = _latest_result()
        sid = meta["scan_id"] if meta else None
    if not sid:
        raise HTTPException(status_code=404, detail="No scan found to attach resources to")

    try:
        ws_resources = await _collect_workspace_resources(host.rstrip("/"), token)
        path = _STORE / sid / "workspace_resources.json"
        path.write_text(json.dumps(ws_resources, default=str, indent=2))
        summary = {k: len(v) for k, v in ws_resources.items() if isinstance(v, list)}
        return {"scan_id": sid, "resources_collected": summary}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/html/{view}", response_class=HTMLResponse)
async def serve_html(
    view: str,
    scan_id: str | None = Query(None),
):
    """Serve a generated HTML dashboard view.

    view options: tree | sunburst | hubspoke | overview | report
    """
    sid = scan_id
    if not sid:
        meta = _latest_result()
        sid = meta["scan_id"] if meta else None

    if not sid:
        return HTMLResponse(
            "<html><body style='font-family:sans-serif;padding:2rem'>"
            "<h2>No assessment results yet</h2>"
            "<p>Run an assessment from the <a href='/assessment/run'>Run Scan</a> page first.</p>"
            "</body></html>",
            status_code=200,
        )

    html_file = _html_path(sid, view)
    if not html_file or not html_file.exists():
        if view == "report":
            msg = (
                "<h2 style='color:#b45309'>No Security Report Available</h2>"
                "<p>This scan was an <strong>inventory-only</strong> scan — no security checks were run.</p>"
                "<p>To generate a security report, go back to "
                "<a href='/assessment/run'>Run Scan</a> and choose "
                "<strong>Full Assessment</strong> or <strong>Security Checks Only</strong>.</p>"
            )
        else:
            msg = (
                f"<h2>{view.title()} view not available</h2>"
                "<p>This view requires a UC inventory scan. Re-run with "
                "<strong>Include UC Inventory</strong> enabled, or check that "
                "the sat_scanner package is installed.</p>"
            )
        return HTMLResponse(
            f"<html><body style='font-family:sans-serif;padding:2rem'>{msg}</body></html>",
            status_code=200,
        )

    html_content = html_file.read_text(encoding="utf-8")
    html_content = _rewrite_nav_links(html_content)
    return HTMLResponse(html_content)


def _rewrite_nav_links(html: str) -> str:
    """Replace relative cross-view file hrefs with portal API endpoint URLs.

    The sat_scanner exporter writes links like:
      href="sat-uc-inventory-2026-06-21-star.html"
    which break when served via /api/assessment/html/{view}.
    Rewrite them to the canonical API paths so in-iframe nav works.
    """
    replacements = [
        (r'href="[^"]*-tree\.html"',     'href="/api/assessment/html/tree"'),
        (r'href="[^"]*-star\.html"',     'href="/api/assessment/html/sunburst"'),
        (r'href="[^"]*-hubspoke\.html"', 'href="/api/assessment/html/hubspoke"'),
        (r'href="[^"]*-overview\.html"', 'href="/api/assessment/html/overview"'),
        (r'href="[^"]*-topology\.html"', 'href="/api/assessment/html/topology"'),
        # "Report" link: sat-uc-inventory-YYYY-MM-DD.html (date-only suffix, no view name)
        (r'href="sat-[^"]*-\d{4}-\d{2}-\d{2}\.html"', 'href="/api/assessment/html/report"'),
    ]
    for pattern, replacement in replacements:
        html = re.sub(pattern, replacement, html)
    return html


@router.get("/remediation/{scan_id}")
async def get_remediation(scan_id: str):
    """Return all remediation statuses for a scan as {check_id: {status, note}}."""
    path = _STORE / scan_id / "remediation.json"
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


@router.put("/remediation/{scan_id}/{check_id}")
async def update_remediation(scan_id: str, check_id: str, body: dict):
    """Persist remediation status for a single finding."""
    scan_dir = _STORE / scan_id
    if not scan_dir.exists():
        raise HTTPException(status_code=404, detail=f"Scan {scan_id} not found")
    path = scan_dir / "remediation.json"
    data: dict = {}
    if path.exists():
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            data = {}
    data[check_id] = {
        "status": body.get("status", "open"),
        "note": body.get("note", ""),
    }
    path.write_text(json.dumps(data, indent=2), encoding="utf-8")
    return data[check_id]


@router.get("/export/{fmt}")
async def export_result(
    fmt: str,
    scan_id: str | None = Query(None),
):
    """Download a scan result in the requested format: json | csv | excel | html."""
    sid = scan_id
    if not sid:
        meta = _latest_result()
        sid = meta["scan_id"] if meta else None

    if not sid:
        raise HTTPException(status_code=404, detail="No assessment results available")

    out_dir = _STORE / sid

    if fmt == "json":
        p = out_dir / "result.json"
        if not p.exists():
            raise HTTPException(status_code=404, detail="Result not found")
        return FileResponse(str(p), media_type="application/json", filename=f"assessment_{sid}.json")

    if fmt == "html":
        p = _html_path(sid, "report")
        if not p or not p.exists():
            raise HTTPException(status_code=404, detail="HTML report not generated — run a full or security-only scan first")
        return FileResponse(str(p), media_type="text/html", filename=f"assessment_{sid}.html")

    # For CSV and Excel we need to regenerate from result
    result_data = _load_result(sid)
    if not result_data:
        raise HTTPException(status_code=404, detail="Result data not found")

    try:
        from sat_scanner.models import SATScanResult, SATFinding
        from sat_scanner.exporters import export_csv, export_excel
        import tempfile

        # SATFinding uses __slots__ (not a dataclass), so filter keys against __slots__
        _slots = set(SATFinding.__slots__)
        findings = [
            SATFinding(**{k: v for k, v in f.items() if k in _slots})
            for f in result_data.get("findings", [])
        ]
        result_obj = SATScanResult(
            workspace_url=result_data.get("workspace_url", ""),
            workspace_name=result_data.get("workspace_name", ""),
            scanned_at=result_data.get("scanned_at", ""),
            overall_score=result_data.get("overall_score", 0),
            total_checks=result_data.get("total_checks", 0),
            passed=result_data.get("passed", 0),
            failed=result_data.get("failed", 0),
            warnings=result_data.get("warnings", 0),
            not_applicable=result_data.get("not_applicable", 0),
            findings=findings,
            category_scores=result_data.get("category_scores", {}),
        )

        # Read file into bytes inside the temp dir before it's cleaned up
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            if fmt == "csv":
                out = export_csv(result_obj, tmp_path)
                content = Path(out).read_bytes()
                return Response(
                    content=content,
                    media_type="text/csv",
                    headers={"Content-Disposition": f'attachment; filename="assessment_{sid}.csv"'},
                )
            elif fmt == "excel":
                out = export_excel(result_obj, tmp_path)
                content = Path(out).read_bytes()
                return Response(
                    content=content,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="assessment_{sid}.xlsx"'},
                )
            else:
                raise HTTPException(status_code=400, detail=f"Unknown format: {fmt}")
    except ImportError:
        raise HTTPException(status_code=503, detail="sat_scanner not available for export")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Export failed: {exc}")


# ---------------------------------------------------------------------------
# Feature 1 — Inventory Drift Detection
# ---------------------------------------------------------------------------

@router.get("/inventory/diff")
async def inventory_diff(
    scan_a: str = Query(..., description="Baseline scan ID"),
    scan_b: str = Query(..., description="Comparison scan ID"),
):
    """Diff two UC inventory snapshots — new/removed/modified tables, schemas, catalogs."""

    def _load_inv(sid: str) -> dict:
        p = _STORE / sid / "inventory.json"
        if not p.exists():
            raise HTTPException(status_code=404, detail=f"Inventory not found for scan {sid}")
        return json.loads(p.read_text())

    inv_a = _load_inv(scan_a)
    inv_b = _load_inv(scan_b)

    def _flatten_tables(inv: dict) -> dict:
        rows: dict = {}
        for cat in inv.get("catalogs", []):
            for sch in cat.get("schemas", []):
                for tbl in sch.get("tables", []):
                    fn = tbl.get("full_name") or f"{cat['name']}.{sch['name']}.{tbl['name']}"
                    rows[fn] = {
                        "owner": (tbl.get("owner") or ""),
                        "comment": (tbl.get("comment") or ""),
                        "table_type": (tbl.get("table_type") or ""),
                        "catalog": cat["name"],
                        "schema": sch["name"],
                        "grants": len(tbl.get("grants", [])),
                        "columns": len(tbl.get("columns", [])),
                    }
        return rows

    a, b = _flatten_tables(inv_a), _flatten_tables(inv_b)
    all_keys = set(a) | set(b)

    added   = [{"full_name": k, **b[k]} for k in sorted(all_keys) if k not in a]
    removed = [{"full_name": k, **a[k]} for k in sorted(all_keys) if k not in b]
    modified = []
    for k in sorted(all_keys):
        if k in a and k in b:
            fields = ("owner", "comment", "grants", "columns", "table_type")
            changes = {
                f: {"before": a[k].get(f), "after": b[k].get(f)}
                for f in fields if a[k].get(f) != b[k].get(f)
            }
            if changes:
                modified.append({
                    "full_name": k,
                    "catalog": b[k].get("catalog", ""),
                    "schema": b[k].get("schema", ""),
                    "changes": changes,
                })

    cats_a  = {c["name"] for c in inv_a.get("catalogs", [])}
    cats_b  = {c["name"] for c in inv_b.get("catalogs", [])}
    schs_a: set[str] = {
        f"{cat['name']}.{sch['name']}"
        for cat in inv_a.get("catalogs", []) for sch in cat.get("schemas", [])
    }
    schs_b: set[str] = {
        f"{cat['name']}.{sch['name']}"
        for cat in inv_b.get("catalogs", []) for sch in cat.get("schemas", [])
    }

    return {
        "scan_a": scan_a,
        "scan_b": scan_b,
        "catalogs_added":   sorted(cats_b - cats_a),
        "catalogs_removed": sorted(cats_a - cats_b),
        "schemas_added":    sorted(schs_b - schs_a),
        "schemas_removed":  sorted(schs_a - schs_b),
        "tables_added":     added,
        "tables_removed":   removed,
        "tables_modified":  modified,
        "summary": {
            "catalogs_added":   len(cats_b - cats_a),
            "catalogs_removed": len(cats_a - cats_b),
            "schemas_added":    len(schs_b - schs_a),
            "schemas_removed":  len(schs_a - schs_b),
            "tables_added":     len(added),
            "tables_removed":   len(removed),
            "tables_modified":  len(modified),
        },
    }


# ---------------------------------------------------------------------------
# Feature 2 — Governance Policy Engine
# ---------------------------------------------------------------------------

_BUILT_IN_POLICIES: list[dict] = [
    {"id": "table_needs_owner",       "name": "Tables must have an owner",           "severity": "high"},
    {"id": "table_needs_description", "name": "Tables must have a description",       "severity": "medium"},
    {"id": "schema_needs_owner",      "name": "Schemas must have an owner",           "severity": "medium"},
    {"id": "no_catalog_all_privs",    "name": "No ALL PRIVILEGES grants on catalogs", "severity": "critical"},
    {"id": "pii_must_be_masked",      "name": "PII-named columns must have masking",  "severity": "high"},
]
_PII_PATTERNS = [
    "email", "ssn", "social_security", "phone", "mobile", "dob", "birth",
    "passport", "salary", "credit_card", "card_number", "cvv", "tax_id",
    "address", "zip", "postal", "ip_address",
]


@router.get("/policies/evaluate")
async def evaluate_policies(scan_id: str | None = Query(None)):
    """Evaluate built-in governance policies against the UC inventory."""
    sid = scan_id or ((_latest_result() or {}).get("scan_id"))
    if not sid:
        raise HTTPException(status_code=404, detail="No scan found")
    p = _STORE / sid / "inventory.json"
    if not p.exists():
        raise HTTPException(status_code=404, detail="Inventory not available — run an inventory scan first")
    inv = json.loads(p.read_text())

    violations: dict[str, list] = {pol["id"]: [] for pol in _BUILT_IN_POLICIES}

    for cat in inv.get("catalogs", []):
        for g in cat.get("grants", []):
            privs = g.get("privileges", [])
            if "ALL PRIVILEGES" in privs or "MANAGE" in privs:
                violations["no_catalog_all_privs"].append({
                    "object": cat["name"],
                    "principal": g.get("principal", ""),
                    "privileges": privs,
                })
        for sch in cat.get("schemas", []):
            fn_sch = sch.get("full_name") or f"{cat['name']}.{sch['name']}"
            if not (sch.get("owner") or "").strip():
                violations["schema_needs_owner"].append({"object": fn_sch})
            for tbl in sch.get("tables", []):
                fn = tbl.get("full_name") or f"{cat['name']}.{sch['name']}.{tbl['name']}"
                if not (tbl.get("owner") or "").strip():
                    violations["table_needs_owner"].append({"object": fn})
                if not (tbl.get("comment") or "").strip():
                    violations["table_needs_description"].append({"object": fn})
                for col in tbl.get("columns", []):
                    col_lower = (col.get("name") or "").lower()
                    if any(pat in col_lower for pat in _PII_PATTERNS) and not col.get("mask"):
                        violations["pii_must_be_masked"].append({
                            "object": fn,
                            "column": col.get("name", ""),
                        })

    # Evaluate custom policies
    custom_policies = _load_custom_policies()
    custom_results = [
        {
            **pol,
            "violations": _evaluate_custom_policy(pol, inv),
            "type": "custom",
        }
        for pol in custom_policies
    ]
    for cr in custom_results:
        cr["count"]  = len(cr["violations"])
        cr["status"] = "pass" if not cr["violations"] else "fail"

    all_policies = [
        {**pol, "violations": violations[pol["id"]], "count": len(violations[pol["id"]]),
         "status": "pass" if not violations[pol["id"]] else "fail", "type": "built_in"}
        for pol in _BUILT_IN_POLICIES
    ] + custom_results

    total_violations = sum(p["count"] for p in all_policies)

    return {
        "scan_id": sid,
        "policies": all_policies,
        "summary": {
            "total":            len(all_policies),
            "passing":          sum(1 for p in all_policies if p["status"] == "pass"),
            "failing":          sum(1 for p in all_policies if p["status"] == "fail"),
            "total_violations": total_violations,
        },
    }


# ---------------------------------------------------------------------------
# Feature 3 — Scan Scheduler endpoints
# ---------------------------------------------------------------------------

@router.get("/schedule")
async def get_schedule():
    """Return current scan schedule config."""
    if _SCHEDULE_PATH.exists():
        try:
            return json.loads(_SCHEDULE_PATH.read_text())
        except Exception:
            pass
    return {"enabled": False}


@router.put("/schedule")
async def update_schedule(body: dict):
    """Save scan schedule. body: {enabled, frequency, hour, host, token, scan_type, workspace_name}"""
    _SCHEDULE_PATH.write_text(json.dumps(body, indent=2))
    return body


@router.delete("/schedule")
async def delete_schedule():
    """Delete scan schedule."""
    if _SCHEDULE_PATH.exists():
        _SCHEDULE_PATH.unlink()
    return {"enabled": False}


# ---------------------------------------------------------------------------
# Feature 4 — Column Lineage proxy
# ---------------------------------------------------------------------------

@router.get("/lineage/table")
async def table_lineage(
    table_name: str = Query(..., description="Full table name: catalog.schema.table"),
    x_databricks_host: str | None = Header(None),
    x_databricks_token: str | None = Header(None),
):
    """Proxy Databricks lineage-tracking/table-lineage for a given table."""
    import httpx
    host  = (x_databricks_host  or "").rstrip("/")
    token = (x_databricks_token or "")
    if not host or not token:
        raise HTTPException(
            status_code=401,
            detail="Databricks credentials required (X-Databricks-Host, X-Databricks-Token headers)",
        )
    try:
        async with httpx.AsyncClient(timeout=15) as c:
            r = await c.get(
                f"{host}/api/2.0/lineage-tracking/table-lineage",
                headers={"Authorization": f"Bearer {token}"},
                params={"table_name": table_name, "include_entity_lineage": "true"},
            )
        if r.status_code == 404:
            return {"upstream_tables": [], "downstream_tables": [], "table_name": table_name}
        if r.status_code != 200:
            raise HTTPException(status_code=r.status_code, detail=r.text[:500])
        data = r.json()
        data["table_name"] = table_name
        return data
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"Databricks lineage API unreachable: {exc}")


# ---------------------------------------------------------------------------
# AI-powered Remediation Plans (Feature — Databricks Foundation Model)
# ---------------------------------------------------------------------------

@router.post("/ai/remediation-plan")
async def ai_remediation_plan(
    body: dict,
    x_databricks_host: str | None = Header(None),
    x_databricks_token: str | None = Header(None),
    model: str = Query("databricks-meta-llama-3-1-70b-instruct"),
):
    """Generate a step-by-step remediation plan for a security finding via Databricks Foundation Model."""
    import httpx
    host  = (x_databricks_host  or "").rstrip("/")
    token = (x_databricks_token or "")
    if not host or not token:
        raise HTTPException(status_code=401, detail="Databricks credentials required")

    finding = body.get("finding", {})
    system_prompt = (
        "You are a Databricks security expert specialising in Unity Catalog governance and workspace security. "
        "Generate a concise, actionable step-by-step remediation plan for the given security finding. "
        "Use markdown. Include specific Databricks SQL or Python snippets where helpful. "
        "Keep the plan focused — no unnecessary padding."
    )
    user_prompt = f"""## Security Finding

**Check ID:** {finding.get('check_id', '')}
**Title:** {finding.get('title', '')}
**Category:** {finding.get('category', '')}
**Severity:** {finding.get('severity', '')}
**Description:** {finding.get('description', '')}
**Current Recommendation:** {finding.get('recommendation', '')}
**Effort:** {finding.get('effort', '')}

Please provide:
1. **Root Cause** (1-2 sentences on why this happens)
2. **Step-by-Step Fix** (numbered steps with Databricks SQL/Python snippets where applicable)
3. **Validation** (how to confirm the fix worked)
4. **Prevention** (how to avoid recurrence)
5. **Quick Win** (if there's a sub-5-minute partial fix, highlight it)"""

    try:
        async with httpx.AsyncClient(timeout=45) as c:
            r = await c.post(
                f"{host}/serving-endpoints/{model}/invocations",
                headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
                json={
                    "messages": [
                        {"role": "system", "content": system_prompt},
                        {"role": "user",   "content": user_prompt},
                    ],
                    "max_tokens": 1500,
                    "temperature": 0.1,
                },
            )
        if r.status_code != 200:
            raise HTTPException(status_code=r.status_code, detail=f"AI model error: {r.text[:500]}")
        content = r.json()["choices"][0]["message"]["content"]
        return {"plan": content, "model": model, "check_id": finding.get("check_id", "")}
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"AI service unreachable: {exc}")


# ---------------------------------------------------------------------------
# Custom Policy CRUD
# ---------------------------------------------------------------------------

_CUSTOM_POLICIES_PATH = _STORE.parent / "custom_policies.json"

# Supported custom rule types and their evaluator keys
_CUSTOM_RULE_TYPES = [
    "tables_need_owner",
    "tables_need_description",
    "schemas_need_owner",
    "no_all_privs_on_catalog",
    "pii_columns_must_be_masked",
]


def _load_custom_policies() -> list[dict]:
    if _CUSTOM_POLICIES_PATH.exists():
        try:
            return json.loads(_CUSTOM_POLICIES_PATH.read_text())
        except Exception:
            pass
    return []


def _evaluate_custom_policy(policy: dict, inv: dict) -> list[dict]:
    """Run a single custom policy rule against an inventory snapshot."""
    rt      = policy.get("rule_type", "")
    scope   = (policy.get("catalog_scope") or "").strip()   # empty = all catalogs
    pat     = (policy.get("column_pattern") or "").lower()  # for PII check

    violations: list[dict] = []

    for cat in inv.get("catalogs", []):
        if scope and cat["name"] != scope:
            continue

        if rt == "no_all_privs_on_catalog":
            for g in cat.get("grants", []):
                if "ALL PRIVILEGES" in g.get("privileges", []) or "MANAGE" in g.get("privileges", []):
                    violations.append({"object": cat["name"], "principal": g.get("principal", "")})
            continue

        for sch in cat.get("schemas", []):
            fn_sch = sch.get("full_name") or f"{cat['name']}.{sch['name']}"

            if rt == "schemas_need_owner" and not (sch.get("owner") or "").strip():
                violations.append({"object": fn_sch})

            for tbl in sch.get("tables", []):
                fn = tbl.get("full_name") or f"{cat['name']}.{sch['name']}.{tbl['name']}"
                if rt == "tables_need_owner" and not (tbl.get("owner") or "").strip():
                    violations.append({"object": fn})
                elif rt == "tables_need_description" and not (tbl.get("comment") or "").strip():
                    violations.append({"object": fn})
                elif rt == "pii_columns_must_be_masked":
                    for col in tbl.get("columns", []):
                        col_lower = (col.get("name") or "").lower()
                        if (not pat or pat in col_lower) and any(p in col_lower for p in _PII_PATTERNS) and not col.get("mask"):
                            violations.append({"object": fn, "column": col.get("name", "")})

    return violations


@router.get("/policies")
async def list_policies():
    """Return built-in and custom policy definitions."""
    return {"built_in": _BUILT_IN_POLICIES, "custom": _load_custom_policies()}


@router.post("/policies")
async def create_policy(body: dict):
    """Create a custom governance policy."""
    import time
    policies = _load_custom_policies()
    new_pol = {
        "id":             f"custom_{int(time.time() * 1000)}",
        "name":           (body.get("name") or "Custom Policy").strip(),
        "severity":       body.get("severity", "medium"),
        "rule_type":      body.get("rule_type", "tables_need_owner"),
        "catalog_scope":  body.get("catalog_scope", ""),
        "column_pattern": body.get("column_pattern", ""),
        "type":           "custom",
    }
    policies.append(new_pol)
    _CUSTOM_POLICIES_PATH.write_text(json.dumps(policies, indent=2))
    return new_pol


@router.delete("/policies/{policy_id}")
async def delete_policy(policy_id: str):
    """Delete a custom policy by ID."""
    policies = [p for p in _load_custom_policies() if p.get("id") != policy_id]
    _CUSTOM_POLICIES_PATH.write_text(json.dumps(policies, indent=2))
    return {"deleted": policy_id}


# ---------------------------------------------------------------------------
# Inventory Timeline
# ---------------------------------------------------------------------------

@router.get("/inventory/timeline")
async def inventory_timeline():
    """Return catalog/schema/table/column counts across all scans over time (oldest-first)."""
    results = list(reversed(_list_results()))  # oldest first
    timeline = []
    for meta in results:
        sid = meta.get("scan_id")
        if not sid:
            continue
        entry: dict = {
            "scan_id":        sid,
            "scanned_at":     meta.get("scanned_at", ""),
            "workspace_name": meta.get("workspace_name") or meta.get("workspace_url", ""),
            "catalogs":       meta.get("catalog_count", 0) or 0,
            "schemas":        meta.get("schema_count",  0) or 0,
            "tables":         meta.get("table_count",   0) or 0,
            "columns":        0,
        }
        # Enrich from inventory.json when available
        inv_path = _STORE / sid / "inventory.json"
        if inv_path.exists():
            try:
                inv = json.loads(inv_path.read_text())
                cats = inv.get("catalogs", [])
                entry["catalogs"] = len(cats)
                entry["schemas"]  = sum(len(c.get("schemas", [])) for c in cats)
                entry["tables"]   = sum(
                    len(s.get("tables", []))
                    for c in cats for s in c.get("schemas", [])
                )
                entry["columns"]  = sum(
                    len(t.get("columns", []))
                    for c in cats for s in c.get("schemas", []) for t in s.get("tables", [])
                )
            except Exception:
                pass
        timeline.append(entry)
    return timeline
