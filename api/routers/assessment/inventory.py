"""UC inventory export, diff, and timeline endpoints."""

from __future__ import annotations

import json
import tempfile
import zipfile
from pathlib import Path

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import FileResponse, Response

from ._storage import _STORE, _html_path, _latest_result, _list_results

router = APIRouter()


@router.get("/inventory/export")
async def export_inventory(
    fmt: str = Query("json"),
    scan_id: str | None = Query(None),
):
    """Download UC inventory: json | csv_tables | csv_columns | excel | html."""
    sid = scan_id or ((_latest_result() or {}).get("scan_id"))
    if not sid:
        raise HTTPException(status_code=404, detail="No scan found")

    inv_path = _STORE / sid / "inventory.json"
    if not inv_path.exists():
        raise HTTPException(
            status_code=404,
            detail="Inventory not available for this scan — run a scan with UC Inventory enabled",
        )

    inv = json.loads(inv_path.read_text())

    if fmt == "json":
        return FileResponse(
            str(inv_path),
            media_type="application/json",
            filename=f"uc_inventory_{sid}.json",
        )

    if fmt == "csv_tables":
        import csv
        import io as _io
        buf = _io.StringIO()
        w   = csv.writer(buf)
        w.writerow([
            "catalog", "catalog_type", "schema", "table", "full_name",
            "table_type", "data_format", "owner", "comment",
            "storage_location", "created_at", "updated_at", "column_count", "grant_count",
        ])
        for cat in inv.get("catalogs", []):
            for sch in cat.get("schemas", []):
                for tbl in sch.get("tables", []):
                    w.writerow([
                        cat["name"], cat.get("catalog_type", ""), sch["name"],
                        tbl["name"], tbl.get("full_name", ""), tbl.get("table_type", ""),
                        tbl.get("data_source_format", ""), tbl.get("owner", ""),
                        tbl.get("comment", ""), tbl.get("storage_location", ""),
                        tbl.get("created_at", ""), tbl.get("updated_at", ""),
                        len(tbl.get("columns", [])), len(tbl.get("grants", [])),
                    ])
        return Response(
            content=buf.getvalue().encode(),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="uc_tables_{sid}.csv"'},
        )

    if fmt == "csv_columns":
        import csv
        import io as _io
        buf = _io.StringIO()
        w   = csv.writer(buf)
        w.writerow([
            "catalog", "schema", "table", "full_table_name",
            "column", "position", "type", "nullable", "comment", "masked",
        ])
        for cat in inv.get("catalogs", []):
            for sch in cat.get("schemas", []):
                for tbl in sch.get("tables", []):
                    for col in tbl.get("columns", []):
                        w.writerow([
                            cat["name"], sch["name"], tbl["name"],
                            tbl.get("full_name", ""), col["name"],
                            col.get("position", ""), col.get("type_text", ""),
                            col.get("nullable", ""), col.get("comment", ""),
                            "yes" if col.get("mask") else "no",
                        ])
        return Response(
            content=buf.getvalue().encode(),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="uc_columns_{sid}.csv"'},
        )

    if fmt == "excel":
        try:
            import io as _io
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise HTTPException(status_code=503, detail="openpyxl not installed — cannot generate Excel export")

        HEADER_FILL = PatternFill("solid", fgColor="E8453C")
        HEADER_FONT = Font(color="FFFFFF", bold=True)

        def _add_sheet(wb, name, headers, rows):
            ws = wb.create_sheet(name)
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            for row in rows:
                ws.append(row)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        _add_sheet(wb, "Catalogs",
            ["Name", "Type", "Owner", "Comment", "Storage Root", "Isolation Mode"],
            [[c["name"], c.get("catalog_type", ""), c.get("owner", ""), c.get("comment", ""),
              c.get("storage_root", ""), c.get("isolation_mode", "")]
             for c in inv.get("catalogs", [])])

        _add_sheet(wb, "Schemas",
            ["Full Name", "Catalog", "Schema", "Owner", "Comment"],
            [[s.get("full_name", ""), cat["name"], s["name"], s.get("owner", ""), s.get("comment", "")]
             for cat in inv.get("catalogs", []) for s in cat.get("schemas", [])])

        _add_sheet(wb, "Tables",
            ["Full Name", "Catalog", "Schema", "Table", "Type", "Format", "Owner", "Comment",
             "Storage Location", "Columns", "Grants", "Created", "Updated"],
            [[t.get("full_name", ""), cat["name"], sch["name"], t["name"],
              t.get("table_type", ""), t.get("data_source_format", ""),
              t.get("owner", ""), t.get("comment", ""), t.get("storage_location", ""),
              len(t.get("columns", [])), len(t.get("grants", [])),
              t.get("created_at", ""), t.get("updated_at", "")]
             for cat in inv.get("catalogs", []) for sch in cat.get("schemas", [])
             for t in sch.get("tables", [])])

        _add_sheet(wb, "Columns",
            ["Table", "Catalog", "Schema", "Column", "Position", "Type", "Nullable", "Comment", "Masked"],
            [[t.get("full_name", ""), cat["name"], sch["name"], col["name"],
              col.get("position", ""), col.get("type_text", ""), col.get("nullable", ""),
              col.get("comment", ""), "yes" if col.get("mask") else "no"]
             for cat in inv.get("catalogs", []) for sch in cat.get("schemas", [])
             for t in sch.get("tables", []) for col in t.get("columns", [])])

        grant_rows: list[list] = []
        for g in inv.get("metastore_grants", []):
            grant_rows.append(["METASTORE", "—", g.get("full_name", ""), g.get("principal", ""),
                                ", ".join(g.get("privileges", [])), g.get("inherited_from", "")])
        for cat in inv.get("catalogs", []):
            for g in cat.get("grants", []):
                grant_rows.append(["CATALOG", cat["name"], g.get("full_name", ""), g.get("principal", ""),
                                    ", ".join(g.get("privileges", [])), g.get("inherited_from", "")])
            for sch in cat.get("schemas", []):
                for g in sch.get("grants", []):
                    grant_rows.append(["SCHEMA", cat["name"], g.get("full_name", ""), g.get("principal", ""),
                                        ", ".join(g.get("privileges", [])), g.get("inherited_from", "")])
                for t in sch.get("tables", []):
                    for g in t.get("grants", []):
                        grant_rows.append(["TABLE", cat["name"], g.get("full_name", ""), g.get("principal", ""),
                                            ", ".join(g.get("privileges", [])), g.get("inherited_from", "")])
        _add_sheet(wb, "Grants",
            ["Level", "Catalog", "Object", "Principal", "Privileges", "Inherited From"],
            grant_rows)

        _add_sheet(wb, "External Locations",
            ["Name", "URL", "Credential", "Read Only", "Owner", "Comment"],
            [[e.get("name", ""), e.get("url", ""), e.get("credential_name", ""),
              str(e.get("read_only", "")), e.get("owner", ""), e.get("comment", "")]
             for e in inv.get("external_locations", [])])

        buf = __import__("io").BytesIO()
        wb.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="uc_inventory_{sid}.xlsx"'},
        )

    if fmt == "html":
        html_files: list[Path] = []
        for view in ("overview", "tree", "sunburst", "hubspoke", "topology"):
            p = _html_path(sid, view)
            if p and p.exists():
                html_files.append(p)
        if not html_files:
            raise HTTPException(
                status_code=404,
                detail="No inventory HTML found — run a scan with UC Inventory enabled",
            )
        if len(html_files) == 1:
            return FileResponse(str(html_files[0]), media_type="text/html",
                                filename=f"uc_inventory_{sid}.html")
        tmp = tempfile.NamedTemporaryFile(suffix=".zip", delete=False)
        tmp.close()
        with zipfile.ZipFile(tmp.name, "w", zipfile.ZIP_DEFLATED) as zf:
            for p in html_files:
                zf.write(str(p), arcname=p.name)
        return FileResponse(
            tmp.name,
            media_type="application/zip",
            filename=f"uc_inventory_{sid}.zip",
        )

    raise HTTPException(status_code=400, detail=f"Unknown format: {fmt}")


@router.get("/inventory/diff")
async def inventory_diff(
    scan_a: str = Query(..., description="Baseline scan ID"),
    scan_b: str = Query(..., description="Comparison scan ID"),
):
    """Diff two UC inventory snapshots — new/removed/modified tables, schemas, catalogs."""

    def _load_inv(sid: str) -> dict:
        p = _STORE / sid / "inventory.json"
        if not p.exists():
            raise HTTPException(status_code=404, detail=f"Inventory not found for scan {sid}")
        return json.loads(p.read_text())

    inv_a = _load_inv(scan_a)
    inv_b = _load_inv(scan_b)

    def _flatten_tables(inv: dict) -> dict:
        rows: dict = {}
        for cat in inv.get("catalogs", []):
            for sch in cat.get("schemas", []):
                for tbl in sch.get("tables", []):
                    fn = tbl.get("full_name") or f"{cat['name']}.{sch['name']}.{tbl['name']}"
                    rows[fn] = {
                        "owner":      (tbl.get("owner")      or ""),
                        "comment":    (tbl.get("comment")    or ""),
                        "table_type": (tbl.get("table_type") or ""),
                        "catalog":    cat["name"],
                        "schema":     sch["name"],
                        "grants":     len(tbl.get("grants",  [])),
                        "columns":    len(tbl.get("columns", [])),
                    }
        return rows

    a, b    = _flatten_tables(inv_a), _flatten_tables(inv_b)
    all_keys = set(a) | set(b)

    added    = [{"full_name": k, **b[k]} for k in sorted(all_keys) if k not in a]
    removed  = [{"full_name": k, **a[k]} for k in sorted(all_keys) if k not in b]
    modified = []
    for k in sorted(all_keys):
        if k in a and k in b:
            fields  = ("owner", "comment", "grants", "columns", "table_type")
            changes = {
                f: {"before": a[k].get(f), "after": b[k].get(f)}
                for f in fields if a[k].get(f) != b[k].get(f)
            }
            if changes:
                modified.append({
                    "full_name": k,
                    "catalog":   b[k].get("catalog", ""),
                    "schema":    b[k].get("schema", ""),
                    "changes":   changes,
                })

    cats_a = {c["name"] for c in inv_a.get("catalogs", [])}
    cats_b = {c["name"] for c in inv_b.get("catalogs", [])}
    schs_a: set[str] = {
        f"{cat['name']}.{sch['name']}"
        for cat in inv_a.get("catalogs", []) for sch in cat.get("schemas", [])
    }
    schs_b: set[str] = {
        f"{cat['name']}.{sch['name']}"
        for cat in inv_b.get("catalogs", []) for sch in cat.get("schemas", [])
    }

    return {
        "scan_a":           scan_a,
        "scan_b":           scan_b,
        "catalogs_added":   sorted(cats_b - cats_a),
        "catalogs_removed": sorted(cats_a - cats_b),
        "schemas_added":    sorted(schs_b - schs_a),
        "schemas_removed":  sorted(schs_a - schs_b),
        "tables_added":     added,
        "tables_removed":   removed,
        "tables_modified":  modified,
        "summary": {
            "catalogs_added":   len(cats_b - cats_a),
            "catalogs_removed": len(cats_a - cats_b),
            "schemas_added":    len(schs_b - schs_a),
            "schemas_removed":  len(schs_a - schs_b),
            "tables_added":     len(added),
            "tables_removed":   len(removed),
            "tables_modified":  len(modified),
        },
    }


@router.get("/inventory/timeline")
async def inventory_timeline():
    """Return catalog/schema/table/column counts across all scans over time (oldest-first)."""
    results  = list(reversed(_list_results()))
    timeline = []
    for meta in results:
        sid = meta.get("scan_id")
        if not sid:
            continue
        entry: dict = {
            "scan_id":        sid,
            "scanned_at":     meta.get("scanned_at", ""),
            "workspace_name": meta.get("workspace_name") or meta.get("workspace_url", ""),
            "catalogs":       meta.get("catalog_count", 0) or 0,
            "schemas":        meta.get("schema_count",  0) or 0,
            "tables":         meta.get("table_count",   0) or 0,
            "columns":        0,
        }
        inv_path = _STORE / sid / "inventory.json"
        if inv_path.exists():
            try:
                inv              = json.loads(inv_path.read_text())
                cats             = inv.get("catalogs", [])
                entry["catalogs"] = len(cats)
                entry["schemas"]  = sum(len(c.get("schemas", [])) for c in cats)
                entry["tables"]   = sum(
                    len(s.get("tables", []))
                    for c in cats for s in c.get("schemas", [])
                )
                entry["columns"]  = sum(
                    len(t.get("columns", []))
                    for c in cats for s in c.get("schemas", []) for t in s.get("tables", [])
                )
            except Exception:
                pass
        timeline.append(entry)
    return timeline
