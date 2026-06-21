"""SAT Scanner — export functions (JSON, CSV, Excel, HTML)."""

from __future__ import annotations

import csv
import io
import json
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from .models import SATFinding, SATScanResult
from .checks import SAT_CHECKS, _get_effort, CHECK_BENEFITS, CHECK_API_ENDPOINTS, EXCEL_CELL_LIMIT, CATEGORY_DEFINITIONS
from .scoring import _build_prioritised_recommendations
from .helpers import (
    _pl, _sanitize_name, _file_prefix, _details_str, _format_scan_items,
    _render_secret_details_html, _render_scan_items_html,
)

def export_json(result: SATScanResult, output_dir: Path) -> str:
    path = output_dir / f"{_file_prefix(result)}.json"
    path.write_text(json.dumps(result.to_dict(), indent=2, default=str))
    return str(path)


def export_api_dump(result: SATScanResult, output_dir: Path) -> str:
    """Export all raw API responses to a standalone JSON file for verification."""
    dump: dict = {
        "workspace_url": result.workspace_url,
        "workspace_name": result.workspace_name,
        "scanned_at": result.scanned_at,
        "api_responses": {},
    }
    for f in result.findings:
        api_resp = f.details.get("api_response")
        if api_resp is not None:
            dump["api_responses"][f.check_id] = {
                "check_id": f.check_id,
                "check_title": f.title,
                "status": f.status,
                "api_endpoint": f.details.get("api_endpoint", ""),
                "response": api_resp,
            }
    path = output_dir / f"{_file_prefix(result)}-api-dump.json"
    path.write_text(json.dumps(dump, indent=2, default=str))
    return str(path)



def export_csv(result: SATScanResult, output_dir: Path, include_api_response: bool = True, show_scan_items: bool = False, show_evidence: bool = False, show_effort: bool = False, show_cost: bool = False) -> str:
    path = output_dir / f"{_file_prefix(result)}.csv"
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        headers = ["Check ID", "Category", "Title", "Status"]
        if show_effort:
            headers.append("Effort")
        headers.extend(["Severity", "Current State", "Recommendation", "Reference URL", "Is API Error"])
        if show_evidence:
            headers.extend(["Evidence Field", "Evidence Value"])
        headers.append("Portal Link")
        headers.append("Why It Matters")
        if include_api_response:
            headers.append("API Response")
        if show_scan_items:
            headers.append("Scanned Items")
        writer.writerow(headers)
        for finding in result.findings:
            row = [finding.check_id, finding.category, finding.title, finding.status]
            if show_effort:
                row.append(finding.effort or "")
            row.extend([finding.severity.upper(), finding.current_state, finding.recommendation, finding.reference_url,
                "Yes" if finding.is_api_error else ""])
            if show_evidence:
                ev = finding.evidence or {}
                row.append(ev.get("field", ""))
                val = ev.get("value", "")
                row.append(json.dumps(val, default=str) if not isinstance(val, str) else val)
            row.append(finding.portal_link or "")
            row.append(finding.benefits or "")
            if include_api_response:
                row.append(_details_str(finding.details))
            if show_scan_items:
                row.append(_format_scan_items(finding.details))
            writer.writerow(row)
    # ── Endpoint Summary → separate CSV (matches Excel separate-sheet approach) ──
    if result.endpoint_summary and result.endpoint_summary.get("endpoints"):
        ep = result.endpoint_summary
        ep_path = output_dir / f"{_file_prefix(result)}_api_endpoints.csv"
        with open(ep_path, "w", newline="", encoding="utf-8-sig") as ef:
            ep_writer = csv.writer(ef)
            ep_writer.writerow(["API Endpoint Summary"])
            ep_writer.writerow(["Total Endpoints", ep["total"]])
            ep_writer.writerow(["With Items", ep["with_items"]])
            ep_writer.writerow(["Config/Settings", ep["config"]])
            ep_writer.writerow(["Empty", ep["empty"]])
            ep_writer.writerow(["Errors", ep.get("error", 0)])
            ep_writer.writerow([])
            ep_writer.writerow(["Endpoint", "Status", "Items Count", "Error Code"])
            for e in ep["endpoints"]:
                ep_writer.writerow([e["endpoint"], e["status"], e["items_count"], e.get("error_code", "") or ""])
    # ── Effort methodology → separate CSV ──
    if show_effort:
        _actionable = [f for f in result.findings if f.status in ("FAIL", "WARN") and not f.is_api_error]
        eff_path = output_dir / f"{_file_prefix(result)}_effort.csv"
        with open(eff_path, "w", newline="", encoding="utf-8-sig") as ef:
            ew = csv.writer(ef)
            ew.writerow(["Remediation Effort Summary"])
            ew.writerow(["Actionable Findings (FAIL + WARN)", len(_actionable)])
            ew.writerow([])
            ew.writerow(["Effort Level", "Count", "Percentage"])
            _eff_counts: dict[str, int] = {}
            for f in _actionable:
                e = f.effort or "Moderate (1\u20134 hrs)"
                _eff_counts[e] = _eff_counts.get(e, 0) + 1
            for eff in ["Quick Fix (5\u201315 min)", "Moderate (1\u20134 hrs)", "Significant (1\u20133 days)", "Project (1+ weeks)"]:
                cnt = _eff_counts.get(eff, 0)
                pct = f"{round(cnt / len(_actionable) * 100)}%" if _actionable else "0%"
                ew.writerow([eff, cnt, pct])
            ew.writerow([])
            ew.writerow(["=== EFFORT LEVEL DEFINITIONS ==="])
            ew.writerow(["Level", "Time Range", "What's Included"])
            ew.writerow(["Quick Fix", "5\u201315 min", "Single configuration toggle or setting change. No cross-team coordination or testing required."])
            ew.writerow(["Moderate", "1\u20134 hrs", "Multi-step configuration, policy creation, or changes that require testing/validation. May involve IaC code updates or build/deploy pipelines."])
            ew.writerow(["Significant", "1\u20133 days", "Architecture changes requiring cross-team coordination, downtime planning, IaC refactoring, and phased rollout."])
            ew.writerow(["Project", "1+ weeks", "Major infrastructure migration or org-wide policy rollout. Requires project planning, stakeholder approval, IaC redesign."])
            ew.writerow([])
            ew.writerow(["=== COMMON PREREQUISITES (not included in estimates) ==="])
            ew.writerow(["Prerequisite", "When Needed", "Typical Additional Time"])
            ew.writerow(["Terraform / IaC Updates", "When infrastructure is managed via Terraform, Pulumi, ARM/Bicep, or other IaC tools.", "+30 min \u2013 2 hrs per change"])
            ew.writerow(["CI/CD Pipeline Runs", "When changes must flow through build, test, and deploy pipelines before reaching production.", "+15 min \u2013 1 hr per deployment"])
            ew.writerow(["Change Management / CAB Approval", "When your organization requires change tickets, CAB review, or ITSM approvals.", "+1 \u2013 5 business days"])
            ew.writerow(["Non-prod Testing", "When changes must be validated in dev/staging environments before production rollout.", "+1 \u2013 4 hrs per environment"])
            ew.writerow(["Provider / Module Upgrades", "When the Terraform azurerm/databricks provider or shared modules must be upgraded.", "+1 \u2013 4 hrs"])
            ew.writerow(["Security Review", "When network, identity, or security teams must review and approve the change.", "+1 \u2013 3 business days"])
            ew.writerow(["Documentation Updates", "When internal runbooks, architecture diagrams, or compliance docs must be updated.", "+30 min \u2013 2 hrs"])
            ew.writerow([])
            ew.writerow(["Note:", "Effort estimates reflect core remediation work only. Add prerequisite time as applicable to your organization."])
    # ── Prioritised Recommendations → separate CSV ──
    prio_items = _build_prioritised_recommendations(result.findings)
    if prio_items:
        prio_path = output_dir / f"{_file_prefix(result)}_prioritised.csv"
        with open(prio_path, "w", newline="", encoding="utf-8-sig") as pf:
            pw = csv.writer(pf)
            pw.writerow(["Prioritised Recommendations"])
            pw.writerow(["Actionable Findings (FAIL + WARN)", len(prio_items)])
            pw.writerow(["Sorted by: Priority Score (severity × effort multiplier, highest first)"])
            if show_cost:
                pw.writerow(["Note: Cost figures are illustrative examples only — actual costs vary with usage, region, and pricing tier."])
            pw.writerow([])
            _csv_prio_hdrs = ["Priority", "Score", "Check ID", "Category", "Severity", "Status", "Effort"]
            if show_cost:
                _csv_prio_hdrs += ["Est. Cost ($/mo Low)", "Est. Cost ($/mo High)", "Cost Reason"]
            _csv_prio_hdrs += ["Title", "Recommendation", "Why It Matters", "Portal Link"]
            pw.writerow(_csv_prio_hdrs)
            for item in prio_items:
                _csv_row = [item["priority_label"], item["priority_score"], item["check_id"],
                    item["category"], item["severity"].upper(), item["status"], item["effort"]]
                if show_cost:
                    _csv_row += [item["cost_low"] or "", item["cost_high"] or "", item["cost_reason"]]
                _csv_row += [item["title"], item["recommendation"], item["benefits"], item["portal_link"]]
                pw.writerow(_csv_row)
    return str(path)


def export_excel(result: SATScanResult, output_dir: Path, include_api_response: bool = True, show_scan_items: bool = False, show_evidence: bool = False, show_effort: bool = False, show_cost: bool = False) -> str:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("ERROR: 'openpyxl' is required for Excel export.  Install with:  pip install sat-scanner[excel]")
        sys.exit(1)

    wb = Workbook()

    # ── Styles ──
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    pass_fill = PatternFill("solid", fgColor="C6EFCE")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")
    warn_fill = PatternFill("solid", fgColor="FFEB9C")
    na_fill = PatternFill("solid", fgColor="D9D9D9")

    def _status_fill(status: str):
        return {"PASS": pass_fill, "FAIL": fail_fill, "WARN": warn_fill}.get(status, na_fill)

    def _set_header(ws, row, cols):
        for c, val in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border

    # ── Sheet 1: Summary ──
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.cell(row=1, column=1, value="Databricks SAT — Summary").font = title_font
    ws_sum.merge_cells("A1:C1")
    _set_header(ws_sum, 3, ["Field", "Value", "Definition"])
    summary_rows = [
        ("Workspace URL", result.workspace_url, "The Databricks workspace endpoint that was scanned."),
        ("Workspace Name", result.workspace_name or "N/A", "Display name of the workspace."),
        ("Scanned At", datetime.fromisoformat(result.scanned_at.replace("Z", "+00:00")).strftime("%m/%d/%Y, %I:%M:%S %p"), "Date and time when this security scan was executed."),
        ("Databricks Version", result.databricks_version or "N/A", "Runtime version reported by the workspace."),
    ]
    for r, (field, val, defn) in enumerate(summary_rows, 4):
        ws_sum.cell(row=r, column=1, value=field).border = thin_border
        ws_sum.cell(row=r, column=2, value=val).border = thin_border
        ws_sum.cell(row=r, column=3, value=defn).border = thin_border

    _set_header(ws_sum, 9, ["Metric", "Value", "Definition"])
    metric_rows = [
        ("Overall Score", result.overall_score, "Weighted security score (0\u2013100). Formula: Score = (1 \u2212 penalty_sum / total_weights) \u00d7 100, where penalty = FAIL \u00d7 full_weight + WARN \u00d7 half_weight. Weights: Critical=10, High=7, Medium=4, Low=2. WARN penalties count at half weight (0.5\u00d7)."),
        ("Total Checks", result.total_checks, "Total number of SAT security checks executed against this workspace."),
        (_pl(result.passed, "Check Passed", "Checks Passed"), result.passed, "Checks where the security control is confirmed to be correctly configured. No action needed."),
        (_pl(result.failed, "Check Failed", "Checks Failed"), result.failed, "Checks that found a concrete security gap. These require remediation action."),
        (_pl(result.warnings, "Warning"), result.warnings, "Checks that returned a borderline result (genuine warnings only, API errors excluded). Requires manual investigation."),
        (_pl(result.not_applicable, "Not Applicable", "Not Applicable"), result.not_applicable, "Checks skipped because the feature is not in use on this workspace. Excluded from scoring."),
        (_pl(result.api_errors, "API Error"), result.api_errors, "Checks that could not be evaluated due to API failures (HTTP 400/401/403/404, timeouts, connection errors). Excluded from scoring. See 'API Errors' sheet for details."),
    ]
    for r, (metric, val, defn) in enumerate(metric_rows, 10):
        ws_sum.cell(row=r, column=1, value=metric).border = thin_border
        ws_sum.cell(row=r, column=2, value=val).border = thin_border
        ws_sum.cell(row=r, column=3, value=defn).border = thin_border

    # Score breakdown rows
    _XSW = {"critical": 10, "high": 7, "medium": 4, "low": 2}
    _x_scorable = [f for f in result.findings if not f.is_api_error]
    _x_applicable = [f for f in _x_scorable if f.status != "NOT_APPLICABLE"]
    _x_fails = [f for f in _x_applicable if f.status == "FAIL"]
    _x_warns = [f for f in _x_applicable if f.status == "WARN"]
    _x_passes = [f for f in _x_applicable if f.status == "PASS"]
    _x_tot = sum(_XSW.get(SAT_CHECKS.get(f.check_id, {}).get("severity", "low"), 2) for f in _x_applicable)
    _x_fp = sum(_XSW.get(SAT_CHECKS.get(f.check_id, {}).get("severity", "low"), 2) for f in _x_fails)
    _x_wp = sum(_XSW.get(SAT_CHECKS.get(f.check_id, {}).get("severity", "low"), 2) * 0.5 for f in _x_warns)
    _x_tp = _x_fp + _x_wp
    _bk_start = 10 + len(metric_rows) + 1
    _set_header(ws_sum, _bk_start, ["Score Breakdown", "Value", "Detail"])
    breakdown_rows = [
        ("Scored Checks", len(_x_applicable), f"Excluding {result.not_applicable} N/A + {result.api_errors} API Error"),
        ("Total Weight Pool", f"{_x_tot:.0f} pts", "Sum of severity weights for all scored checks"),
        ("FAIL Penalty", f"\u2212{_x_fp:.0f} pts", f"{len(_x_fails)} checks \u00d7 full weight"),
        ("WARN Penalty", f"\u2212{_x_wp:.1f} pts", f"{len(_x_warns)} checks \u00d7 half weight (0.5\u00d7)"),
        ("PASS", "0 pts", f"{len(_x_passes)} checks \u2014 no penalty"),
        ("Formula", f"(1 \u2212 {_x_tp:.1f} / {_x_tot:.0f}) \u00d7 100 = {result.overall_score}", ""),
        ("Grade", "Good" if result.overall_score >= 80 else ("Needs Improvement" if result.overall_score >= 60 else "Critical"),
         "Good = 80\u2013100, Needs Improvement = 60\u201379, Critical = 0\u201359"),
    ]
    for r, (label, val, detail) in enumerate(breakdown_rows, _bk_start + 1):
        ws_sum.cell(row=r, column=1, value=label).border = thin_border
        ws_sum.cell(row=r, column=2, value=val).border = thin_border
        ws_sum.cell(row=r, column=3, value=detail).border = thin_border

    ws_sum.column_dimensions["A"].width = 25
    ws_sum.column_dimensions["B"].width = 50
    ws_sum.column_dimensions["C"].width = 80

    # ── Sheet 2: Definitions ──
    ws_def = wb.create_sheet("Definitions")
    def_rows = [
        ("Databricks SAT \u2014 Definitions & Reference", "", ""),
        ("", "", ""),
        ("=== FINDING STATUSES ===", "", ""),
        ("Status", "Label", "Definition"),
        ("PASS", "Compliant", "The check ran successfully and confirmed the security control is in place. No action needed."),
        ("FAIL", "Action Required", "The check found a concrete security gap \u2014 confirmed bad configuration that must be fixed. Examples: no audit logging, plaintext credentials in cluster config, public cluster access enabled."),
        ("WARN", "Review Needed", "The check returned a borderline result that needs manual investigation. Genuine warnings only (API errors are listed separately)."),
        ("NOT_APPLICABLE", "Skipped", "The feature being checked is not in use on this workspace (e.g. no Delta Sharing recipients, no MLflow models, Unity Catalog not enabled). Excluded from the score."),
        ("API ERROR", "Could Not Evaluate", "The check could not run due to an API failure (HTTP 400/401/403/404, timeout, or connection error). Excluded from score. Fix the underlying issue and re-scan. See the 'API Errors' sheet."),
        ("", "", ""),
        ("=== SEVERITY LEVELS ===", "", ""),
        ("Severity", "Weight", "Definition"),
        ("CRITICAL", "10", "Immediate security risk that could lead to data breach, unauthorized access, or compliance violation. Must be remediated immediately."),
        ("HIGH", "7", "Significant security weakness that could be exploited. Should be remediated within days."),
        ("MEDIUM", "4", "Moderate risk that weakens overall security posture. Should be remediated within the current sprint/cycle."),
        ("LOW", "2", "Minor improvement opportunity or best-practice deviation. Plan remediation for upcoming cycles."),
        ("", "", ""),
        ("=== GRADE DEFINITIONS ===", "", ""),
        ("Grade", "Score Range", "Definition"),
        ("Good", "80\u2013100", "Strong security posture. Address remaining findings as maintenance items."),
        ("Needs Improvement", "60\u201379", "Gaps exist that weaken security posture. Prioritize High and Critical findings."),
        ("Critical", "0\u201359", "Significant security risks are present. Immediate remediation is required."),
        ("", "", ""),
        ("=== SCORING FORMULA ===", "", ""),
        ("Component", "Detail", ""),
        ("Weight per check", "Critical=10, High=7, Medium=4, Low=2", ""),
        ("FAIL penalty", "Full weight (Critical=10, High=7, Medium=4, Low=2)", ""),
        ("WARN penalty", "Half weight (Critical=5, High=3.5, Medium=2, Low=1)", ""),
        ("PASS penalty", "0 (no penalty)", ""),
        ("Excluded", "NOT_APPLICABLE and API Error findings are excluded from weight totals", ""),
        ("Formula", "Score = (1 \u2212 penalty_sum / total_weights) \u00d7 100, where penalty = FAIL \u00d7 full_weight + WARN \u00d7 half_weight", ""),
        ("Interpretation", "A FAIL on a Critical check costs 10 pts; a WARN on a Critical check costs 5 pts. A FAIL on Low costs 2 pts; a WARN on Low costs 1 pt.", ""),
        ("", "", ""),
        ("=== COLUMN DEFINITIONS (Findings Sheets) ===", "", ""),
        ("Column", "Definition", ""),
        ("Check ID", "Unique identifier for this SAT check (e.g. SAT-IAM-001). Use this when referencing specific findings.", ""),
        ("Category", "Security domain this check belongs to (e.g. Identity & Access, Network Security, Data Protection).", ""),
        ("Severity", "Risk level if this check fails. Determines score weight and remediation priority.", ""),
        ("Status", "Result of the check: PASS, FAIL, WARN, or NOT_APPLICABLE. See status definitions above.", ""),
        ("Title", "Short descriptive name of the security control being verified.", ""),
        ("Description", "Detailed explanation of what the check evaluates and why it matters.", ""),
        ("Current State", "What the scanner found on your workspace \u2014 the actual configuration state detected.", ""),
        ("Recommendation", "Specific remediation steps to fix a FAIL or WARN finding.", ""),
        ("Reference URL", "Link to official Databricks documentation for the security control.", ""),
        ("", "", ""),
        ("=== CATEGORY DEFINITIONS ===", "", ""),
        ("Category", "Definition", ""),
        *((cat, defn, "") for cat, defn in CATEGORY_DEFINITIONS),
        ("", "", ""),
        ("=== HTTP ERROR CODES IN FINDINGS ===", "", ""),
        ("Code", "Meaning", "What to do"),
        ("401", "Unauthorized \u2014 token is invalid, expired, or revoked.", "Regenerate a new PAT token in Databricks \u2192 User Settings \u2192 Access Tokens, or re-login via Azure to refresh."),
        ("403", "Permission Denied \u2014 token authenticated but lacks the required admin role for this endpoint.", "Use a Workspace Admin PAT token. Admin-only endpoints: token-management, ip-access-lists, workspace-conf, Settings API."),
        ("404", "Not Found \u2014 the API endpoint does not exist on this workspace (feature not enabled or requires Premium pricing tier).", "Check the Azure Databricks Account Console for account-level settings. Verify your workspace pricing tier (Premium required for many features)."),
        ("400", "Bad Request \u2014 feature is not configured or workspace does not support that configuration key.", "Review raw details. This usually means the feature is managed differently (e.g. Unity Catalog instead of Table ACLs)."),
    ]
    if show_effort:
        def_rows.extend([
            ("", "", ""),
            ("=== REMEDIATION EFFORT LEVELS ===", "", ""),
            ("Level", "Time Range", "What\u2019s Included"),
            ("Quick Fix", "5\u201315 min", "Single configuration toggle or setting change in the admin console. No cross-team coordination or testing required."),
            ("Moderate", "1\u20134 hrs", "Multi-step configuration, policy creation, or changes that require testing and validation. May involve creating new resources, updating IaC code, or running build/deploy pipelines."),
            ("Significant", "1\u20133 days", "Architecture changes requiring cross-team coordination, downtime planning, IaC refactoring, and phased rollout. Includes prerequisite work like updating Terraform providers, modifying ARM/Bicep templates, and change management approvals."),
            ("Project", "1+ weeks", "Major infrastructure migration or org-wide policy rollout. Requires project planning, stakeholder approval, IaC redesign, multi-phase implementation, and extensive testing across environments."),
            ("", "", ""),
            ("=== COMMON PREREQUISITES (not included in estimates) ===", "", ""),
            ("Prerequisite", "When Needed", "Typical Additional Time"),
            ("Terraform / IaC Updates", "When infrastructure is managed via Terraform, Pulumi, ARM/Bicep, or other IaC tools. Changes must be codified, reviewed, and applied through the IaC pipeline.", "+30 min \u2013 2 hrs per change"),
            ("CI/CD Pipeline Runs", "When changes must flow through build, test, and deploy pipelines before reaching production.", "+15 min \u2013 1 hr per deployment"),
            ("Change Management / CAB Approval", "When your organization requires change tickets, CAB review, or ITSM approvals before production changes.", "+1 \u2013 5 business days"),
            ("Non-prod Testing", "When changes must be validated in dev/staging environments before production rollout.", "+1 \u2013 4 hrs per environment"),
            ("Provider / Module Upgrades", "When the Terraform azurerm/databricks provider or shared modules must be upgraded to support the required resource type or argument.", "+1 \u2013 4 hrs"),
            ("Security Review", "When network, identity, or security teams must review and approve the change before implementation.", "+1 \u2013 3 business days"),
            ("Documentation Updates", "When internal runbooks, architecture diagrams, or compliance documentation must be updated to reflect the change.", "+30 min \u2013 2 hrs"),
            ("", "", ""),
            ("=== HOW EFFORT IS ESTIMATED ===", "", ""),
            ("Factor", "Description", ""),
            ("Configuration Steps", "Number of settings, APIs, or UI steps required to remediate the finding.", ""),
            ("Access Requirements", "Whether changes need admin console access, API/IaC changes, or Azure Portal configuration.", ""),
            ("Testing & Validation", "Whether the fix needs functional testing, security validation, or user acceptance testing.", ""),
            ("Cross-team Coordination", "Whether network, security, identity, or platform teams need to be involved.", ""),
            ("Blast Radius", "Whether the change affects a single workspace, multiple workspaces, or the entire organization.", ""),
            ("", "", ""),
            ("Note:", "Effort estimates reflect core remediation work and are approximate. Actual time will vary based on your organization\u2019s change management processes, IaC maturity, CI/CD pipeline complexity, and team familiarity with Databricks administration.", ""),
        ])
    for r, (a, b, c) in enumerate(def_rows, 1):
        for col_idx, val in enumerate((a, b, c), 1):
            cell = ws_def.cell(row=r, column=col_idx, value=val)
            if isinstance(val, str) and val.startswith("="):
                cell.data_type = "s"
    ws_def.column_dimensions["A"].width = 30
    ws_def.column_dimensions["B"].width = 55
    ws_def.column_dimensions["C"].width = 80

    # ── Sheet 3: Category Scores ──
    ws_cat = wb.create_sheet("Category Scores")
    _set_header(ws_cat, 1, ["Category", "Score", "Grade", "Grade Definition"])
    for r, (cat, score) in enumerate(sorted(result.category_scores.items(), key=lambda x: x[1]), 2):
        grade = "Good" if score >= 80 else ("Needs Improvement" if score >= 60 else "Critical")
        grade_def = ("Strong security posture. Address remaining findings as maintenance items." if score >= 80
            else "Gaps exist that weaken security posture. Prioritize High and Critical findings." if score >= 60
            else "Significant security risks are present. Immediate remediation is required.")
        ws_cat.cell(row=r, column=1, value=cat).border = thin_border
        c2 = ws_cat.cell(row=r, column=2, value=score)
        c2.border = thin_border
        c2.fill = pass_fill if score >= 80 else (warn_fill if score >= 60 else fail_fill)
        ws_cat.cell(row=r, column=3, value=grade).border = thin_border
        ws_cat.cell(row=r, column=4, value=grade_def).border = thin_border
    ws_cat.column_dimensions["A"].width = 30
    ws_cat.column_dimensions["B"].width = 10
    ws_cat.column_dimensions["C"].width = 20
    ws_cat.column_dimensions["D"].width = 70

    # ── Sheet 4: All Findings ──
    ws_all = wb.create_sheet("All Findings")
    headers = ["Check ID", "Category", "Severity", "Status"]
    if show_effort:
        headers.append("Effort")
    headers.extend(["Title", "Description", "Current State", "Recommendation", "Reference URL"])
    if show_evidence:
        headers.extend(["Evidence Field", "Evidence Value"])
    headers.append("Portal Link")
    headers.append("Why It Matters")
    if include_api_response:
        headers.append("API Response")
    if show_scan_items:
        headers.append("Scanned Items")
    _set_header(ws_all, 1, headers)
    for r, f in enumerate(result.findings, 2):
        col = 1
        ws_all.cell(row=r, column=col, value=f.check_id).border = thin_border; col += 1
        ws_all.cell(row=r, column=col, value=f.category).border = thin_border; col += 1
        ws_all.cell(row=r, column=col, value=f.severity.upper()).border = thin_border; col += 1
        c4 = ws_all.cell(row=r, column=col, value=f.status)
        c4.border = thin_border
        c4.fill = _status_fill(f.status); col += 1
        if show_effort:
            ws_all.cell(row=r, column=col, value=f.effort or "").border = thin_border; col += 1
        ws_all.cell(row=r, column=col, value=f.title).border = thin_border; col += 1
        ws_all.cell(row=r, column=col, value=f.description).border = thin_border; col += 1
        ws_all.cell(row=r, column=col, value=f.current_state).border = thin_border; col += 1
        ws_all.cell(row=r, column=col, value=f.recommendation).border = thin_border; col += 1
        ws_all.cell(row=r, column=col, value=f.reference_url).border = thin_border; col += 1
        if show_evidence:
            ev = f.evidence or {}
            ws_all.cell(row=r, column=col, value=ev.get("field", "")).border = thin_border; col += 1
            val = ev.get("value", "")
            ws_all.cell(row=r, column=col, value=json.dumps(val, default=str) if not isinstance(val, str) else val).border = thin_border; col += 1
        portal_cell = ws_all.cell(row=r, column=col, value=f.portal_link or "")
        portal_cell.border = thin_border
        if f.portal_link:
            portal_cell.hyperlink = f.portal_link
            portal_cell.font = Font(color="0000FF", underline="single")
        col += 1
        ws_all.cell(row=r, column=col, value=f.benefits or "").border = thin_border; col += 1
        if include_api_response:
            ws_all.cell(row=r, column=col, value=_details_str(f.details, excel_safe=True)).border = thin_border; col += 1
        if show_scan_items:
            items_text = _format_scan_items(f.details)
            if len(items_text) > EXCEL_CELL_LIMIT:
                items_text = items_text[:EXCEL_CELL_LIMIT - 30] + "... [truncated for Excel]"
            ws_all.cell(row=r, column=col, value=items_text).border = thin_border
    col_widths = [14, 22, 10, 10, 40, 50, 40, 50, 40]
    if show_evidence:
        col_widths.extend([25, 30])
    col_widths.append(40)  # Portal Link
    col_widths.append(50)  # Why It Matters
    if include_api_response:
        col_widths.append(50)
    if show_scan_items:
        col_widths.append(50)
    for c, w in enumerate(col_widths, 1):
        ws_all.column_dimensions[chr(64 + c)].width = w

    # ── Sheet 5+: Per category ──
    by_cat: dict[str, list[SATFinding]] = {}
    for f in result.findings:
        by_cat.setdefault(f.category, []).append(f)
    for cat, findings in by_cat.items():
        import re
        sheet_name = re.sub(r'[\[\]*?/\\:]', '', cat)[:31]
        ws_c = wb.create_sheet(sheet_name)
        cat_headers = ["Check ID", "Severity", "Status"]
        if show_effort:
            cat_headers.append("Effort")
        cat_headers.extend(["Title", "Description", "Current State", "Recommendation", "Reference URL", "Portal Link", "Why It Matters"])
        if include_api_response:
            cat_headers.append("API Response")
        if show_scan_items:
            cat_headers.append("Scanned Items")
        _set_header(ws_c, 1, cat_headers)
        for r, f in enumerate(findings, 2):
            col = 1
            ws_c.cell(row=r, column=col, value=f.check_id).border = thin_border; col += 1
            ws_c.cell(row=r, column=col, value=f.severity.upper()).border = thin_border; col += 1
            c3 = ws_c.cell(row=r, column=col, value=f.status)
            c3.border = thin_border
            c3.fill = _status_fill(f.status); col += 1
            if show_effort:
                ws_c.cell(row=r, column=col, value=f.effort or "").border = thin_border; col += 1
            ws_c.cell(row=r, column=col, value=f.title).border = thin_border; col += 1
            ws_c.cell(row=r, column=col, value=f.description).border = thin_border; col += 1
            ws_c.cell(row=r, column=col, value=f.current_state).border = thin_border; col += 1
            ws_c.cell(row=r, column=col, value=f.recommendation).border = thin_border; col += 1
            ws_c.cell(row=r, column=col, value=f.reference_url).border = thin_border; col += 1
            portal_cell = ws_c.cell(row=r, column=col, value=f.portal_link or ""); portal_cell.border = thin_border
            if f.portal_link:
                portal_cell.hyperlink = f.portal_link
                portal_cell.font = Font(color="0000FF", underline="single")
            col += 1
            ws_c.cell(row=r, column=col, value=f.benefits or "").border = thin_border; col += 1
            if include_api_response:
                ws_c.cell(row=r, column=col, value=_details_str(f.details, excel_safe=True)).border = thin_border; col += 1
            if show_scan_items:
                items_text = _format_scan_items(f.details)
                if len(items_text) > EXCEL_CELL_LIMIT:
                    items_text = items_text[:EXCEL_CELL_LIMIT - 30] + "... [truncated for Excel]"
                ws_c.cell(row=r, column=col, value=items_text).border = thin_border

    # ── Failed Checks sheet ──
    failed = [f for f in result.findings if f.status == "FAIL"]
    if failed:
        ws_fail = wb.create_sheet("Failed Checks")
        fail_headers = ["Check ID", "Category", "Severity"]
        if show_effort:
            fail_headers.append("Effort")
        fail_headers.extend(["Title", "Current State", "Recommendation", "Reference URL", "Portal Link", "Why It Matters"])
        if include_api_response:
            fail_headers.append("API Response")
        if show_scan_items:
            fail_headers.append("Scanned Items")
        _set_header(ws_fail, 1, fail_headers)
        for r, f in enumerate(failed, 2):
            col = 1
            ws_fail.cell(row=r, column=col, value=f.check_id).border = thin_border; col += 1
            ws_fail.cell(row=r, column=col, value=f.category).border = thin_border; col += 1
            ws_fail.cell(row=r, column=col, value=f.severity.upper()).border = thin_border; col += 1
            if show_effort:
                ws_fail.cell(row=r, column=col, value=f.effort or "").border = thin_border; col += 1
            ws_fail.cell(row=r, column=col, value=f.title).border = thin_border; col += 1
            ws_fail.cell(row=r, column=col, value=f.current_state).border = thin_border; col += 1
            ws_fail.cell(row=r, column=col, value=f.recommendation).border = thin_border; col += 1
            ws_fail.cell(row=r, column=col, value=f.reference_url).border = thin_border; col += 1
            portal_cell = ws_fail.cell(row=r, column=col, value=f.portal_link or ""); portal_cell.border = thin_border
            if f.portal_link:
                portal_cell.hyperlink = f.portal_link
                portal_cell.font = Font(color="0000FF", underline="single")
            col += 1
            ws_fail.cell(row=r, column=col, value=f.benefits or "").border = thin_border; col += 1
            if include_api_response:
                ws_fail.cell(row=r, column=col, value=_details_str(f.details, excel_safe=True)).border = thin_border; col += 1
            if show_scan_items:
                items_text = _format_scan_items(f.details)
                if len(items_text) > EXCEL_CELL_LIMIT:
                    items_text = items_text[:EXCEL_CELL_LIMIT - 30] + "... [truncated for Excel]"
                ws_fail.cell(row=r, column=col, value=items_text).border = thin_border

    # ── Warnings sheet ──
    warns = [f for f in result.findings if f.status == "WARN" and not f.is_api_error]
    if warns:
        ws_warn = wb.create_sheet("Warnings")
        warn_headers = ["Check ID", "Category", "Severity"]
        if show_effort:
            warn_headers.append("Effort")
        warn_headers.extend(["Title", "Current State", "Recommendation", "Reference URL", "Portal Link", "Why It Matters"])
        _set_header(ws_warn, 1, warn_headers)
        for r, f in enumerate(warns, 2):
            col = 1
            ws_warn.cell(row=r, column=col, value=f.check_id).border = thin_border; col += 1
            ws_warn.cell(row=r, column=col, value=f.category).border = thin_border; col += 1
            ws_warn.cell(row=r, column=col, value=f.severity.upper()).border = thin_border; col += 1
            if show_effort:
                ws_warn.cell(row=r, column=col, value=f.effort or "").border = thin_border; col += 1
            ws_warn.cell(row=r, column=col, value=f.title).border = thin_border; col += 1
            ws_warn.cell(row=r, column=col, value=f.current_state).border = thin_border; col += 1
            ws_warn.cell(row=r, column=col, value=f.recommendation).border = thin_border; col += 1
            ws_warn.cell(row=r, column=col, value=f.reference_url).border = thin_border; col += 1
            portal_cell = ws_warn.cell(row=r, column=col, value=f.portal_link or ""); portal_cell.border = thin_border
            if f.portal_link:
                portal_cell.hyperlink = f.portal_link
                portal_cell.font = Font(color="0000FF", underline="single")
            col += 1
            ws_warn.cell(row=r, column=col, value=f.benefits or "").border = thin_border

    # ── N/A (Not Applicable) sheet ──
    na_findings = [f for f in result.findings if f.status == "NOT_APPLICABLE" and not f.is_api_error]
    if na_findings:
        ws_na = wb.create_sheet("Not Applicable")
        na_headers = ["Check ID", "Category", "Severity", "Title", "Reason", "Portal Link", "Why It Matters"]
        _set_header(ws_na, 1, na_headers)
        for r, f in enumerate(na_findings, 2):
            col = 1
            ws_na.cell(row=r, column=col, value=f.check_id).border = thin_border; col += 1
            ws_na.cell(row=r, column=col, value=f.category).border = thin_border; col += 1
            ws_na.cell(row=r, column=col, value=f.severity.upper()).border = thin_border; col += 1
            ws_na.cell(row=r, column=col, value=f.title).border = thin_border; col += 1
            ws_na.cell(row=r, column=col, value=f.current_state).border = thin_border; col += 1
            portal_cell = ws_na.cell(row=r, column=col, value=f.portal_link or ""); portal_cell.border = thin_border
            if f.portal_link:
                portal_cell.hyperlink = f.portal_link
                portal_cell.font = Font(color="0000FF", underline="single")
            col += 1
            ws_na.cell(row=r, column=col, value=f.benefits or "").border = thin_border

    # ── Passed Checks sheet ──
    passed = [f for f in result.findings if f.status == "PASS"]
    if passed:
        ws_pass = wb.create_sheet("Passed Checks")
        pass_headers = ["Check ID", "Category", "Severity", "Title", "Current State", "Portal Link", "Why It Matters"]
        _set_header(ws_pass, 1, pass_headers)
        for r, f in enumerate(passed, 2):
            col = 1
            ws_pass.cell(row=r, column=col, value=f.check_id).border = thin_border; col += 1
            ws_pass.cell(row=r, column=col, value=f.category).border = thin_border; col += 1
            ws_pass.cell(row=r, column=col, value=f.severity.upper()).border = thin_border; col += 1
            ws_pass.cell(row=r, column=col, value=f.title).border = thin_border; col += 1
            ws_pass.cell(row=r, column=col, value=f.current_state).border = thin_border; col += 1
            portal_cell = ws_pass.cell(row=r, column=col, value=f.portal_link or ""); portal_cell.border = thin_border
            if f.portal_link:
                portal_cell.hyperlink = f.portal_link
                portal_cell.font = Font(color="0000FF", underline="single")
            col += 1
            ws_pass.cell(row=r, column=col, value=f.benefits or "").border = thin_border

    # ── API Errors sheet ──
    api_errs = [f for f in result.findings if f.is_api_error]
    if api_errs:
        api_err_fill = PatternFill("solid", fgColor="D6DCE4")
        ws_api = wb.create_sheet("API Errors")
        api_headers = ["Check ID", "Category", "Severity", "Title", "HTTP Status", "Error Detail", "Justification", "Recommendation", "Reference URL", "Why It Matters"]
        _set_header(ws_api, 1, api_headers)
        for r, f in enumerate(api_errs, 2):
            http_code = f.details.get("http_status", "") if f.details else ""
            err_detail = f.current_state
            justification = f.details.get("justification", "") if f.details else ""
            ws_api.cell(row=r, column=1, value=f.check_id).border = thin_border
            ws_api.cell(row=r, column=2, value=f.category).border = thin_border
            ws_api.cell(row=r, column=3, value=f.severity.upper()).border = thin_border
            ws_api.cell(row=r, column=4, value=f.title).border = thin_border
            c5 = ws_api.cell(row=r, column=5, value=http_code if http_code else "Exception")
            c5.border = thin_border
            c5.fill = api_err_fill
            ws_api.cell(row=r, column=6, value=err_detail).border = thin_border
            ws_api.cell(row=r, column=7, value=justification).border = thin_border
            ws_api.cell(row=r, column=8, value=f.recommendation).border = thin_border
            ws_api.cell(row=r, column=9, value=f.reference_url).border = thin_border
            ws_api.cell(row=r, column=10, value=f.benefits or "").border = thin_border
        ws_api.column_dimensions["A"].width = 18
        ws_api.column_dimensions["B"].width = 22
        ws_api.column_dimensions["C"].width = 10
        ws_api.column_dimensions["D"].width = 40
        ws_api.column_dimensions["E"].width = 14
        ws_api.column_dimensions["F"].width = 60
        ws_api.column_dimensions["G"].width = 60
        ws_api.column_dimensions["H"].width = 50
        ws_api.column_dimensions["I"].width = 40
        ws_api.column_dimensions["J"].width = 50

    # ── Prioritised Recommendations sheet ──
    prio_items = _build_prioritised_recommendations(result.findings)
    if prio_items:
        p1_fill = PatternFill("solid", fgColor="FFC7CE")  # red
        p2_fill = PatternFill("solid", fgColor="FFEB9C")  # amber
        p3_fill = PatternFill("solid", fgColor="C6EFCE")  # green
        p4_fill = PatternFill("solid", fgColor="D9D9D9")  # grey
        _prio_fills = {"P1": p1_fill, "P2": p2_fill, "P3": p3_fill, "P4": p4_fill}
        ws_prio = wb.create_sheet("Prioritised Recommendations")
        ws_prio.cell(row=1, column=1, value="Prioritised Recommendations — Fix High-Impact, Low-Effort Items First").font = title_font
        ws_prio.merge_cells("A1:F1")
        ws_prio.cell(row=2, column=1, value=f"Actionable findings: {len(prio_items)} (FAIL + WARN) — sorted by Priority Score (severity × effort multiplier)").font = Font(italic=True, size=10, color="64748B")
        _prio_hdr_start = 4
        if show_cost:
            ws_prio.cell(row=3, column=1, value="Note: Cost figures are illustrative examples only — actual costs vary with usage, region, and pricing tier.").font = Font(italic=True, size=9, color="94A3B8")
        prio_hdrs = ["Priority", "Score", "Check ID", "Category", "Severity", "Status", "Effort"]
        if show_cost:
            prio_hdrs += ["Est. Cost Low ($/mo)", "Est. Cost High ($/mo)", "Cost Reason"]
        prio_hdrs += ["Title", "Recommendation", "Why It Matters", "Portal Link"]
        _set_header(ws_prio, _prio_hdr_start, prio_hdrs)
        usd_fmt = '#,##0'
        for r, item in enumerate(prio_items, _prio_hdr_start + 1):
            col = 1
            prio_cell = ws_prio.cell(row=r, column=col, value=item["priority_label"])
            prio_cell.border = thin_border
            prio_cell.font = Font(bold=True)
            prio_prefix = item["priority_label"][:2]
            prio_cell.fill = _prio_fills.get(prio_prefix, p4_fill)
            col += 1
            ws_prio.cell(row=r, column=col, value=item["priority_score"]).border = thin_border; col += 1
            ws_prio.cell(row=r, column=col, value=item["check_id"]).border = thin_border; col += 1
            ws_prio.cell(row=r, column=col, value=item["category"]).border = thin_border; col += 1
            ws_prio.cell(row=r, column=col, value=item["severity"].upper()).border = thin_border; col += 1
            status_cell = ws_prio.cell(row=r, column=col, value=item["status"])
            status_cell.border = thin_border
            status_cell.fill = _status_fill(item["status"]); col += 1
            ws_prio.cell(row=r, column=col, value=item["effort"]).border = thin_border; col += 1
            if show_cost:
                cost_low_cell = ws_prio.cell(row=r, column=col, value=item["cost_low"] if item["cost_low"] else None)
                cost_low_cell.border = thin_border
                if item["cost_low"]:
                    cost_low_cell.number_format = usd_fmt
                col += 1
                cost_high_cell = ws_prio.cell(row=r, column=col, value=item["cost_high"] if item["cost_high"] else None)
                cost_high_cell.border = thin_border
                if item["cost_high"]:
                    cost_high_cell.number_format = usd_fmt
                col += 1
                ws_prio.cell(row=r, column=col, value=item["cost_reason"]).border = thin_border; col += 1
            ws_prio.cell(row=r, column=col, value=item["title"]).border = thin_border; col += 1
            ws_prio.cell(row=r, column=col, value=item["recommendation"]).border = thin_border; col += 1
            ws_prio.cell(row=r, column=col, value=item["benefits"]).border = thin_border; col += 1
            portal_cell = ws_prio.cell(row=r, column=col, value=item["portal_link"])
            portal_cell.border = thin_border
            if item["portal_link"]:
                portal_cell.hyperlink = item["portal_link"]
                portal_cell.font = Font(color="0000FF", underline="single")
        if show_cost:
            # Total estimated cost row
            _cost_items = [i for i in prio_items if i["cost_low"]]
            if _cost_items:
                total_row = _prio_hdr_start + 1 + len(prio_items) + 1
                ws_prio.cell(row=total_row, column=7, value="TOTAL EST. COST").font = Font(bold=True, size=10)
                total_low = sum(i["cost_low"] for i in _cost_items)
                total_high = sum(i["cost_high"] for i in _cost_items)
                tl = ws_prio.cell(row=total_row, column=8, value=total_low)
                tl.font = Font(bold=True); tl.number_format = usd_fmt
                th = ws_prio.cell(row=total_row, column=9, value=total_high)
                th.font = Font(bold=True); th.number_format = usd_fmt
        ws_prio.column_dimensions["A"].width = 24
        ws_prio.column_dimensions["B"].width = 8
        ws_prio.column_dimensions["C"].width = 22
        ws_prio.column_dimensions["D"].width = 22
        ws_prio.column_dimensions["E"].width = 10
        ws_prio.column_dimensions["F"].width = 8
        ws_prio.column_dimensions["G"].width = 22
        if show_cost:
            ws_prio.column_dimensions["H"].width = 18
            ws_prio.column_dimensions["I"].width = 18
            ws_prio.column_dimensions["J"].width = 50
            ws_prio.column_dimensions["K"].width = 40
            ws_prio.column_dimensions["L"].width = 50
            ws_prio.column_dimensions["M"].width = 50
            ws_prio.column_dimensions["N"].width = 40
        else:
            ws_prio.column_dimensions["H"].width = 40
            ws_prio.column_dimensions["I"].width = 50
            ws_prio.column_dimensions["J"].width = 50
        ws_prio.column_dimensions["K"].width = 40
        ws_prio.column_dimensions["L"].width = 50
        ws_prio.column_dimensions["M"].width = 50
        ws_prio.column_dimensions["N"].width = 40

    # ── Shared fill for critical severity (used in Endpoints + Reference sheets) ──
    error_fill = PatternFill("solid", fgColor="FFC7CE")

    # ── API Endpoints sheet ──
    if result.endpoint_summary and result.endpoint_summary.get("endpoints"):
        ep = result.endpoint_summary
        ws_ep = wb.create_sheet("API Endpoints")
        ws_ep.cell(row=1, column=1, value="API Endpoint Summary").font = title_font
        ws_ep.merge_cells("A1:C1")
        _set_header(ws_ep, 3, ["Metric", "Value"])
        summary_kv = [
            ("Total Endpoints", ep["total"]),
            ("With Items", ep["with_items"]),
            ("Config/Settings", ep["config"]),
            ("Empty", ep["empty"]),
            ("Errors", ep.get("error", 0)),
        ]
        for r, (metric, val) in enumerate(summary_kv, 4):
            ws_ep.cell(row=r, column=1, value=metric).border = thin_border
            ws_ep.cell(row=r, column=2, value=val).border = thin_border

        ep_start = 4 + len(summary_kv) + 1
        _set_header(ws_ep, ep_start, ["Endpoint", "Status", "Items Count", "Error Code"])
        items_fill = PatternFill("solid", fgColor="C6EFCE")
        config_fill = PatternFill("solid", fgColor="FFEB9C")
        empty_fill = PatternFill("solid", fgColor="D9D9D9")
        error_fill = PatternFill("solid", fgColor="FFC7CE")
        for r, e in enumerate(ep["endpoints"], ep_start + 1):
            ws_ep.cell(row=r, column=1, value=e["endpoint"]).border = thin_border
            status_cell = ws_ep.cell(row=r, column=2, value=e["status"])
            status_cell.border = thin_border
            if e["status"] == "items":
                status_cell.fill = items_fill
            elif e["status"] == "config":
                status_cell.fill = config_fill
            elif e["status"] == "error":
                status_cell.fill = error_fill
            else:
                status_cell.fill = empty_fill
            ws_ep.cell(row=r, column=3, value=e["items_count"]).border = thin_border
            err_code = e.get("error_code", 0)
            err_cell = ws_ep.cell(row=r, column=4, value=err_code if err_code else "")
            err_cell.border = thin_border
            if err_code:
                err_cell.fill = error_fill
        ws_ep.column_dimensions["A"].width = 55
        ws_ep.column_dimensions["B"].width = 18
        ws_ep.column_dimensions["C"].width = 14
        ws_ep.column_dimensions["D"].width = 14

    # ── All Checks Reference sheet ──
    ws_ref = wb.create_sheet("All Checks Reference")
    ref_hdrs = ["Check ID", "Category", "Severity"]
    if show_effort:
        ref_hdrs.append("Effort")
    ref_hdrs.extend(["Title", "Description", "Recommendation"])
    ws_ref.cell(row=1, column=1, value=f"All {len(SAT_CHECKS)} Checks Reference").font = title_font
    ws_ref.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(ref_hdrs))
    _set_header(ws_ref, 3, ref_hdrs)
    _sev_sort_xl = {"critical": 0, "high": 1, "medium": 2, "low": 3}
    _ref_items = sorted(SAT_CHECKS.items(), key=lambda x: (x[1].get("category", ""), _sev_sort_xl.get(x[1].get("severity", "low"), 3), x[0]))
    for idx, (cid, ck) in enumerate(_ref_items, start=4):
        col = 1
        ws_ref.cell(row=idx, column=col, value=cid).border = thin_border; col += 1
        ws_ref.cell(row=idx, column=col, value=ck.get("category", "")).border = thin_border; col += 1
        sev_cell = ws_ref.cell(row=idx, column=col, value=ck.get("severity", "").upper())
        sev_cell.border = thin_border
        sev_val = ck.get("severity", "low")
        if sev_val == "critical":
            sev_cell.fill = error_fill
            sev_cell.font = Font(bold=True, color="FFFFFF")
        elif sev_val == "high":
            sev_cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        elif sev_val == "medium":
            sev_cell.fill = warn_fill
        col += 1
        if show_effort:
            ws_ref.cell(row=idx, column=col, value=_get_effort(cid)).border = thin_border; col += 1
        ws_ref.cell(row=idx, column=col, value=ck.get("title", "")).border = thin_border; col += 1
        ws_ref.cell(row=idx, column=col, value=ck.get("description", "")).border = thin_border; col += 1
        ws_ref.cell(row=idx, column=col, value=ck.get("recommendation", "")).border = thin_border
    ws_ref.column_dimensions["A"].width = 22
    ws_ref.column_dimensions["B"].width = 24
    ws_ref.column_dimensions["C"].width = 12
    _ref_col = 4
    if show_effort:
        ws_ref.column_dimensions[chr(64 + _ref_col)].width = 22; _ref_col += 1
    ws_ref.column_dimensions[chr(64 + _ref_col)].width = 50; _ref_col += 1
    ws_ref.column_dimensions[chr(64 + _ref_col)].width = 70; _ref_col += 1
    ws_ref.column_dimensions[chr(64 + _ref_col)].width = 70

    # ── Remediation Plan sheet ──
    _prio_items = _build_prioritised_recommendations(result.findings)
    if _prio_items:
        ws_rp = wb.create_sheet("Remediation Plan")
        rp_headers = ["Check ID", "Title", "Category", "Severity", "Priority",
                       "Prerequisites", "Pre-Checks", "Remediation Steps",
                       "Post-Validation", "Rollback", "Downtime", "Blast Radius",
                       "Est. Hours", "Stakeholders", "Change Type", "Approval Required"]
        for col_i, h in enumerate(rp_headers, 1):
            c = ws_rp.cell(row=1, column=col_i, value=h)
            c.font = header_font; c.fill = header_fill; c.border = thin_border
        for ri, item in enumerate(_prio_items, 2):
            plan = item.get("remediation_plan", {})
            cl = plan.get("checklist", {})
            ia = plan.get("impact_assessment", {})
            cm = plan.get("change_management", {})
            vals = [
                item["check_id"], item["title"], item["category"],
                item["severity"].upper(), item.get("priority_label", ""),
                "; ".join(plan.get("prerequisites", [])),
                "; ".join(cl.get("pre_checks", [])),
                "; ".join(cl.get("steps", [])),
                "; ".join(cl.get("post_validation", [])),
                "; ".join(cl.get("rollback", [])),
                ia.get("downtime", ""), ia.get("blast_radius", ""),
                plan.get("estimated_duration_hours", ""),
                "; ".join(plan.get("stakeholders", [])),
                cm.get("change_type", ""),
                "Yes" if cm.get("approval_required") else "No",
            ]
            for col_i, v in enumerate(vals, 1):
                ws_rp.cell(row=ri, column=col_i, value=v).border = thin_border
        for ci, w in enumerate([18, 40, 20, 10, 10, 40, 40, 50, 40, 40, 10, 14, 10, 30, 12, 12], 1):
            ws_rp.column_dimensions[chr(64 + ci) if ci <= 26 else "A" + chr(64 + ci - 26)].width = w

        # ── Remediation Timeline sheet ──
        from .remediation import build_remediation_timeline
        timeline = build_remediation_timeline(_prio_items)
        ws_rt = wb.create_sheet("Remediation Timeline")
        # Summary header
        ws_rt.cell(row=1, column=1, value="Remediation Roadmap").font = Font(bold=True, size=14)
        _ts = timeline["summary"]
        ws_rt.cell(row=2, column=1, value="Total Findings"); ws_rt.cell(row=2, column=2, value=_ts["total_findings"])
        ws_rt.cell(row=3, column=1, value="Total Effort (hours)"); ws_rt.cell(row=3, column=2, value=_ts["total_effort_hours"])
        ws_rt.cell(row=4, column=1, value="Total Working Days"); ws_rt.cell(row=4, column=2, value=_ts["total_working_days"])
        ws_rt.cell(row=5, column=1, value="Estimated Resources"); ws_rt.cell(row=5, column=2, value=_ts["estimated_resources"])
        ws_rt.cell(row=6, column=1, value="Categories"); ws_rt.cell(row=6, column=2, value=_ts["categories"])
        ws_rt.cell(row=7, column=1, value="Note"); ws_rt.cell(row=7, column=2, value="Resources = parallel team members needed per phase (8h/day). P1: 1wk, P2: 2wks, P3: 5wks, P4: flexible.")
        rt_row = 9
        rt_headers = ["Phase", "Category", "Check ID", "Title", "Severity", "Effort", "Est. Hours", "Working Days", "Est. Resources"]
        for col_i, h in enumerate(rt_headers, 1):
            c = ws_rt.cell(row=rt_row, column=col_i, value=h)
            c.font = header_font; c.fill = header_fill; c.border = thin_border
        rt_row += 1
        for phase in timeline["phases"]:
            for cat, cat_data in phase["categories"].items():
                for f in cat_data["findings"]:
                    vals = [phase["phase"], cat, f["check_id"], f["title"],
                            f["severity"], f["effort"], f["effort_hours"], f["working_days"], ""]
                    for col_i, v in enumerate(vals, 1):
                        ws_rt.cell(row=rt_row, column=col_i, value=v).border = thin_border
                    rt_row += 1
            # Phase subtotal row
            if phase["total_hours"]:
                ws_rt.cell(row=rt_row, column=1, value=f"{phase['phase']} — Total").font = Font(bold=True)
                ws_rt.cell(row=rt_row, column=7, value=phase["total_hours"]).font = Font(bold=True)
                ws_rt.cell(row=rt_row, column=8, value=phase["total_working_days"]).font = Font(bold=True)
                ws_rt.cell(row=rt_row, column=9, value=phase["estimated_resources"]).font = Font(bold=True)
                rt_row += 1
        for ci, w in enumerate([30, 22, 18, 40, 10, 22, 12, 14, 14], 1):
            ws_rt.column_dimensions[chr(64 + ci)].width = w

    # ── Enrich with Unity Catalog + Azure inventory (when --with-inventory) ──
    inv_obj = getattr(result, "inventory_obj", None)
    if inv_obj is not None:
        try:
            _build_inventory_sheets(wb, inv_obj, prefix="UC ")
        except Exception:
            pass

    path = output_dir / f"{_file_prefix(result)}.xlsx"
    wb.save(str(path))
    return str(path)


def export_html(result: SATScanResult, output_dir: Path, include_api_response: bool = True, summary_link: str = "", show_scan_items: bool = False, show_evidence: bool = False, show_effort: bool = False, show_cost: bool = False) -> str:
    import html as _html
    _esc = _html.escape

    ts = datetime.fromisoformat(result.scanned_at.replace("Z", "+00:00")).strftime("%Y-%m-%d %H:%M:%S UTC")

    def _icon(name: str, size: int = 16) -> str:
        return f'<i data-lucide="{name}" style="width:{size}px;height:{size}px;vertical-align:middle;margin-right:4px"></i>'

    def _portal_label(link: str) -> str:
        if "portal.azure.com" in link:
            return "Azure Portal"
        if "accounts.azuredatabricks.net" in link:
            return "Account Console"
        return "Open"

    def _portal_label_long(link: str) -> str:
        if "portal.azure.com" in link:
            return "Open in Azure Portal"
        if "accounts.azuredatabricks.net" in link:
            return "Open Account Console"
        return "Open in Workspace"

    _SEV_TIPS = {
        "critical": "Immediate security risk — must be remediated immediately (weight: 10)",
        "high": "Significant security weakness — remediate within days (weight: 7)",
        "medium": "Moderate risk — remediate within the current sprint (weight: 4)",
        "low": "Minor improvement opportunity — plan for upcoming cycles (weight: 2)",
        "pass": "Check passed — no action needed",
    }
    _STATUS_TIPS = {
        "PASS": "Compliant — security control is correctly configured",
        "FAIL": "Action Required — confirmed security gap that must be fixed",
        "WARN": "Review Needed — borderline result that needs manual investigation",
        "NOT_APPLICABLE": "Skipped — feature not in use on this workspace, excluded from score",
        "API ERROR": "Could Not Evaluate — API failure prevented this check from running",
    }

    def sev_badge(s: str) -> str:
        colors = {"critical": "#dc2626", "high": "#ea580c", "medium": "#ca8a04", "low": "#2563eb", "pass": "#16a34a"}
        tip = _esc(_SEV_TIPS.get(s, ""))
        return f'<span class="tip" style="background:{colors.get(s,"#6b7280")};color:#fff;padding:2px 8px;border-radius:9999px;font-size:11px;font-weight:600;text-transform:uppercase"><span class="tip-text">{tip}</span>{_esc(s)}</span>'

    def status_badge(s: str) -> str:
        colors = {"PASS": "#16a34a", "FAIL": "#dc2626", "WARN": "#ca8a04", "NOT_APPLICABLE": "#6b7280"}
        tip = _esc(_STATUS_TIPS.get(s, ""))
        return f'<span class="tip" style="background:{colors.get(s,"#6b7280")};color:#fff;padding:2px 8px;border-radius:9999px;font-size:11px;font-weight:600"><span class="tip-text">{tip}</span>{_esc(s)}</span>'

    # ── Score gauge (SVG) ──
    score = result.overall_score
    gauge_color = "#16a34a" if score >= 80 else ("#ca8a04" if score >= 60 else "#dc2626")
    grade = "Good" if score >= 80 else ("Needs Improvement" if score >= 60 else "Critical")
    dash = round(score / 100 * 439.8, 1)
    score_gauge = f"""<div style="text-align:center;margin-bottom:24px">
<svg width="160" height="160" viewBox="0 0 160 160">
  <circle cx="80" cy="80" r="70" fill="none" stroke="#e2e8f0" stroke-width="12"/>
  <circle cx="80" cy="80" r="70" fill="none" stroke="{gauge_color}" stroke-width="12"
    stroke-dasharray="{dash} 440" stroke-dashoffset="0"
    transform="rotate(-90 80 80)" stroke-linecap="round"/>
  <text x="80" y="72" text-anchor="middle" font-size="32" font-weight="700" fill="{gauge_color}">{score}</text>
  <text x="80" y="96" text-anchor="middle" font-size="13" fill="#64748b">{_esc(grade)}</text>
</svg></div>"""

    # ── Score breakdown ──
    _SEV_W = {"critical": 10, "high": 7, "medium": 4, "low": 2}
    _scorable = [f for f in result.findings if not f.is_api_error]
    _applicable = [f for f in _scorable if f.status != "NOT_APPLICABLE"]
    _fail_checks = [f for f in _applicable if f.status == "FAIL"]
    _warn_checks = [f for f in _applicable if f.status == "WARN"]
    _pass_checks = [f for f in _applicable if f.status == "PASS"]
    _total_wt = sum(_SEV_W.get(SAT_CHECKS.get(f.check_id, {}).get("severity", "low"), 2) for f in _applicable)
    _fail_pen = sum(_SEV_W.get(SAT_CHECKS.get(f.check_id, {}).get("severity", "low"), 2) for f in _fail_checks)
    _warn_pen = sum(_SEV_W.get(SAT_CHECKS.get(f.check_id, {}).get("severity", "low"), 2) * 0.5 for f in _warn_checks)
    _total_pen = _fail_pen + _warn_pen
    score_breakdown = f"""<div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:12px;padding:20px;margin:0 auto 24px;max-width:560px">
<h4 style="margin:0 0 12px;font-size:14px;font-weight:700;color:#0f172a">Score Breakdown</h4>
<table style="font-size:13px;border:none">
<tr><td style="border:none;padding:4px 12px 4px 0;color:#64748b">Scored checks</td><td style="border:none;padding:4px 0;font-weight:600">{len(_applicable)} <span style="color:#94a3b8;font-weight:400">(excl. {result.not_applicable} N/A + {result.api_errors} API Error)</span></td></tr>
<tr><td style="border:none;padding:4px 12px 4px 0;color:#64748b">Total weight pool</td><td style="border:none;padding:4px 0;font-weight:600">{_total_wt:.0f} pts</td></tr>
<tr><td style="border:none;padding:4px 12px 4px 0;color:#dc2626">FAIL penalty <span style="color:#94a3b8">({len(_fail_checks)} checks &times; full weight)</span></td><td style="border:none;padding:4px 0;font-weight:600;color:#dc2626">&minus;{_fail_pen:.0f} pts</td></tr>
<tr><td style="border:none;padding:4px 12px 4px 0;color:#ca8a04">WARN penalty <span style="color:#94a3b8">({len(_warn_checks)} checks &times; half weight)</span></td><td style="border:none;padding:4px 0;font-weight:600;color:#ca8a04">&minus;{_warn_pen:.1f} pts</td></tr>
<tr><td style="border:none;padding:4px 12px 4px 0;color:#16a34a">PASS <span style="color:#94a3b8">({len(_pass_checks)} checks)</span></td><td style="border:none;padding:4px 0;font-weight:600;color:#16a34a">0 pts</td></tr>
<tr style="border-top:1px solid #e2e8f0"><td style="border:none;padding:8px 12px 4px 0;color:#0f172a;font-weight:600">Formula</td><td style="border:none;padding:8px 0 4px;font-family:monospace;font-size:12px">(1 &minus; {_total_pen:.1f} / {_total_wt:.0f}) &times; 100 = <strong style="color:{gauge_color};font-size:14px">{score}</strong></td></tr>
</table>
<div style="display:flex;gap:8px;margin-top:14px;font-size:12px">
<span style="padding:3px 10px;border-radius:9999px;font-weight:600;{'background:#dcfce7;color:#16a34a' if score >= 80 else 'background:#f1f5f9;color:#94a3b8'}">Good 80&ndash;100</span>
<span style="padding:3px 10px;border-radius:9999px;font-weight:600;{'background:#fef9c3;color:#ca8a04' if 60 <= score < 80 else 'background:#f1f5f9;color:#94a3b8'}">Needs Improvement 60&ndash;79</span>
<span style="padding:3px 10px;border-radius:9999px;font-weight:600;{'background:#fee2e2;color:#dc2626' if score < 60 else 'background:#f1f5f9;color:#94a3b8'}">Critical 0&ndash;59</span>
</div></div>"""

    # ── Category scores table with progress bars ──
    cat_score_rows = ""
    for cat, cat_score in sorted(result.category_scores.items(), key=lambda x: x[1]):
        c = "#16a34a" if cat_score >= 80 else ("#ca8a04" if cat_score >= 60 else "#dc2626")
        g = "Good" if cat_score >= 80 else ("Needs Improvement" if cat_score >= 60 else "Critical")
        bar_w = max(2, cat_score)
        cat_score_rows += f"""<tr><td>{_esc(cat)}</td>
<td style="font-weight:700;color:{c};text-align:right">{cat_score}</td>
<td><div style="background:#e2e8f0;border-radius:4px;height:8px;width:100%"><div style="background:{c};border-radius:4px;height:8px;width:{bar_w}%"></div></div></td>
<td style="color:{c};font-size:12px">{_esc(g)}</td></tr>"""
    # ── Effort summary (only when --effort) ──
    effort_summary_html = ""
    if show_effort:
        _actionable = [f for f in result.findings if f.status in ("FAIL", "WARN") and not f.is_api_error]
        _eff_counts: dict[str, int] = {}
        for f in _actionable:
            e = f.effort or "Moderate (1–4 hrs)"
            _eff_counts[e] = _eff_counts.get(e, 0) + 1
        _eff_order = ["Quick Fix (5–15 min)", "Moderate (1–4 hrs)", "Significant (1–3 days)", "Project (1+ weeks)"]
        _eff_colors = {"Quick Fix (5–15 min)": "#16a34a", "Moderate (1–4 hrs)": "#2563eb", "Significant (1–3 days)": "#ca8a04", "Project (1+ weeks)": "#dc2626"}
        _eff_icons = {"Quick Fix (5–15 min)": "zap", "Moderate (1–4 hrs)": "wrench", "Significant (1–3 days)": "hard-hat", "Project (1+ weeks)": "building"}
        _eff_bars = ""
        for eff in _eff_order:
            cnt = _eff_counts.get(eff, 0)
            if cnt == 0:
                continue
            pct = round(cnt / len(_actionable) * 100) if _actionable else 0
            c = _eff_colors.get(eff, "#6b7280")
            _eff_bars += f"""<tr><td style="white-space:nowrap;font-size:13px;padding:6px 12px 6px 0">{_icon(_eff_icons.get(eff, 'clock'), 14)} {_esc(eff)}</td>
<td style="font-weight:700;color:{c};text-align:right;padding:6px 12px 6px 0;font-size:14px">{cnt}</td>
<td style="width:200px;padding:6px 0"><div style="background:#e2e8f0;border-radius:4px;height:8px;width:100%"><div style="background:{c};border-radius:4px;height:8px;width:{max(2, pct)}%"></div></div></td>
<td style="color:#64748b;font-size:12px;padding:6px 0 6px 8px">{pct}%</td></tr>"""
        if _actionable:
            effort_summary_html = f"""
<div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:12px;padding:20px;margin:24px auto;max-width:560px">
<h4 style="margin:0 0 12px;font-size:14px;font-weight:700;color:#0f172a">{_icon('clock', 16)} Remediation Effort Summary</h4>
<p style="font-size:12px;color:#64748b;margin:0 0 12px">{len(_actionable)} actionable finding{"s" if len(_actionable) != 1 else ""} (FAIL + WARN) by estimated remediation effort:</p>
<table style="border:none;width:100%">{_eff_bars}</table>
<div style="background:#fffbeb;border:1px solid #fcd34d;border-radius:8px;padding:12px 14px;margin-top:14px">
<p style="margin:0 0 8px;font-size:12px;font-weight:600;color:#92400e">{_icon('alert-triangle', 13)} Prerequisites Not Included in Estimates</p>
<p style="margin:0 0 6px;font-size:11px;color:#78350f">The times above reflect core remediation work only. Factor in additional time for:</p>
<div style="display:flex;flex-wrap:wrap;gap:6px;font-size:11px">
<span style="background:#fef3c7;color:#92400e;padding:2px 8px;border-radius:9999px">Terraform / IaC updates</span>
<span style="background:#fef3c7;color:#92400e;padding:2px 8px;border-radius:9999px">CI/CD pipeline runs</span>
<span style="background:#fef3c7;color:#92400e;padding:2px 8px;border-radius:9999px">Change management / CAB approval</span>
<span style="background:#fef3c7;color:#92400e;padding:2px 8px;border-radius:9999px">Non-prod testing</span>
<span style="background:#fef3c7;color:#92400e;padding:2px 8px;border-radius:9999px">Provider / module upgrades</span>
<span style="background:#fef3c7;color:#92400e;padding:2px 8px;border-radius:9999px">Security review</span>
<span style="background:#fef3c7;color:#92400e;padding:2px 8px;border-radius:9999px">Documentation updates</span>
</div>
<p style="margin:6px 0 0;font-size:11px;color:#78350f;font-style:italic">See the <strong>Definitions</strong> tab for full prerequisite details and typical additional times.</p>
</div>
</div>"""

    summary_content = f"""{score_gauge}
{score_breakdown}
<table><tr><th>Category</th><th style="text-align:right">Score</th><th style="width:200px">Progress</th><th>Grade</th></tr>
{cat_score_rows}</table>
{effort_summary_html}"""

    # ── All Findings table ──
    all_findings_html = ""
    for f in result.findings:
        af_portal = f'<a href="{_esc(f.portal_link)}" target="_blank" style="color:#7c3aed;font-size:12px;font-weight:600">{_portal_label(f.portal_link)}&nbsp;&#8599;</a>' if f.portal_link else ""
        all_findings_html += f"""<tr>
<td style="font-family:monospace;font-size:12px">{_esc(f.check_id)}</td>
<td style="font-size:12px">{_esc(f.category)}</td>
<td>{sev_badge(f.severity)}</td><td>{status_badge(f.status)}</td>
{'<td style="font-size:12px;white-space:nowrap">' + _esc(f.effort) + '</td>' if show_effort else ''}
<td>{_esc(f.title)}</td>
<td style="font-size:12px;max-width:250px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">{_esc(f.current_state)}</td>
<td>{af_portal}</td>
<td style="font-size:12px;max-width:300px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">{_esc(f.benefits) if f.benefits else ''}</td></tr>"""
    all_findings_content = f"""<div style="overflow-x:auto"><table style="min-width:700px">
<tr><th>Check ID</th><th>Category</th><th>Severity</th><th>Status</th>{'<th>Effort</th>' if show_effort else ''}<th>Title</th><th>Current State</th><th>Portal</th><th>Why It Matters</th></tr>
{all_findings_html}</table></div>"""

    # ── Failed Checks table ──
    failed = [f for f in result.findings if f.status == "FAIL"]
    if failed:
        failed_rows = ""
        for f in failed:
            ref = f'<a href="{_esc(f.reference_url)}" target="_blank" style="color:#3b82f6;font-size:12px">Docs</a>' if f.reference_url else ""
            portal = f'<a href="{_esc(f.portal_link)}" target="_blank" style="color:#7c3aed;font-size:12px;font-weight:600">{_portal_label(f.portal_link)}&nbsp;&#8599;</a>' if f.portal_link else ""
            secret_detail_row = ""
            _fail_cols = 10 if show_effort else 9
            if f.details and "findings" in f.details:
                secret_detail_row = f'<tr><td colspan="{_fail_cols}" style="padding:0 8px 12px">{_render_secret_details_html(f.details)}</td></tr>'
            failed_rows += f"""<tr>
<td style="font-family:monospace;font-size:12px">{_esc(f.check_id)}</td>
<td style="font-size:12px">{_esc(f.category)}</td><td>{sev_badge(f.severity)}</td>
{'<td style="font-size:12px;white-space:nowrap">' + _esc(f.effort) + '</td>' if show_effort else ''}
<td>{_esc(f.title)}</td><td style="font-size:12px">{_esc(f.current_state)}</td>
<td style="font-size:12px">{_esc(f.recommendation)}</td><td>{ref}</td><td>{portal}</td>
<td style="font-size:12px">{_esc(f.benefits) if f.benefits else ''}</td></tr>{secret_detail_row}"""
        failed_content = f"""<div style="overflow-x:auto"><table style="min-width:800px">
<tr><th>Check ID</th><th>Category</th><th>Severity</th>{'<th>Effort</th>' if show_effort else ''}<th>Title</th><th>Current State</th><th>Recommendation</th><th>Ref</th><th>Portal</th><th>Why It Matters</th></tr>
{failed_rows}</table></div>"""
    else:
        failed_content = '<p style="color:#16a34a;font-size:14px;font-weight:600">No failed checks. All checks passed or are warnings.</p>'

    # ── Warnings table (genuine warnings only, API errors excluded) ──
    warns = [f for f in result.findings if f.status == "WARN" and not f.is_api_error]
    if warns:
        warn_rows = ""
        for f in warns:
            ref = f'<a href="{_esc(f.reference_url)}" target="_blank" style="color:#3b82f6;font-size:12px">Docs</a>' if f.reference_url else ""
            portal = f'<a href="{_esc(f.portal_link)}" target="_blank" style="color:#7c3aed;font-size:12px;font-weight:600">{_portal_label(f.portal_link)}&nbsp;&#8599;</a>' if f.portal_link else ""
            warn_rows += f"""<tr>
<td style="font-family:monospace;font-size:12px">{_esc(f.check_id)}</td>
<td style="font-size:12px">{_esc(f.category)}</td><td>{sev_badge(f.severity)}</td>
{'<td style="font-size:12px;white-space:nowrap">' + _esc(f.effort) + '</td>' if show_effort else ''}
<td>{_esc(f.title)}</td><td style="font-size:12px">{_esc(f.current_state)}</td>
<td style="font-size:12px">{_esc(f.recommendation)}</td><td>{ref}</td><td>{portal}</td>
<td style="font-size:12px">{_esc(f.benefits) if f.benefits else ''}</td></tr>"""
        warnings_content = f"""<div style="overflow-x:auto"><table style="min-width:800px">
<tr><th>Check ID</th><th>Category</th><th>Severity</th>{'<th>Effort</th>' if show_effort else ''}<th>Title</th><th>Current State</th><th>Recommendation</th><th>Ref</th><th>Portal</th><th>Why It Matters</th></tr>
{warn_rows}</table></div>"""
    else:
        warnings_content = '<p style="color:#16a34a;font-size:14px;font-weight:600">No warnings found.</p>'

    # ── API Errors table ──
    api_errs = [f for f in result.findings if f.is_api_error]
    if api_errs:
        api_err_rows = ""
        for f in api_errs:
            http_code = f.details.get("http_status", "") if f.details else ""
            http_label = f"HTTP {http_code}" if http_code else "Exception"
            http_badge = f'<span style="background:#7c3aed;color:#fff;padding:2px 8px;border-radius:9999px;font-size:11px;font-weight:600">{_esc(http_label)}</span>'
            justification = f.details.get("justification", "") if f.details else ""
            justification_html = f'<div style="color:#6b7280;font-size:11px;margin-top:4px;font-style:italic">{_esc(justification)}</div>' if justification else ""
            ae_portal = f'<a href="{_esc(f.portal_link)}" target="_blank" style="color:#7c3aed;font-size:12px;font-weight:600">{_portal_label(f.portal_link)}&nbsp;&#8599;</a>' if f.portal_link else ""
            api_err_rows += f"""<tr>
<td style="font-family:monospace;font-size:12px">{_esc(f.check_id)}</td>
<td style="font-size:12px">{_esc(f.category)}</td>
<td>{http_badge}</td>
<td>{_esc(f.title)}</td>
<td style="font-size:12px">{_esc(f.current_state)}{justification_html}</td>
<td style="font-size:12px">{_esc(f.recommendation)}</td>
<td>{ae_portal}</td>
<td style="font-size:12px">{_esc(f.benefits) if f.benefits else ''}</td></tr>"""
        api_errors_content = f"""<p style="color:#7c3aed;font-size:13px;margin-bottom:16px;font-weight:600">{_pl(len(api_errs), 'check')} could not be evaluated due to API failures. They are excluded from the security score. Fix the underlying issues and re-scan.</p>
<div style="overflow-x:auto"><table style="min-width:800px">
<tr><th>Check ID</th><th>Category</th><th>HTTP Status</th><th>Title</th><th>Error Detail</th><th>Recommendation</th><th>Portal</th><th>Why It Matters</th></tr>
{api_err_rows}</table></div>"""
    else:
        ep_err_count = result.endpoint_summary.get("error", 0) if result.endpoint_summary else 0
        if ep_err_count:
            api_errors_content = (
                f'<p style="color:#16a34a;font-size:14px;font-weight:600">All checks produced results.</p>'
                f'<p style="color:#6b7280;font-size:13px;margin-top:4px">'
                f'{ep_err_count} API endpoint{"s" if ep_err_count != 1 else ""} returned HTTP errors '
                f'but checks handled them with fallback logic (see endpoint summary above).</p>'
            )
        else:
            api_errors_content = '<p style="color:#16a34a;font-size:14px;font-weight:600">No API errors. All checks were evaluated successfully.</p>'

    # ── N/A (Not Applicable) table ──
    na_findings = [f for f in result.findings if f.status == "NOT_APPLICABLE" and not f.is_api_error]
    if na_findings:
        na_rows = ""
        for f in na_findings:
            orig_sev = SAT_CHECKS.get(f.check_id, {}).get("severity", "low")
            ref = f'<a href="{_esc(f.reference_url)}" target="_blank" style="color:#3b82f6;font-size:12px">Docs</a>' if f.reference_url else ""
            portal = f'<a href="{_esc(f.portal_link)}" target="_blank" style="color:#7c3aed;font-size:12px;font-weight:600">{_portal_label(f.portal_link)}&nbsp;&#8599;</a>' if f.portal_link else ""
            na_rows += f"""<tr>
<td style="font-family:monospace;font-size:12px">{_esc(f.check_id)}</td>
<td style="font-size:12px">{_esc(f.category)}</td><td>{sev_badge(orig_sev)}</td>
<td>{_esc(f.title)}</td><td style="font-size:12px">{_esc(f.current_state)}</td>
<td style="font-size:12px">{_esc(f.recommendation)}</td><td>{ref}</td><td>{portal}</td>
<td style="font-size:12px">{_esc(f.benefits) if f.benefits else ''}</td></tr>"""
        na_content = f"""<p style="color:#64748b;font-size:13px;margin-bottom:16px">These checks were skipped because the feature or resource is not in use on this workspace. They are excluded from the security score.</p>
<div style="overflow-x:auto"><table style="min-width:800px">
<tr><th>Check ID</th><th>Category</th><th>Severity</th><th>Title</th><th>Reason</th><th>Recommendation</th><th>Ref</th><th>Portal</th><th>Why It Matters</th></tr>
{na_rows}</table></div>"""
    else:
        na_content = '<p style="color:#16a34a;font-size:14px;font-weight:600">No skipped checks. All checks were evaluated.</p>'

    # ── Passed Checks table ──
    passed = [f for f in result.findings if f.status == "PASS"]
    if passed:
        pass_rows = ""
        for f in passed:
            orig_sev = SAT_CHECKS.get(f.check_id, {}).get("severity", "low")
            ref = f'<a href="{_esc(f.reference_url)}" target="_blank" style="color:#3b82f6;font-size:12px">Docs</a>' if f.reference_url else ""
            portal = f'<a href="{_esc(f.portal_link)}" target="_blank" style="color:#7c3aed;font-size:12px;font-weight:600">{_portal_label(f.portal_link)}&nbsp;&#8599;</a>' if f.portal_link else ""
            pass_rows += f"""<tr>
<td style="font-family:monospace;font-size:12px">{_esc(f.check_id)}</td>
<td style="font-size:12px">{_esc(f.category)}</td><td>{sev_badge(orig_sev)}</td>
<td>{_esc(f.title)}</td><td style="font-size:12px">{_esc(f.current_state)}</td>
<td style="font-size:12px">{_esc(f.recommendation)}</td><td>{ref}</td><td>{portal}</td>
<td style="font-size:12px">{_esc(f.benefits) if f.benefits else ''}</td></tr>"""
        passed_content = f"""<p style="color:#16a34a;font-size:13px;margin-bottom:16px;font-weight:600">{len(passed)} check(s) passed successfully.</p>
<div style="overflow-x:auto"><table style="min-width:800px">
<tr><th>Check ID</th><th>Category</th><th>Severity</th><th>Title</th><th>Current State</th><th>Recommendation</th><th>Ref</th><th>Portal</th><th>Why It Matters</th></tr>
{pass_rows}</table></div>"""
    else:
        passed_content = '<p style="color:#ca8a04;font-size:14px;font-weight:600">No checks passed.</p>'

    # ── Per-category finding cards ──
    by_cat: dict[str, list[SATFinding]] = {}
    for f in result.findings:
        by_cat.setdefault(f.category, []).append(f)

    cat_content: dict[str, str] = {}
    for cat, findings in by_cat.items():
        items = ""
        for f in findings:
            rec_block = f"""<div style="background:#eff6ff;border-left:4px solid #3b82f6;padding:10px 14px;border-radius:0 6px 6px 0;margin-top:8px"><strong style="color:#1d4ed8">Recommendation:</strong><br><span style="color:#374151;font-size:13px">{_esc(f.recommendation)}</span></div>""" if f.status != "PASS" else ""
            ref_link = f'<p style="margin:8px 0 0;font-size:12px"><a href="{_esc(f.reference_url)}" style="color:#2563eb">Docs &#8599;</a></p>' if f.reference_url else ""
            portal_link = f'<p style="margin:4px 0 0;font-size:12px"><a href="{_esc(f.portal_link)}" target="_blank" style="color:#7c3aed;font-weight:600">{_portal_label_long(f.portal_link)} &#8599;</a></p>' if f.portal_link else ""
            # Evidence block (amber callout)
            evidence_block = ""
            if show_evidence and f.evidence and f.evidence.get("field") != "current_state":
                ev = f.evidence
                ev_val = _esc(json.dumps(ev["value"], default=str) if not isinstance(ev.get("value"), str) else str(ev["value"]))
                evidence_block = f'''<div style="background:#fef3c7;border-left:4px solid #f59e0b;padding:8px 12px;border-radius:0 6px 6px 0;margin:8px 0;font-size:12px"><strong style="color:#92400e">Evidence:</strong> <code style="background:#fefce8;padding:2px 6px;border-radius:3px">{_esc(str(ev["field"]))}</code> = <code style="background:#fefce8;padding:2px 6px;border-radius:3px">{ev_val}</code> <span style="color:#a16207;font-size:11px;margin-left:8px">({_esc(ev.get("source", ""))})</span></div>'''
            # Secret scan findings: formatted secrets table
            # Scanned items table: always shown when items exist
            # Fallback: raw JSON details
            details_block = ""
            if f.details and "findings" in f.details:
                details_block = _render_secret_details_html(f.details)
            if f.details and "items" in f.details:
                scan_items_html = _render_scan_items_html(f.details)
                details_block += f"""<details style="margin-top:8px"><summary style="cursor:pointer;font-size:12px;color:#6b7280;font-weight:600">Scanned Items</summary>{scan_items_html}</details>"""
            elif include_api_response and f.details:
                details_json = _esc(_details_str(f.details))
                details_block += f"""<details style="margin-top:8px"><summary style="cursor:pointer;font-size:12px;color:#6b7280;font-weight:600">API Response Details</summary><pre style="background:#f3f4f6;border:1px solid #e5e7eb;border-radius:6px;padding:10px;font-size:11px;overflow-x:auto;margin-top:4px;white-space:pre-wrap;word-break:break-all">{details_json}</pre></details>"""
            items += f"""<div class="finding-card" style="background:#f9fafb;border:1px solid #e5e7eb;border-radius:8px;padding:16px;margin-bottom:12px">
<div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:8px">
<div><span style="font-size:11px;color:#6b7280;font-weight:600">{_esc(f.check_id)}</span>
<h4 style="font-size:15px;font-weight:600;color:#111827;margin:4px 0 0">{_esc(f.title)}</h4></div>
<div style="display:flex;gap:6px;flex-shrink:0">{status_badge(f.status)} {sev_badge(f.severity)}</div></div>
<p style="color:#374151;font-size:13px;margin:0 0 8px"><strong>Current:</strong> {_esc(f.current_state)}</p>
{evidence_block}<p style="color:#374151;font-size:13px;margin:0 0 8px">{_esc(f.description)}</p>
{f'<div style="background:#ecfdf5;border-left:4px solid #10b981;padding:10px 14px;border-radius:0 6px 6px 0;margin:8px 0"><strong style="color:#065f46">Why it matters:</strong><br><span style="color:#374151;font-size:13px">' + _esc(f.benefits) + '</span></div>' if f.benefits else ''}
{rec_block}{ref_link}{portal_link}
{details_block}</div>"""
        cat_pass = sum(1 for f in findings if f.status == "PASS")
        cat_fail = sum(1 for f in findings if f.status == "FAIL")
        cat_warn = sum(1 for f in findings if f.status == "WARN")
        cat_na = sum(1 for f in findings if f.status == "NOT_APPLICABLE")
        cat_summary = f"""<div style="display:flex;gap:16px;margin-bottom:16px;font-size:13px">
<span style="color:#16a34a;font-weight:600">{_pl(cat_pass, 'Passed', 'Passed')}</span>
<span style="color:#dc2626;font-weight:600">{_pl(cat_fail, 'Failed', 'Failed')}</span>
<span style="color:#ca8a04;font-weight:600">{_pl(cat_warn, 'Warning')}</span>
<span style="color:#6b7280">{cat_na} N/A</span></div>"""
        cat_content[cat] = cat_summary + items

    # ── Definitions tab ──
    definitions_content = """
<h3 style="font-size:15px;font-weight:700;color:#0f172a;margin:0 0 12px">Finding Statuses</h3>
<table>
<tr><th>Status</th><th>Label</th><th>Definition</th></tr>
<tr><td><span style="background:#16a34a;color:#fff;padding:2px 8px;border-radius:9999px;font-size:11px;font-weight:600">PASS</span></td><td>Compliant</td><td>The check ran successfully and confirmed the security control is in place. No action needed.</td></tr>
<tr><td><span style="background:#dc2626;color:#fff;padding:2px 8px;border-radius:9999px;font-size:11px;font-weight:600">FAIL</span></td><td>Action Required</td><td>The check found a concrete security gap &mdash; confirmed bad configuration that must be fixed.</td></tr>
<tr><td><span style="background:#ca8a04;color:#fff;padding:2px 8px;border-radius:9999px;font-size:11px;font-weight:600">WARN</span></td><td>Review Needed</td><td>The check returned a borderline result that needs manual investigation. Genuine security warnings only (API errors are separated).</td></tr>
<tr><td><span style="background:#6b7280;color:#fff;padding:2px 8px;border-radius:9999px;font-size:11px;font-weight:600">N/A</span></td><td>Skipped</td><td>The feature being checked is not in use on this workspace (e.g. no Delta Sharing recipients, no MLflow models, Unity Catalog not enabled). Excluded from score.</td></tr>
<tr><td><span style="background:#7c3aed;color:#fff;padding:2px 8px;border-radius:9999px;font-size:11px;font-weight:600">API ERROR</span></td><td>Could Not Evaluate</td><td>The check could not run due to an API failure (HTTP 400/401/403/404, timeout, or connection error). Excluded from score. Fix the underlying issue and re-scan.</td></tr>
</table>
<h3 style="font-size:15px;font-weight:700;color:#0f172a;margin:24px 0 12px">Severity Levels</h3>
<table>
<tr><th>Severity</th><th>Weight</th><th>Definition</th></tr>
<tr><td><span style="background:#dc2626;color:#fff;padding:2px 8px;border-radius:9999px;font-size:11px;font-weight:600">CRITICAL</span></td><td>10</td><td>Immediate security risk that could lead to data breach, unauthorized access, or compliance violation. Must be remediated immediately.</td></tr>
<tr><td><span style="background:#ea580c;color:#fff;padding:2px 8px;border-radius:9999px;font-size:11px;font-weight:600">HIGH</span></td><td>7</td><td>Significant security weakness that could be exploited. Should be remediated within days.</td></tr>
<tr><td><span style="background:#ca8a04;color:#fff;padding:2px 8px;border-radius:9999px;font-size:11px;font-weight:600">MEDIUM</span></td><td>4</td><td>Moderate risk that weakens overall security posture. Should be remediated within the current sprint/cycle.</td></tr>
<tr><td><span style="background:#2563eb;color:#fff;padding:2px 8px;border-radius:9999px;font-size:11px;font-weight:600">LOW</span></td><td>2</td><td>Minor improvement opportunity or best-practice deviation. Plan remediation for upcoming cycles.</td></tr>
</table>
<h3 style="font-size:15px;font-weight:700;color:#0f172a;margin:24px 0 12px">Grade Definitions</h3>
<table>
<tr><th>Grade</th><th>Score Range</th><th>Definition</th></tr>
<tr><td style="color:#16a34a;font-weight:700">Good</td><td>80&ndash;100</td><td>Strong security posture. Address remaining findings as maintenance items.</td></tr>
<tr><td style="color:#ca8a04;font-weight:700">Needs Improvement</td><td>60&ndash;79</td><td>Gaps exist that weaken security posture. Prioritize High and Critical findings.</td></tr>
<tr><td style="color:#dc2626;font-weight:700">Critical</td><td>0&ndash;59</td><td>Significant security risks are present. Immediate remediation is required.</td></tr>
</table>
<h3 style="font-size:15px;font-weight:700;color:#0f172a;margin:24px 0 12px">Scoring Formula</h3>
<table>
<tr><th>Component</th><th>Detail</th></tr>
<tr><td>Weight per check</td><td>Critical=10, High=7, Medium=4, Low=2</td></tr>
<tr><td>FAIL penalty</td><td>Full weight &mdash; Critical=10, High=7, Medium=4, Low=2</td></tr>
<tr><td>WARN penalty</td><td>Half weight &mdash; Critical=5, High=3.5, Medium=2, Low=1</td></tr>
<tr><td>PASS penalty</td><td>0 (no penalty)</td></tr>
<tr><td>Excluded</td><td>NOT_APPLICABLE and API Error findings are excluded from weight totals</td></tr>
<tr><td>Formula</td><td>Score = (1 &minus; penalty_sum / total_weights) &times; 100<br><em>penalty_sum = &sum;(FAIL &times; full_weight) + &sum;(WARN &times; half_weight)</em></td></tr>
<tr><td>Interpretation</td><td>A FAIL on a Critical check costs 10 pts; a WARN on a Critical check costs 5 pts. A FAIL on Low costs 2 pts; a WARN on Low costs 1 pt.</td></tr>
</table>
<h3 style="font-size:15px;font-weight:700;color:#0f172a;margin:24px 0 12px">Category Definitions</h3>
<table>
<tr><th>Category</th><th>Definition</th></tr>
""" + "".join(f'<tr><td>{_esc(cat)}</td><td>{_esc(defn)}</td></tr>' for cat, defn in CATEGORY_DEFINITIONS) + """
</table>
<h3 style="font-size:15px;font-weight:700;color:#0f172a;margin:24px 0 12px">HTTP Error Codes in Findings</h3>
<table>
<tr><th>Code</th><th>Meaning</th><th>What to do</th></tr>
<tr><td>401</td><td>Unauthorized &mdash; token is invalid, expired, or revoked.</td><td>Regenerate a new PAT token in Databricks &rarr; User Settings &rarr; Access Tokens, or re-login via Azure to refresh.</td></tr>
<tr><td>403</td><td>Permission Denied &mdash; token authenticated but lacks the required admin role for this endpoint.</td><td>Use a Workspace Admin PAT token. Admin-only endpoints: token-management, ip-access-lists, workspace-conf, Settings API.</td></tr>
<tr><td>404</td><td>Not Found &mdash; the API endpoint does not exist on this workspace (feature not enabled or requires Premium pricing tier).</td><td>Check the Azure Databricks Account Console for account-level settings. Verify your workspace pricing tier (Premium required for many features).</td></tr>
<tr><td>400</td><td>Bad Request &mdash; feature is not configured or workspace does not support that configuration key.</td><td>Review raw details. This usually means the feature is managed differently (e.g. Unity Catalog instead of Table ACLs).</td></tr>
</table>"""
    if show_effort:
        definitions_content += """
<h3 style="font-size:15px;font-weight:700;color:#0f172a;margin:24px 0 12px">Remediation Effort Levels</h3>
<table>
<tr><th>Level</th><th>Time Range</th><th>What&rsquo;s Included</th><th>Examples</th></tr>
<tr><td style="color:#16a34a;font-weight:700">Quick Fix</td><td>5&ndash;15 min</td><td>Single configuration toggle or setting change in the admin console. No cross-team coordination or testing required.</td><td>Enable a workspace flag, toggle an admin setting, update a single policy value</td></tr>
<tr><td style="color:#2563eb;font-weight:700">Moderate</td><td>1&ndash;4 hrs</td><td>Multi-step configuration, policy creation, or changes that require testing and validation. May involve creating new resources, updating IaC code, or running build/deploy pipelines.</td><td>Create cluster policies, configure IP access lists, set up secret scopes, update Terraform modules, run CI/CD pipelines</td></tr>
<tr><td style="color:#ca8a04;font-weight:700">Significant</td><td>1&ndash;3 days</td><td>Architecture changes requiring cross-team coordination, downtime planning, IaC refactoring, and phased rollout. Includes prerequisite work like updating Terraform providers, modifying ARM/Bicep templates, and change management approvals.</td><td>Enable Unity Catalog, configure VNet peering, migrate to private endpoints, implement SSO/SCIM, refactor Terraform state</td></tr>
<tr><td style="color:#dc2626;font-weight:700">Project</td><td>1+ weeks</td><td>Major infrastructure migration or org-wide policy rollout. Requires project planning, stakeholder approval, IaC redesign, multi-phase implementation, and extensive testing across environments.</td><td>Full workspace migration, enterprise-wide compliance overhaul, network architecture redesign, Terraform module rewrite</td></tr>
</table>
<h3 style="font-size:15px;font-weight:700;color:#0f172a;margin:24px 0 12px">Common Prerequisites (not included in estimates)</h3>
<p style="font-size:12px;color:#64748b;margin:0 0 10px">The effort estimates above reflect the <strong>core remediation work</strong>. Depending on your organization, you may also need to account for these additional prerequisites:</p>
<table>
<tr><th>Prerequisite</th><th>When Needed</th><th>Typical Additional Time</th></tr>
<tr><td>Terraform / IaC Updates</td><td>When infrastructure is managed via Terraform, Pulumi, ARM/Bicep, or other IaC tools. Changes must be codified, reviewed, and applied through the IaC pipeline rather than the UI.</td><td>+30 min &ndash; 2 hrs per change</td></tr>
<tr><td>CI/CD Pipeline Runs</td><td>When changes must flow through build, test, and deploy pipelines before reaching production.</td><td>+15 min &ndash; 1 hr per deployment</td></tr>
<tr><td>Change Management / CAB Approval</td><td>When your organization requires change tickets, CAB review, or ITSM approvals before production changes.</td><td>+1 &ndash; 5 business days</td></tr>
<tr><td>Non-prod Testing</td><td>When changes must be validated in dev/staging environments before production rollout.</td><td>+1 &ndash; 4 hrs per environment</td></tr>
<tr><td>Provider / Module Upgrades</td><td>When the Terraform azurerm/databricks provider or shared modules must be upgraded to support the required resource type or argument.</td><td>+1 &ndash; 4 hrs</td></tr>
<tr><td>Security Review</td><td>When network, identity, or security teams must review and approve the change before implementation.</td><td>+1 &ndash; 3 business days</td></tr>
<tr><td>Documentation Updates</td><td>When internal runbooks, architecture diagrams, or compliance documentation must be updated to reflect the change.</td><td>+30 min &ndash; 2 hrs</td></tr>
</table>
<h3 style="font-size:15px;font-weight:700;color:#0f172a;margin:24px 0 12px">How Effort Is Estimated</h3>
<table>
<tr><th>Factor</th><th>Description</th></tr>
<tr><td>Configuration Steps</td><td>Number of settings, APIs, or UI steps required to remediate the finding.</td></tr>
<tr><td>Access Requirements</td><td>Whether changes need admin console access, API/IaC changes, or Azure Portal configuration.</td></tr>
<tr><td>Testing &amp; Validation</td><td>Whether the fix needs functional testing, security validation, or user acceptance testing.</td></tr>
<tr><td>Cross-team Coordination</td><td>Whether network, security, identity, or platform teams need to be involved.</td></tr>
<tr><td>Blast Radius</td><td>Whether the change affects a single workspace, multiple workspaces, or the entire organization.</td></tr>
</table>
<p style="font-size:12px;color:#64748b;margin-top:12px;font-style:italic">Note: Effort estimates reflect core remediation work and are approximate. Actual time will vary based on your organization&rsquo;s change management processes, IaC maturity, CI/CD pipeline complexity, and team familiarity with Databricks administration. Add prerequisite time from the table above as applicable.</p>"""

    # ── Build tabs list ──
    CAT_ICONS = {
        "Identity & Access": "user-check", "Network Security": "network",
        "Data Protection": "database", "Compute Security": "server",
        "SQL Warehouses": "warehouse", "Secrets & Credentials": "key",
        "Audit & Logging": "eye", "Governance": "landmark",
        "Informational": "info", "Operations": "activity",
        "Feature Adoption": "sparkles", "Secret Scanning": "search",
        "AI / ML Governance": "brain", "Operations": "activity",
        "Performance": "zap", "Cost Optimization": "dollar-sign",
        "Reliability": "shield", "Data Architecture": "layers",
        "Ops Excellence": "settings", "Governance Data Quality": "bar-chart",
        "Account Governance": "users",
        "Advanced Governance": "landmark", "Advanced Performance": "zap",
        "Data Residency": "map-pin", "Table Optimization": "gauge",
        "Data Quality": "check-square",
    }
    # Preferred display order for category tabs
    _CAT_ORDER = [
        "Identity & Access", "Network Security", "Data Protection",
        "Compute Security", "SQL Warehouses", "Secrets & Credentials",
        "Audit & Logging", "Governance", "AI / ML Governance",
        "Informational", "Secret Scanning", "Operations",
        "Performance", "Cost Optimization", "Reliability",
        "Data Architecture", "Ops Excellence", "Governance Data Quality",
        "Feature Adoption", "Account Governance",
        "Advanced Governance", "Advanced Performance", "Data Residency",
        "Table Optimization", "Data Quality",
    ]
    tabs: list[tuple[str, str, str, str]] = [
        ("summary", "shield", "Summary", summary_content),
        ("all-findings", "list", "All Findings", all_findings_content),
        ("failed-checks", "x-circle", f"{_pl(result.failed, 'Failed Check')}", failed_content),
        ("warnings", "alert-triangle", f"{_pl(result.warnings, 'Warning')}", warnings_content),
        ("api-errors", "plug-zap", f"{_pl(result.api_errors, 'API Error')}", api_errors_content),
        ("na-checks", "minus-circle", f"{_pl(result.not_applicable, 'N/A', 'N/As')}", na_content),
        ("passed-checks", "check-circle", f"{_pl(result.passed, 'Passed', 'Passed')}", passed_content),
    ]
    # ── Prioritised Recommendations tab ──
    prio_items = _build_prioritised_recommendations(result.findings)
    if prio_items:
        _PRIO_COLORS = {"P1": "#dc2626", "P2": "#ea580c", "P3": "#ca8a04", "P4": "#6b7280"}
        _PRIO_BG = {"P1": "#fef2f2", "P2": "#fff7ed", "P3": "#fefce8", "P4": "#f8fafc"}
        # Priority distribution summary
        _prio_counts: dict[str, int] = {}
        for item in prio_items:
            prefix = item["priority_label"][:2]
            _prio_counts[prefix] = _prio_counts.get(prefix, 0) + 1
        prio_dist = ""
        for p in ["P1", "P2", "P3", "P4"]:
            cnt = _prio_counts.get(p, 0)
            if cnt == 0:
                continue
            prio_dist += f'<div style="background:{_PRIO_BG[p]};border:1px solid {_PRIO_COLORS[p]}33;border-radius:8px;padding:12px 18px;text-align:center"><div style="font-size:24px;font-weight:700;color:{_PRIO_COLORS[p]}">{cnt}</div><div style="font-size:11px;color:#64748b;text-transform:uppercase">{p}</div></div>'
        prio_rows_html = ""
        _cost_total_low = 0
        _cost_total_high = 0
        _cost_count = 0
        for item in prio_items:
            prefix = item["priority_label"][:2]
            pc = _PRIO_COLORS.get(prefix, "#6b7280")
            prio_badge = f'<span style="background:{pc};color:#fff;padding:2px 8px;border-radius:9999px;font-size:11px;font-weight:600">{_esc(item["priority_label"])}</span>'
            ref = f'<a href="{_esc(item["reference_url"])}" target="_blank" style="color:#3b82f6;font-size:12px">Docs</a>' if item["reference_url"] else ""
            portal = f'<a href="{_esc(item["portal_link"])}" target="_blank" style="color:#7c3aed;font-size:12px;font-weight:600">{_portal_label(item["portal_link"])}&nbsp;&#8599;</a>' if item["portal_link"] else ""
            cost_td = ""
            if show_cost:
                if item["cost_low"]:
                    cost_td = f'<td style="font-size:12px;text-align:right;white-space:nowrap;color:#b45309">${item["cost_low"]:,} &ndash; ${item["cost_high"]:,}</td>'
                    _cost_total_low += item["cost_low"]
                    _cost_total_high += item["cost_high"]
                    _cost_count += 1
                else:
                    cost_td = '<td style="font-size:11px;color:#94a3b8;text-align:center">&mdash;</td>'
            prio_rows_html += f"""<tr>
<td>{prio_badge}</td>
<td style="font-weight:700;color:{pc};text-align:center">{item['priority_score']}</td>
<td style="font-family:monospace;font-size:12px">{_esc(item['check_id'])}</td>
<td style="font-size:12px">{_esc(item['category'])}</td>
<td>{sev_badge(item['severity'])}</td><td>{status_badge(item['status'])}</td>
<td style="font-size:12px;white-space:nowrap">{_esc(item['effort'])}</td>
{cost_td}
<td>{_esc(item['title'])}</td>
<td style="font-size:12px">{_esc(item['recommendation'])}</td>
<td>{ref}</td><td>{portal}</td>
<td style="font-size:12px">{_esc(item['benefits']) if item['benefits'] else ''}</td></tr>"""
        # Cost summary card
        _cost_card = ""
        if show_cost and _cost_count:
            _cost_card = f'<div style="background:#fffbeb;border:1px solid #fbbf2433;border-radius:8px;padding:12px 18px;text-align:center"><div style="font-size:20px;font-weight:700;color:#b45309">${_cost_total_low:,} &ndash; ${_cost_total_high:,}</div><div style="font-size:11px;color:#64748b;text-transform:uppercase">Est. Monthly Cost</div></div>'
        _cost_explain = ""
        if show_cost:
            _cost_explain = """<br><strong>Est. Cost</strong> = estimated monthly cloud operational cost of the misconfiguration (per workspace).
<em>Cost figures are illustrative examples only &mdash; actual costs vary with usage, region, and pricing tier.</em>"""
        _cost_th = '<th>Est. Cost ($/mo)</th>' if show_cost else ''
        prio_content = f"""<p style="font-size:13px;color:#475569;margin-bottom:16px">
{len(prio_items)} actionable finding{"s" if len(prio_items) != 1 else ""} ranked by <strong>Priority Score</strong>
(severity weight &times; effort multiplier). High-severity quick fixes appear first.
</p>
<div style="display:flex;gap:16px;margin-bottom:20px;flex-wrap:wrap">{prio_dist}{_cost_card}</div>
<div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:14px 18px;margin-bottom:18px;font-size:12px;color:#64748b">
<strong>How it works:</strong> Priority Score = Severity Weight (Critical=10, High=7, Medium=4, Low=2)
&times; Effort Multiplier (Quick Fix=4&times;, Moderate=3&times;, Significant=2&times;, Project=1&times;).
Higher score = fix first.
<br><strong>P1</strong> &ge;28 &middot; <strong>P2</strong> 16&ndash;27 &middot; <strong>P3</strong> 7&ndash;15 &middot; <strong>P4</strong> &lt;7
{_cost_explain}
</div>
<div style="overflow-x:auto"><table style="min-width:{'1100' if show_cost else '1000'}px">
<tr><th>Priority</th><th>Score</th><th>Check ID</th><th>Category</th><th>Severity</th><th>Status</th><th>Effort</th>{_cost_th}<th>Title</th><th>Recommendation</th><th>Ref</th><th>Portal</th><th>Why It Matters</th></tr>
{prio_rows_html}</table></div>"""
        tabs.append(("prioritised", "arrow-up-circle", f"Prioritised ({len(prio_items)})", prio_content))
    # Add category tabs dynamically — ordered list first, then any extras
    _seen_cats: set[str] = set()
    for cat in _CAT_ORDER:
        if cat in cat_content:
            _seen_cats.add(cat)
            tid = cat.lower().replace(" & ", "-").replace("/", "-").replace(" ", "-")
            tabs.append((tid, CAT_ICONS.get(cat, "folder"), cat, cat_content[cat]))
    for cat in sorted(cat_content.keys()):
        if cat not in _seen_cats:
            tid = cat.lower().replace(" & ", "-").replace("/", "-").replace(" ", "-")
            tabs.append((tid, CAT_ICONS.get(cat, "folder"), cat, cat_content[cat]))

    # ── API Endpoints tab ──
    if result.endpoint_summary and result.endpoint_summary.get("endpoints"):
        ep = result.endpoint_summary
        ep_rows = ""
        for e in ep["endpoints"]:
            if e["status"] == "items":
                icon = '<span style="color:#16a34a;font-weight:700">&#10003;</span>'
                count_text = f'{e["items_count"]} item{"s" if e["items_count"] != 1 else ""}'
                count_style = "color:#16a34a;font-weight:600"
            elif e["status"] == "config":
                icon = '<span style="color:#ca8a04;font-weight:700">&#9881;</span>'
                count_text = "config/settings"
                count_style = "color:#ca8a04;font-weight:600"
            elif e["status"] == "error":
                icon = '<span style="color:#dc2626;font-weight:700">&#10007;</span>'
                count_text = f'HTTP {e["error_code"]}'
                count_style = "color:#dc2626;font-weight:600"
            else:
                icon = '<span style="color:#94a3b8">&#9675;</span>'
                count_text = "0 items"
                count_style = "color:#94a3b8"
            ep_rows += f'<tr><td>{icon}</td><td style="font-family:monospace;font-size:12px">{_esc(e["endpoint"])}</td><td style="{count_style}">{count_text}</td></tr>'
        ep_summary_content = f"""<div style="display:flex;gap:24px;margin-bottom:20px;font-size:14px;flex-wrap:wrap">
<div style="background:#f0fdf4;border:1px solid #bbf7d0;border-radius:8px;padding:12px 18px;text-align:center"><div style="font-size:24px;font-weight:700;color:#16a34a">{ep['with_items']}</div><div style="font-size:11px;color:#64748b;text-transform:uppercase">With Items</div></div>
<div style="background:#fefce8;border:1px solid #fef08a;border-radius:8px;padding:12px 18px;text-align:center"><div style="font-size:24px;font-weight:700;color:#ca8a04">{ep['config']}</div><div style="font-size:11px;color:#64748b;text-transform:uppercase">Config/Settings</div></div>
<div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:12px 18px;text-align:center"><div style="font-size:24px;font-weight:700;color:#94a3b8">{ep['empty']}</div><div style="font-size:11px;color:#64748b;text-transform:uppercase">Empty</div></div>
<div style="background:#fef2f2;border:1px solid #fca5a5;border-radius:8px;padding:12px 18px;text-align:center"><div style="font-size:24px;font-weight:700;color:#dc2626">{ep.get('error', 0)}</div><div style="font-size:11px;color:#64748b;text-transform:uppercase">Errors</div></div>
<div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;padding:12px 18px;text-align:center"><div style="font-size:24px;font-weight:700;color:#3b82f6">{ep['total']}</div><div style="font-size:11px;color:#64748b;text-transform:uppercase">Total</div></div>
</div>
<div style="overflow-x:auto"><table style="min-width:500px">
<tr><th style="width:30px"></th><th>Endpoint</th><th>Result</th></tr>
{ep_rows}</table></div>"""
        tabs.append(("api-endpoints", "globe", "API Endpoints", ep_summary_content))

    # ── All Checks Reference tab ──
    _cat_order_ref: list[str] = []
    _cat_checks_ref: dict[str, list[str]] = {}
    for cid, cdata in SAT_CHECKS.items():
        cat = cdata.get("category", "Other")
        if cat not in _cat_checks_ref:
            _cat_order_ref.append(cat)
            _cat_checks_ref[cat] = []
        _cat_checks_ref[cat].append(cid)
    ref_rows = ""
    _sev_sort = {"critical": 0, "high": 1, "medium": 2, "low": 3}
    for cat in _cat_order_ref:
        sorted_ids = sorted(_cat_checks_ref[cat], key=lambda c: (_sev_sort.get(SAT_CHECKS[c].get("severity", "low"), 3), c))
        for cid in sorted_ids:
            ck = SAT_CHECKS[cid]
            ref_url = ck.get("reference_url", "")
            ref_link = f'<a href="{_esc(ref_url)}" target="_blank" style="color:#3b82f6;text-decoration:none;font-size:11px">Docs</a>' if ref_url else ""
            _effort_td = f'<td style="font-size:12px;white-space:nowrap">{_esc(_get_effort(cid))}</td>' if show_effort else ''
            ref_rows += f'<tr><td style="font-family:monospace;font-size:12px;white-space:nowrap">{_esc(cid)}</td><td>{_esc(cat)}</td><td>{sev_badge(ck.get("severity", "low"))}</td>{_effort_td}<td><strong>{_esc(ck.get("title", ""))}</strong></td><td style="font-size:12px">{_esc(ck.get("description", ""))}</td><td style="font-size:12px">{_esc(ck.get("recommendation", ""))}</td><td>{ref_link}</td></tr>\n'
    _sev_counts = {}
    for ck in SAT_CHECKS.values():
        s = ck.get("severity", "low")
        _sev_counts[s] = _sev_counts.get(s, 0) + 1
    sev_chips = " ".join(f'{sev_badge(s)} <span style="font-size:13px;margin-right:12px">{_sev_counts.get(s,0)}</span>' for s in ["critical", "high", "medium", "low"])
    checks_ref_content = f"""<div style="margin-bottom:16px;display:flex;gap:16px;align-items:center;flex-wrap:wrap">
<div style="font-size:15px;font-weight:600;color:#1e293b">Total: {len(SAT_CHECKS)} checks</div>
{sev_chips}
</div>
<div style="overflow-x:auto"><table style="min-width:900px">
<tr><th>Check ID</th><th>Category</th><th>Severity</th>{'<th>Effort</th>' if show_effort else ''}<th>Title</th><th>Description</th><th>Recommendation</th><th>Docs</th></tr>
{ref_rows}</table></div>"""
    tabs.append(("checks-reference", "list-checks", "All Checks ({})".format(len(SAT_CHECKS)), checks_ref_content))

    tabs.append(("definitions", "book-open", "Definitions", definitions_content))

    # ── Build tab buttons & panels ──
    tab_buttons = ""
    for i, (tid, icon_name, label, _) in enumerate(tabs):
        active = " active" if i == 0 else ""
        tab_buttons += f'<button class="tab-btn{active}" data-tab="{tid}">{_icon(icon_name, 15)} {_esc(label)}</button>\n'

    tab_panels = ""
    for i, (tid, icon_name, label, content) in enumerate(tabs):
        display = "block" if i == 0 else "none"
        tab_panels += f'<div class="tab-panel" id="panel-{tid}" style="display:{display}">\n  <div class="card">\n    <h2>{_icon(icon_name, 18)} {_esc(label)}</h2>\n    {content}\n  </div>\n</div>\n'

    # ── KPI cards ──
    score_color = "#16a34a" if result.overall_score >= 80 else ("#ca8a04" if result.overall_score >= 60 else "#dc2626")
    _sub = '<div style="font-size:12px;color:#64748b;margin-top:4px">'
    na_kpi = f"""<div class="kpi"><div class="label">N/A</div><div class="value" style="color:#64748b">{result.not_applicable}</div>{_sub}excluded from score</div></div>""" if result.not_applicable else ""
    api_err_kpi = f"""<div class="kpi"><div class="label">{"API Error" if result.api_errors == 1 else "API Errors"}</div><div class="value" style="color:#7c3aed">{result.api_errors}</div>{_sub}excluded from score</div></div>""" if result.api_errors else ""
    kpis_html = f"""<div class="kpis">
<div class="kpi"><div class="label">Overall Score</div><div class="value" style="color:{score_color}">{result.overall_score}/100</div>{_sub}{_esc(grade)}</div></div>
<div class="kpi"><div class="label">Total Checks</div><div class="value" style="color:#374151">{result.total_checks}</div>{_sub}{len(_applicable)} scored</div></div>
<div class="kpi"><div class="label">Passed</div><div class="value" style="color:#16a34a">{result.passed}</div>{_sub}no penalty</div></div>
<div class="kpi"><div class="label">Failed</div><div class="value" style="color:#dc2626">{result.failed}</div>{_sub}full penalty</div></div>
<div class="kpi"><div class="label">{"Warning" if result.warnings == 1 else "Warnings"}</div><div class="value" style="color:#ca8a04">{result.warnings}</div>{_sub}half penalty</div></div>
{na_kpi}{api_err_kpi}</div>"""

    # ── Full HTML ──
    ws_name_line = f" \u2014 {_esc(result.workspace_name)}" if result.workspace_name else ""
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Databricks SAT Report{ws_name_line}</title>
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
          background: #f1f5f9; color: #1e293b; line-height: 1.5; -webkit-font-smoothing: antialiased; }}

  /* ── Header ── */
  .header {{ background: linear-gradient(135deg, #0f172a 0%, #1e3a5f 50%, #1e40af 100%);
             color: white; padding: 28px 36px; position: relative; overflow: hidden; }}
  .header::before {{ content: ''; position: absolute; top: -50%; right: -10%; width: 400px; height: 400px;
    background: radial-gradient(circle, rgba(59,130,246,.15) 0%, transparent 70%); pointer-events: none; }}
  .header h1 {{ font-size: 24px; font-weight: 700; margin-bottom: 6px; letter-spacing: -.01em;
                position: relative; z-index: 1; }}
  .header .sub {{ opacity: .75; font-size: 13px; position: relative; z-index: 1; }}
  .header .sub a {{ color: #93c5fd; transition: color .2s; }}
  .header .sub a:hover {{ color: #bfdbfe; }}

  /* ── Layout ── */
  .layout {{ display: flex; min-height: calc(100vh - 88px); }}

  /* ── Sidebar ── */
  .sidebar {{ width: 220px; min-width: 220px; background: #0f172a;
              overflow-y: auto; padding: 16px 0; scrollbar-width: thin; scrollbar-color: #334155 transparent; }}
  .sidebar::-webkit-scrollbar {{ width: 6px; }}
  .sidebar::-webkit-scrollbar-track {{ background: transparent; }}
  .sidebar::-webkit-scrollbar-thumb {{ background: #334155; border-radius: 3px; }}
  .tab-btn {{ display: flex; align-items: center; width: 100%; padding: 11px 20px;
              border: none; background: none; color: #94a3b8; font-size: 13px;
              cursor: pointer; text-align: left; transition: all .2s ease; gap: 6px;
              border-left: 3px solid transparent; position: relative; }}
  .tab-btn:hover {{ background: rgba(30,58,95,.6); color: #e2e8f0; border-left-color: #475569; }}
  .tab-btn.active {{ background: linear-gradient(90deg, rgba(59,130,246,.15) 0%, transparent 100%);
                     color: #ffffff; font-weight: 600; border-left-color: #3b82f6; }}

  /* ── Main ── */
  .main {{ flex: 1; overflow-y: auto; padding: 28px 36px; }}

  /* ── KPI Cards ── */
  .kpis {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); gap: 16px; margin-bottom: 28px; }}
  .kpi {{ background: white; border: 1px solid #e2e8f0; border-radius: 12px; padding: 20px 24px;
          box-shadow: 0 1px 3px rgba(0,0,0,.04); transition: transform .2s ease, box-shadow .2s ease; }}
  .kpi:hover {{ transform: translateY(-2px); box-shadow: 0 4px 12px rgba(0,0,0,.08); }}
  .kpi .value {{ font-size: 28px; font-weight: 700; margin: 6px 0; }}
  .kpi .label {{ font-size: 11px; color: #64748b; text-transform: uppercase; letter-spacing: .06em; font-weight: 600; }}

  /* ── Content Card ── */
  .card {{ background: white; border: 1px solid #e2e8f0; border-radius: 12px; padding: 28px;
           box-shadow: 0 1px 3px rgba(0,0,0,.04); }}
  h2 {{ font-size: 17px; font-weight: 700; margin-bottom: 20px; color: #0f172a; letter-spacing: -.01em; }}

  /* ── Tables ── */
  table {{ width: 100%; border-collapse: separate; border-spacing: 0; font-size: 13px; }}
  th {{ text-align: left; padding: 10px 12px; background: #f8fafc;
        font-size: 11px; text-transform: uppercase; letter-spacing: .05em; color: #475569; font-weight: 600;
        border-bottom: 2px solid #e2e8f0; position: sticky; top: 0; z-index: 1; }}
  td {{ padding: 10px 12px; border-bottom: 1px solid #f1f5f9; transition: background .15s; }}
  tbody tr:nth-child(even) td {{ background: #fafbfc; }}
  tr:hover td {{ background: #eff6ff; }}

  /* ── Finding Cards ── */
  .finding-card {{ background: #fff; border: 1px solid #e2e8f0; border-radius: 10px; padding: 18px 20px;
                   margin-bottom: 14px; box-shadow: 0 1px 3px rgba(0,0,0,.04);
                   transition: box-shadow .2s ease, border-color .2s ease; }}
  .finding-card:hover {{ box-shadow: 0 4px 16px rgba(0,0,0,.08); border-color: #cbd5e1; }}

  /* ── Details / Accordion ── */
  details {{ transition: all .2s ease; }}
  details[open] summary {{ color: #1d4ed8; }}
  details summary {{ padding: 6px 0; transition: color .15s; }}
  details summary:hover {{ color: #2563eb; }}
  details p {{ margin: 6px 0; line-height: 1.6; }}

  /* ── Footer ── */
  .footer {{ text-align: center; color: #94a3b8; font-size: 11px; padding: 20px; margin-top: 12px; }}

  /* ── Badges ── */
  .tip {{ position: relative; cursor: help; transition: transform .15s; }}
  .tip:hover {{ transform: scale(1.05); }}
  .tip .tip-text {{ visibility: hidden; opacity: 0; position: absolute; bottom: calc(100% + 8px);
    left: 50%; transform: translateX(-50%); background: #1e293b; color: #f1f5f9;
    padding: 8px 12px; border-radius: 8px; font-size: 11px; font-weight: 400;
    white-space: nowrap; z-index: 100; pointer-events: none;
    transition: opacity .2s, transform .2s; text-transform: none; letter-spacing: normal;
    box-shadow: 0 4px 12px rgba(0,0,0,.15); }}
  .tip .tip-text::after {{ content: ''; position: absolute; top: 100%; left: 50%;
    margin-left: -5px; border: 5px solid transparent; border-top-color: #1e293b; }}
  .tip:hover .tip-text {{ visibility: visible; opacity: 1; }}

  /* ── Search ── */
  .search-bar {{ position: relative; margin-bottom: 20px; }}
  .search-bar input {{ width: 100%; padding: 12px 18px 12px 44px; border: 1px solid #e2e8f0;
    border-radius: 10px; font-size: 14px; outline: none; transition: all .2s ease;
    background: #fff url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='16' height='16' fill='none' stroke='%2394a3b8' stroke-width='2' stroke-linecap='round'%3E%3Ccircle cx='7' cy='7' r='5'/%3E%3Cline x1='11' y1='11' x2='15' y2='15'/%3E%3C/svg%3E") 14px center no-repeat; }}
  .search-bar input:focus {{ border-color: #3b82f6; box-shadow: 0 0 0 4px rgba(59,130,246,.12); }}
  .search-count {{ position: absolute; right: 14px; top: 50%; transform: translateY(-50%);
    font-size: 12px; color: #64748b; font-weight: 500; }}
  .search-badge {{ display: none; margin-left: auto; padding: 2px 8px; border-radius: 10px;
    font-size: 11px; font-weight: 600; background: #3b82f6; color: #fff; }}

  /* ── Responsive ── */
  @media (max-width: 900px) {{
    .layout {{ flex-direction: column; min-height: auto; }}
    .sidebar {{ width: 100%; min-width: unset; display: flex; flex-wrap: wrap;
                padding: 10px; gap: 6px; }}
    .tab-btn {{ width: auto; padding: 8px 14px; font-size: 12px;
                border-radius: 8px; border-left: none !important; }}
    .tab-btn.active {{ border-left: none !important; background: #1e3a5f; border-radius: 8px; }}
    .main {{ padding: 16px; }}
    .kpis {{ grid-template-columns: repeat(2, 1fr); }}
    .card {{ padding: 16px; }}
  }}

  /* ── Print ── */
  @media print {{
    body {{ background: white; }}
    .sidebar {{ display: none !important; }}
    .layout {{ display: block; }}
    .tab-panel {{ display: block !important; page-break-inside: avoid; margin-bottom: 24px; }}
    .main {{ padding: 0; overflow: visible; }}
    .header {{ background: #0f172a !important; -webkit-print-color-adjust: exact; print-color-adjust: exact; padding: 16px 24px; }}
    .kpi {{ box-shadow: none; border: 1px solid #ccc; }}
    .card {{ box-shadow: none; border: 1px solid #ccc; }}
    .search-bar {{ display: none; }}
    th {{ background: #f1f5f9 !important; -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
    .finding-card {{ box-shadow: none; break-inside: avoid; }}
  }}

  /* ── Scrollbar (main) ── */
  .main::-webkit-scrollbar {{ width: 8px; }}
  .main::-webkit-scrollbar-track {{ background: transparent; }}
  .main::-webkit-scrollbar-thumb {{ background: #cbd5e1; border-radius: 4px; }}
  .main::-webkit-scrollbar-thumb:hover {{ background: #94a3b8; }}

  /* ── Smooth panel transitions ── */
  .tab-panel {{ animation: fadeIn .25s ease; }}
  @keyframes fadeIn {{ from {{ opacity: 0; transform: translateY(6px); }} to {{ opacity: 1; transform: translateY(0); }} }}
</style>
<script src="https://unpkg.com/lucide@latest"></script>
</head>
<body>
<div class="header">
  <h1>{_icon('shield', 24)} Databricks SAT Report{ws_name_line}</h1>
  <div class="sub">Security Analysis Tool &middot; {_esc(result.workspace_url)} &middot; {ts}{f'&nbsp;&nbsp;&middot;&nbsp;&nbsp;<a href="{_esc(summary_link)}" style="color:#93c5fd;text-decoration:none;font-weight:600">{_icon("arrow-left",13)} Back to Summary</a>' if summary_link else ''}</div>
</div>
<div class="layout">
  <nav class="sidebar">
    {tab_buttons}
  </nav>
  <div class="main">
    {kpis_html}
    <div class="search-bar">
      <input type="text" id="searchInput" placeholder="Search checks, categories, keywords..." oninput="doSearch(this.value)">
      <span id="searchCount" class="search-count"></span>
    </div>
    {tab_panels}
    <div class="footer">Generated by SAT Scanner CLI &middot; {datetime.now().strftime('%Y-%m-%d')}</div>
  </div>
</div>  
<script>
lucide.createIcons();
function doSearch(q) {{
  q = q.toLowerCase().trim();
  var total = 0;
  document.querySelectorAll('.tab-panel').forEach(function(panel) {{
    var tabCount = 0;
    var rows = panel.querySelectorAll('tr');
    for (var i = 0; i < rows.length; i++) {{
      if (rows[i].querySelector('th')) continue;
      var text = rows[i].textContent.toLowerCase();
      if (!q || text.indexOf(q) >= 0) {{
        rows[i].style.display = '';
        if (q) tabCount++;
      }} else {{
        rows[i].style.display = 'none';
      }}
    }}
    var cards = panel.querySelectorAll('.finding-card');
    for (var j = 0; j < cards.length; j++) {{
      var ct = cards[j].textContent.toLowerCase();
      if (!q || ct.indexOf(q) >= 0) {{
        cards[j].style.display = '';
        if (q) tabCount++;
      }} else {{
        cards[j].style.display = 'none';
      }}
    }}
    total += tabCount;
    var tabId = panel.id.replace('panel-', '');
    var btn = document.querySelector('.tab-btn[data-tab="' + tabId + '"]');
    if (btn) {{
      var b = btn.querySelector('.search-badge');
      if (!b) {{ b = document.createElement('span'); b.className = 'search-badge'; btn.appendChild(b); }}
      b.textContent = (q && tabCount) ? tabCount : '';
      b.style.display = (q && tabCount) ? 'inline-block' : 'none';
    }}
  }});
  var badge = document.getElementById('searchCount');
  if (badge) badge.textContent = q ? total + ' match' + (total !== 1 ? 'es' : '') : '';
}}
document.querySelectorAll('.tab-btn').forEach(function(btn) {{
  btn.addEventListener('click', function() {{
    document.querySelectorAll('.tab-btn').forEach(function(b) {{ b.classList.remove('active'); }});
    document.querySelectorAll('.tab-panel').forEach(function(p) {{ p.style.display = 'none'; }});
    btn.classList.add('active');
    var panel = document.getElementById('panel-' + btn.getAttribute('data-tab'));
    if (panel) panel.style.display = 'block';
    var si = document.getElementById('searchInput');
    if (si) doSearch(si.value);
  }});
}});
</script>
</body></html>"""

    path = output_dir / f"{_file_prefix(result)}.html"
    path.write_text(html, encoding="utf-8")
    return str(path)


# ─────────────────────────────────────────────────────────────────────────────
# Recommendation Summary — standalone HTML report
# ─────────────────────────────────────────────────────────────────────────────

def export_recommendation_summary(result: SATScanResult | None, output_dir: Path,
                                  show_cost: bool = False,
                                  findings: list[SATFinding] | None = None,
                                  show_architecture: bool = False) -> str:
    """Generate a standalone Recommendation Summary HTML with Overview + P1–P4 tabs.

    Pass ``findings`` directly for combined multi-workspace reports (result can be None).
    """
    import html as _html
    import json as _json
    _esc = _html.escape

    src = findings if findings is not None else (result.findings if result else [])
    prio_items = _build_prioritised_recommendations(src)
    if not prio_items:
        return ""

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    date_str = datetime.now().strftime("%Y-%m-%d")

    # ── Severity / status / effort badge helpers ──
    _SEV_COLORS = {"critical": "#dc2626", "high": "#ea580c", "medium": "#ca8a04", "low": "#2563eb"}
    _STATUS_COLORS = {"FAIL": "#dc2626", "WARN": "#ca8a04"}
    _EFFORT_COLORS = {
        "Quick Fix (5\u201315 min)": "#16a34a",
        "Moderate (1\u20134 hrs)": "#ca8a04",
        "Significant (1\u20133 days)": "#ea580c",
        "Project (1+ weeks)": "#6b7280",
    }

    def _badge(text: str, color: str) -> str:
        return (f'<span style="background:{color};color:#fff;padding:2px 10px;border-radius:9999px;'
                f'font-size:11px;font-weight:600;text-transform:uppercase;white-space:nowrap">'
                f'{_esc(text)}</span>')

    # ── Counts ──
    total = len(prio_items)
    fail_count = sum(1 for i in prio_items if i["status"] == "FAIL")
    warn_count = sum(1 for i in prio_items if i["status"] == "WARN")

    # Group by priority bucket
    _PRIO_BUCKETS = {"P1": [], "P2": [], "P3": [], "P4": []}
    for item in prio_items:
        bucket = item["priority_label"][:2]
        _PRIO_BUCKETS.setdefault(bucket, []).append(item)

    p1_count = len(_PRIO_BUCKETS["P1"])
    p2_count = len(_PRIO_BUCKETS["P2"])
    p3_count = len(_PRIO_BUCKETS["P3"])
    p4_count = len(_PRIO_BUCKETS["P4"])

    # ── Category Breakdown ──
    cat_stats: dict[str, dict[str, int]] = {}
    for item in prio_items:
        cat = item["category"]
        if cat not in cat_stats:
            cat_stats[cat] = {"fail": 0, "warn": 0}
        if item["status"] == "FAIL":
            cat_stats[cat]["fail"] += 1
        else:
            cat_stats[cat]["warn"] += 1
    # Sort by total descending
    sorted_cats = sorted(cat_stats.items(), key=lambda x: x[1]["fail"] + x[1]["warn"], reverse=True)

    cat_rows = ""
    for cat, counts in sorted_cats:
        cat_total = counts["fail"] + counts["warn"]
        pct = max(2, round(counts["fail"] / cat_total * 100)) if cat_total else 0
        cat_rows += (
            f'<tr><td style="font-weight:600">{_esc(cat)}</td>'
            f'<td style="text-align:center;color:#dc2626;font-weight:700">{counts["fail"]}</td>'
            f'<td style="text-align:center;color:#ca8a04;font-weight:700">{counts["warn"]}</td>'
            f'<td style="text-align:center;font-weight:700">{cat_total}</td>'
            f'<td><div style="background:#e2e8f0;border-radius:4px;height:8px;width:100%">'
            f'<div style="background:#dc2626;border-radius:4px;height:8px;width:{pct}%"></div></div></td></tr>'
        )

    # ── Effort Breakdown ──
    effort_counts: dict[str, int] = {}
    for item in prio_items:
        eff = item.get("effort", "Moderate (1\u20134 hrs)")
        effort_counts[eff] = effort_counts.get(eff, 0) + 1
    effort_order = ["Quick Fix (5\u201315 min)", "Moderate (1\u20134 hrs)", "Significant (1\u20133 days)", "Project (1+ weeks)"]
    effort_rows = ""
    for eff in effort_order:
        cnt = effort_counts.get(eff, 0)
        if cnt:
            ec = _EFFORT_COLORS.get(eff, "#6b7280")
            effort_rows += (
                f'<tr><td>{_badge(eff, ec)}</td>'
                f'<td style="text-align:center;font-weight:700;font-size:18px">{cnt}</td></tr>'
            )

    # ── Cost summary for overview (optional) ──
    cost_kpi = ""
    if show_cost:
        total_low = sum(i["cost_low"] for i in prio_items if i.get("cost_low"))
        total_high = sum(i["cost_high"] for i in prio_items if i.get("cost_high"))
        if total_low:
            cost_kpi = (
                f'<div class="kpi" style="border-top:3px solid #b45309">'
                f'<div class="kpi-label">Est. Monthly Cost</div>'
                f'<div class="kpi-value" style="color:#b45309;font-size:20px">${total_low:,} &ndash; ${total_high:,}</div></div>'
            )

    # ── Overview content ──
    overview = f"""<div class="kpis">
<div class="kpi" style="border-top:3px solid #374151"><div class="kpi-label">Total Findings</div><div class="kpi-value" style="color:#374151">{total}</div></div>
<div class="kpi" style="border-top:3px solid #dc2626"><div class="kpi-label">Failures</div><div class="kpi-value" style="color:#dc2626">{fail_count}</div></div>
<div class="kpi" style="border-top:3px solid #ca8a04"><div class="kpi-label">Warnings</div><div class="kpi-value" style="color:#ca8a04">{warn_count}</div></div>
<div class="kpi" style="border-top:3px solid #dc2626"><div class="kpi-label">P1 - Immediate</div><div class="kpi-value" style="color:#dc2626">{p1_count}</div></div>
<div class="kpi" style="border-top:3px solid #ea580c"><div class="kpi-label">P2 - This Sprint</div><div class="kpi-value" style="color:#ea580c">{p2_count}</div></div>
<div class="kpi" style="border-top:3px solid #ca8a04"><div class="kpi-label">P3 - Next Sprint</div><div class="kpi-value" style="color:#ca8a04">{p3_count}</div></div>
<div class="kpi" style="border-top:3px solid #6b7280"><div class="kpi-label">P4 - Backlog</div><div class="kpi-value" style="color:#6b7280">{p4_count}</div></div>
{cost_kpi}
</div>
<div style="display:grid;grid-template-columns:1fr 1fr;gap:24px;margin-top:24px">
<div>
<h3 style="font-size:15px;font-weight:700;color:#0f172a;margin:0 0 12px"><i data-lucide="layers" style="width:16px;height:16px;vertical-align:middle;margin-right:6px"></i>Category Breakdown</h3>
<table><tr><th>Category</th><th style="text-align:center">Fail</th><th style="text-align:center">Warn</th><th style="text-align:center">Total</th><th style="width:120px">Fail %</th></tr>
{cat_rows}</table>
</div>
<div>
<h3 style="font-size:15px;font-weight:700;color:#0f172a;margin:0 0 12px"><i data-lucide="timer" style="width:16px;height:16px;vertical-align:middle;margin-right:6px"></i>Effort Breakdown</h3>
<table><tr><th>Effort Level</th><th style="text-align:center">Count</th></tr>
{effort_rows}</table>
</div>
</div>"""

    # ── P1–P4 tab content ──
    _PRIO_LABELS = {
        "P1": ("Fix Immediately", "#dc2626", "Critical and high-severity findings that pose immediate security risk. Quick fixes with highest impact."),
        "P2": ("Fix This Sprint", "#ea580c", "High and medium-severity findings to address within the current sprint."),
        "P3": ("Plan for Next Sprint", "#ca8a04", "Medium-severity findings and architectural improvements for upcoming sprints."),
        "P4": ("Backlog", "#6b7280", "Low-severity findings and best-practice improvements to schedule when capacity allows."),
    }

    prio_panels = ""
    for px in ["P1", "P2", "P3", "P4"]:
        items = _PRIO_BUCKETS[px]
        _, color, desc = _PRIO_LABELS[px]
        if not items:
            panel_body = '<p style="color:#64748b;font-style:italic;padding:24px">No findings in this priority level.</p>'
        else:
            # Group by category
            by_cat: dict[str, list[dict]] = {}
            for item in items:
                by_cat.setdefault(item["category"], []).append(item)

            panel_body = ""
            for cat, cat_items in by_cat.items():
                panel_body += (
                    f'<h3 style="font-size:14px;font-weight:700;color:#0f172a;margin:20px 0 8px;'
                    f'border-bottom:1px solid #e2e8f0;padding-bottom:6px;display:flex;align-items:center;justify-content:space-between">'
                    f'<span><i data-lucide="folder" style="width:14px;height:14px;vertical-align:middle;margin-right:6px"></i>'
                    f'{_esc(cat)}</span>'
                    f'<button class="jira-story-btn" data-category="{_esc(cat)}" '
                    f'style="font-size:11px;padding:3px 10px;border:1px solid #3b82f6;color:#3b82f6;'
                    f'background:#fff;border-radius:4px;cursor:pointer;display:none" '
                    f'title="Copy Story details to clipboard">&#128203; Copy Story</button>'
                    f'</h3>'
                )
                # Table header
                cost_th = '<th style="width:120px">Est. Cost</th>' if show_cost else ''
                panel_body += (
                    f'<table><tr><th style="width:140px">Check ID</th><th style="width:80px">Severity</th>'
                    f'<th style="width:70px">Status</th><th style="width:140px">Effort</th>{cost_th}<th>Title</th></tr>'
                )
                for item in cat_items:
                    sev_c = _SEV_COLORS.get(item["severity"], "#6b7280")
                    status_c = _STATUS_COLORS.get(item["status"], "#6b7280")
                    eff_c = _EFFORT_COLORS.get(item.get("effort", ""), "#6b7280")
                    cost_td = ""
                    if show_cost:
                        if item.get("cost_low"):
                            cost_td = f'<td style="font-size:12px;color:#b45309;white-space:nowrap">${item["cost_low"]:,} &ndash; ${item["cost_high"]:,}/mo</td>'
                        else:
                            cost_td = '<td style="font-size:11px;color:#94a3b8;text-align:center">&mdash;</td>'
                    cur_state = item.get("current_state", "")
                    cur_state_html = (
                        f'<div style="font-size:11px;color:#64748b;margin-top:2px;font-weight:400">{_esc(cur_state)}</div>'
                        if cur_state else ''
                    )
                    panel_body += (
                        f'<tr><td style="font-family:monospace;font-size:12px">{_esc(item["check_id"])}</td>'
                        f'<td>{_badge(item["severity"], sev_c)}</td>'
                        f'<td>{_badge(item["status"], status_c)}</td>'
                        f'<td>{_badge(item.get("effort", "Moderate (1\u20134 hrs)"), eff_c)}</td>'
                        f'{cost_td}'
                        f'<td style="font-weight:500">{_esc(item["title"])}{cur_state_html}</td></tr>'
                    )
                panel_body += '</table>'

                # Expandable details per check
                for item in cat_items:
                    rec = item.get("recommendation", "")
                    why = item.get("benefits", "")
                    cost_detail = ""
                    if show_cost and item.get("cost_reason"):
                        cost_detail = f'<p><strong>Estimated cost impact:</strong> {_esc(item["cost_reason"])}</p>'
                    _jira_data = _json.dumps({
                        "check_id": item["check_id"], "category": item["category"],
                        "severity": item["severity"], "status": item["status"],
                        "title": item["title"], "current_state": item.get("current_state", ""),
                        "recommendation": rec, "benefits": why,
                        "priority_label": item.get("priority_label", ""),
                        "effort": item.get("effort", ""),
                        "reference_url": item.get("reference_url", ""),
                        "remediation_plan": item.get("remediation_plan", {}),
                    }).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")
                    panel_body += (
                        f'<details style="margin:6px 0 12px 8px;font-size:13px" data-jira="{_jira_data}">'
                        f'<summary style="cursor:pointer;font-weight:600;color:#1e40af;display:flex;align-items:center;justify-content:space-between">'
                        f'<span><code>{_esc(item["check_id"])}</code> &mdash; {_esc(item["title"])}</span>'
                        f'<button class="jira-task-btn" '
                        f'style="font-size:10px;padding:2px 8px;border:1px solid #16a34a;color:#16a34a;'
                        f'background:#fff;border-radius:4px;cursor:pointer;display:none;margin-left:8px" '
                        f'title="Copy Task details to clipboard">&#128203; Copy Task</button>'
                        f'</summary>'
                        f'<div style="padding:8px 16px;background:#f8fafc;border-radius:6px;margin-top:4px">'
                    )
                    cur_state_detail = item.get("current_state", "")
                    if cur_state_detail:
                        panel_body += f'<p><strong>Finding Details:</strong> {_esc(cur_state_detail)}</p>'
                    panel_body += f'<p><strong>Recommendation:</strong> {_esc(rec)}</p>'
                    if why:
                        panel_body += f'<p><strong>Why it matters:</strong> {_esc(why)}</p>'
                    panel_body += f'{cost_detail}</div></details>'

        prio_panels += (
            f'<div id="panel-{px.lower()}" class="tab-panel">'
            f'<div style="background:{color}10;border-left:4px solid {color};padding:12px 16px;'
            f'border-radius:0 8px 8px 0;margin-bottom:20px;font-size:13px;color:#374151">{desc}</div>'
            f'{panel_body}</div>\n'
        )

    # ── Count categories ──
    unique_cats = len(cat_stats)

    # ── Sidebar ──
    _arch_btn = ('<button class="tab-btn" data-tab="architecture">'
                 '<i data-lucide="boxes" style="width:15px;height:15px;vertical-align:middle;margin-right:6px"></i>Architecture</button>\n') if show_architecture else ''
    sidebar = (
        '<button class="tab-btn active" data-tab="overview">'
        '<i data-lucide="bar-chart-3" style="width:15px;height:15px;vertical-align:middle;margin-right:6px"></i>Overview</button>\n'
        f'<button class="tab-btn" data-tab="p1">'
        f'<i data-lucide="alert-circle" style="width:15px;height:15px;vertical-align:middle;margin-right:6px"></i>P1 &mdash; Fix Immediately</button>\n'
        f'<button class="tab-btn" data-tab="p2">'
        f'<i data-lucide="clock" style="width:15px;height:15px;vertical-align:middle;margin-right:6px"></i>P2 &mdash; Fix This Sprint</button>\n'
        f'<button class="tab-btn" data-tab="p3">'
        f'<i data-lucide="calendar" style="width:15px;height:15px;vertical-align:middle;margin-right:6px"></i>P3 &mdash; Plan for Next Sprint</button>\n'
        f'<button class="tab-btn" data-tab="p4">'
        f'<i data-lucide="archive" style="width:15px;height:15px;vertical-align:middle;margin-right:6px"></i>P4 &mdash; Backlog</button>\n'
        f'<hr style="margin:12px 16px;border:none;border-top:1px solid #e2e8f0">\n'
        f'<button class="tab-btn" data-tab="jira">'
        f'<i data-lucide="ticket" style="width:15px;height:15px;vertical-align:middle;margin-right:6px"></i>Jira Export</button>\n'
        f'<button class="tab-btn" data-tab="ado">'
        f'<i data-lucide="layout-list" style="width:15px;height:15px;vertical-align:middle;margin-right:6px"></i>Azure DevOps</button>\n'
        f'<button class="tab-btn" data-tab="roadmap">'
        f'<i data-lucide="map" style="width:15px;height:15px;vertical-align:middle;margin-right:6px"></i>Roadmap</button>\n'
        f'<button class="tab-btn" data-tab="change-mgmt">'
        f'<i data-lucide="clipboard-check" style="width:15px;height:15px;vertical-align:middle;margin-right:6px"></i>Change Mgmt</button>\n'
        f'{_arch_btn}'
        f'<hr style="margin:12px 16px;border:none;border-top:1px solid #e2e8f0">\n'
        f'<button id="excel-export-btn" style="display:flex;align-items:center;width:100%;padding:10px 20px;border:none;'
        f'background:none;cursor:pointer;font-size:13px;color:#16a34a;text-align:left;font-weight:600;transition:all .15s">'
        f'<i data-lucide="file-spreadsheet" style="width:15px;height:15px;vertical-align:middle;margin-right:6px"></i>'
        f'Export to Excel</button>\n'
    )

    # ── Build remediation timeline ──
    from .remediation import build_remediation_timeline
    timeline = build_remediation_timeline(prio_items)

    # ── Roadmap panel ──
    roadmap_panel = '<div id="panel-roadmap" class="tab-panel">'
    roadmap_panel += '<div style="background:#f0fdf4;border-left:4px solid #16a34a;padding:12px 16px;border-radius:0 8px 8px 0;margin-bottom:20px;font-size:13px;color:#374151">Remediation roadmap grouped by priority phase and category. Effort estimates help plan sprint capacity.</div>'

    # Summary KPIs
    ts = timeline["summary"]
    roadmap_panel += f'<div class="kpis" style="margin-bottom:24px">'
    roadmap_panel += f'<div class="kpi" style="border-top:3px solid #374151"><div class="kpi-label">Total Findings</div><div class="kpi-value" style="color:#374151">{ts["total_findings"]}</div></div>'
    roadmap_panel += f'<div class="kpi" style="border-top:3px solid #16a34a"><div class="kpi-label">Total Effort</div><div class="kpi-value" style="color:#16a34a">{ts["total_effort_hours"]}h</div></div>'
    roadmap_panel += f'<div class="kpi" style="border-top:3px solid #8b5cf6"><div class="kpi-label">Working Days</div><div class="kpi-value" style="color:#8b5cf6">{ts["total_working_days"]}d</div></div>'
    roadmap_panel += f'<div class="kpi" style="border-top:3px solid #0891b2"><div class="kpi-label">Est. Resources</div><div class="kpi-value" style="color:#0891b2">{ts["estimated_resources"]}</div></div>'
    roadmap_panel += f'<div class="kpi" style="border-top:3px solid #3b82f6"><div class="kpi-label">Categories</div><div class="kpi-value" style="color:#3b82f6">{ts["categories"]}</div></div>'

    # Per-phase effort KPIs
    _phase_colors = {"P1": "#dc2626", "P2": "#ea580c", "P3": "#ca8a04", "P4": "#6b7280"}
    for phase in timeline["phases"]:
        px = phase["priority"]
        pc = _phase_colors.get(px, "#6b7280")
        _res_label = f' / {phase["estimated_resources"]} res' if phase["estimated_resources"] else ''
        roadmap_panel += f'<div class="kpi" style="border-top:3px solid {pc}"><div class="kpi-label">{_esc(phase["phase"][:20])}</div><div class="kpi-value" style="color:{pc}">{phase["total_working_days"]}d{_res_label}</div></div>'
    roadmap_panel += '</div>'

    # Resource estimation note
    roadmap_panel += '<div style="background:#f0f9ff;border-left:4px solid #0891b2;padding:10px 14px;border-radius:0 8px 8px 0;margin-bottom:16px;font-size:12px;color:#374151">'
    roadmap_panel += '<strong>Resource Estimation:</strong> Based on 8 hours/working day. Resources = parallel team members needed to complete the phase within its timeline window (P1: 1 wk, P2: 2 wks, P3: 5 wks, P4: flexible).'
    roadmap_panel += '</div>'

    # Phase details
    for phase in timeline["phases"]:
        px = phase["priority"]
        pc = _phase_colors.get(px, "#6b7280")
        _res_txt = f' &bull; {phase["estimated_resources"]} resource(s)' if phase["estimated_resources"] else ''
        roadmap_panel += f'<h3 style="font-size:15px;font-weight:700;color:#0f172a;margin:24px 0 12px;border-left:4px solid {pc};padding-left:12px">{_esc(phase["phase"])} &mdash; {phase["total_working_days"]}d ({phase["total_hours"]}h){_res_txt}</h3>'

        if not phase["categories"]:
            roadmap_panel += '<p style="color:#64748b;font-style:italic;padding:8px 16px">No findings in this phase.</p>'
            continue

        for cat, cat_data in phase["categories"].items():
            roadmap_panel += f'<details style="margin:8px 0 12px 8px;font-size:13px" open>'
            roadmap_panel += f'<summary style="cursor:pointer;font-weight:600;color:#1e40af;padding:4px 0">'
            roadmap_panel += f'<i data-lucide="folder" style="width:14px;height:14px;vertical-align:middle;margin-right:6px"></i>'
            roadmap_panel += f'{_esc(cat)} &mdash; {cat_data["subtotal_working_days"]}d ({cat_data["subtotal_hours"]}h) &bull; {len(cat_data["findings"])} findings</summary>'
            roadmap_panel += '<table style="margin:8px 0"><tr><th>Check ID</th><th>Title</th><th>Severity</th><th>Effort</th><th>Est. Hours</th><th>Working Days</th></tr>'
            for f in cat_data["findings"]:
                sev_c = _SEV_COLORS.get(f["severity"], "#6b7280")
                roadmap_panel += f'<tr><td style="font-family:monospace;font-size:12px">{_esc(f["check_id"])}</td>'
                roadmap_panel += f'<td style="font-weight:500">{_esc(f["title"])}</td>'
                roadmap_panel += f'<td>{_badge(f["severity"], sev_c)}</td>'
                roadmap_panel += f'<td style="font-size:12px">{_esc(f["effort"])}</td>'
                roadmap_panel += f'<td style="text-align:center;font-weight:600">{f["effort_hours"]}</td>'
                roadmap_panel += f'<td style="text-align:center;font-weight:600">{f["working_days"]}</td></tr>'
            roadmap_panel += '</table></details>'
    roadmap_panel += '</div>\n'

    # ── Change Management panel ──
    change_panel = '<div id="panel-change-mgmt" class="tab-panel">'
    change_panel += '<div style="background:#faf5ff;border-left:4px solid #7c3aed;padding:12px 16px;border-radius:0 8px 8px 0;margin-bottom:20px;font-size:13px;color:#374151">Change management templates for each finding. Includes prerequisites, checklists, rollback plans, and approval requirements.</div>'

    for px in ["P1", "P2", "P3", "P4"]:
        px_items = [i for i in prio_items if i.get("priority_label", "").startswith(px)]
        if not px_items:
            continue
        _, pcolor, _ = _PRIO_LABELS[px]
        change_panel += f'<h3 style="font-size:15px;font-weight:700;color:#0f172a;margin:24px 0 12px;border-left:4px solid {pcolor};padding-left:12px">{px} Findings</h3>'

        for item in px_items:
            plan = item.get("remediation_plan", {})
            cl = plan.get("checklist", {})
            ia = plan.get("impact_assessment", {})
            cm = plan.get("change_management", {})

            change_panel += f'<details style="margin:8px 0 12px 0;font-size:13px;border:1px solid #e2e8f0;border-radius:8px;padding:0">'
            change_panel += f'<summary style="cursor:pointer;font-weight:600;color:#1e40af;padding:12px 16px;background:#f8fafc;border-radius:8px">'
            change_panel += f'<code>{_esc(item["check_id"])}</code> &mdash; {_esc(item["title"])}</summary>'
            change_panel += '<div style="padding:12px 16px">'

            # Change type badge
            ct_color = "#dc2626" if cm.get("change_type") == "Emergency" else ("#ca8a04" if cm.get("change_type") == "Standard" else "#16a34a")
            change_panel += f'<div style="display:flex;gap:12px;flex-wrap:wrap;margin-bottom:12px">'
            change_panel += f'{_badge(cm.get("change_type", "Standard"), ct_color)}'
            change_panel += f'{_badge("Approval Required" if cm.get("approval_required") else "No Approval", "#7c3aed" if cm.get("approval_required") else "#16a34a")}'
            change_panel += f'{_badge(ia.get("downtime", "none") + " downtime", "#6b7280")}'
            change_panel += f'{_badge(ia.get("blast_radius", "workspace") + " scope", "#3b82f6")}'
            change_panel += f'{_badge(str(plan.get("estimated_duration_hours", "")) + "h est.", "#0f172a")}'
            change_panel += '</div>'

            # Prerequisites
            if plan.get("prerequisites"):
                change_panel += '<p style="font-weight:600;margin:8px 0 4px;color:#475569">Prerequisites</p><ul style="margin:0 0 8px 20px;padding:0">'
                for p in plan["prerequisites"]:
                    change_panel += f'<li style="margin:2px 0">{_esc(p)}</li>'
                change_panel += '</ul>'

            # Checklist sections
            for section_name, section_key in [("Pre-Checks", "pre_checks"), ("Remediation Steps", "steps"), ("Post-Validation", "post_validation"), ("Rollback", "rollback")]:
                items_list = cl.get(section_key, [])
                if items_list:
                    change_panel += f'<p style="font-weight:600;margin:8px 0 4px;color:#475569">{section_name}</p><ul style="margin:0 0 8px 20px;padding:0;list-style:none">'
                    for s in items_list:
                        icon = "&#9744; " if section_key != "rollback" else "&#8226; "
                        change_panel += f'<li style="margin:2px 0">{icon}{_esc(s)}</li>'
                    change_panel += '</ul>'

            # Stakeholders & Window
            change_panel += f'<div style="display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-top:8px;font-size:12px;color:#64748b">'
            change_panel += f'<div><strong>Stakeholders:</strong> {_esc(", ".join(plan.get("stakeholders", [])))}</div>'
            change_panel += f'<div><strong>Change Window:</strong> {_esc(cm.get("suggested_change_window", ""))}</div>'
            change_panel += f'<div><strong>Testing:</strong> {_esc(cm.get("testing_plan", ""))}</div>'
            change_panel += f'<div><strong>Communication:</strong> {_esc(cm.get("communication_plan", ""))}</div>'
            change_panel += '</div>'

            change_panel += '</div></details>'
    change_panel += '</div>\n'

    # ── Architecture panel (opt-in via --architecture) ──
    arch_panel = ''
    if show_architecture:
        from .checks import SAT_CHECKS, CATEGORY_DEFINITIONS
        _total_checks = len(SAT_CHECKS)
        _cat_counts: dict[str, int] = {}
        _sev_counts: dict[str, int] = {"critical": 0, "high": 0, "medium": 0, "low": 0}
        for _ck in SAT_CHECKS.values():
            _cat_counts[_ck["category"]] = _cat_counts.get(_ck["category"], 0) + 1
            _sev_counts[_ck["severity"]] = _sev_counts.get(_ck["severity"], 0) + 1
        _n_cats = len(_cat_counts)
        _n_yamls = len(list((_CHECKS_DIR_ARCH := __import__("pathlib").Path(__file__).parent / "checks").glob("*.yaml")))

        arch_panel = '<div id="panel-architecture" class="tab-panel">'
        arch_panel += '<div style="background:#f0f9ff;border-left:4px solid #0284c7;padding:12px 16px;border-radius:0 8px 8px 0;margin-bottom:20px;font-size:13px;color:#374151">SAT Scanner architecture — data flow, scan execution sequence, security domain coverage, scoring methodology, remediation pipeline, and external integrations.</div>'

        # ── Key Stats KPIs (dynamic) ──
        arch_panel += '<div class="kpis" style="margin-bottom:24px">'
        arch_panel += f'<div class="kpi" style="border-top:3px solid #1e40af"><div class="kpi-label">Total Checks</div><div class="kpi-value" style="color:#1e40af">{_total_checks}</div></div>'
        arch_panel += f'<div class="kpi" style="border-top:3px solid #dc2626"><div class="kpi-label">Critical</div><div class="kpi-value" style="color:#dc2626">{_sev_counts["critical"]}</div></div>'
        arch_panel += f'<div class="kpi" style="border-top:3px solid #ea580c"><div class="kpi-label">High</div><div class="kpi-value" style="color:#ea580c">{_sev_counts["high"]}</div></div>'
        arch_panel += f'<div class="kpi" style="border-top:3px solid #ca8a04"><div class="kpi-label">Medium</div><div class="kpi-value" style="color:#ca8a04">{_sev_counts["medium"]}</div></div>'
        arch_panel += f'<div class="kpi" style="border-top:3px solid #16a34a"><div class="kpi-label">Low</div><div class="kpi-value" style="color:#16a34a">{_sev_counts["low"]}</div></div>'
        arch_panel += f'<div class="kpi" style="border-top:3px solid #7c3aed"><div class="kpi-label">Categories</div><div class="kpi-value" style="color:#7c3aed">{_n_cats}</div></div>'
        arch_panel += f'<div class="kpi" style="border-top:3px solid #0284c7"><div class="kpi-label">YAML Files</div><div class="kpi-value" style="color:#0284c7">{_n_yamls}</div></div>'
        arch_panel += '<div class="kpi" style="border-top:3px solid #374151"><div class="kpi-label">Export Formats</div><div class="kpi-value" style="color:#374151">7</div></div>'
        arch_panel += '</div>'

        # ── 1. High-Level Architecture ──
        arch_panel += '<h3 style="font-size:16px;font-weight:700;color:#0f172a;margin:20px 0 12px;border-left:4px solid #0284c7;padding-left:12px">1. High-Level Architecture</h3>'
        arch_panel += '''<pre class="mermaid" style="background:#fff;border:1px solid #e2e8f0;border-radius:8px;padding:16px">
    flowchart TD
        subgraph Input["Input Layer"]
            A["CLI<br/><code>cli.py</code>"]
            AUTH{"Auth Method"}
            PAT["PAT Token"]
            AAD["Azure AD<br/><code>azure_auth.py</code>"]
        end

        subgraph Core["Core Engine"]
            direction TB
            CHECKS["Check Registry<br/><code>checks.py</code><br/>35 YAML files"]
            API["API Client<br/><code>api.py</code><br/>httpx async"]
            SCAN["Scanner<br/><code>scanner.py</code><br/>Orchestrator"]
            SECRET["Secret Scanner<br/><code>secret_scan.py</code><br/>TruffleHog"]
        end

        subgraph Processing["Processing Layer"]
            MODEL["Data Models<br/><code>models.py</code><br/>SATFinding / SATScanResult"]
            SCORE["Scoring Engine<br/><code>scoring.py</code><br/>Severity-weighted"]
            REMED["Remediation Planner<br/><code>remediation.py</code><br/>Timeline + Change Mgmt"]
        end

        subgraph Export["Export Layer"]
            direction TB
            EXP["Exporters<br/><code>exporters.py</code>"]
            COMB["Multi-Workspace<br/><code>combined.py</code>"]
            DELTA["Delta Tables<br/><code>delta.py</code>"]
            DASH["Lakeview<br/><code>dashboard.py</code>"]
        end

        subgraph Outputs["Output Formats"]
            JSON["JSON"]
            CSV["CSV"]
            XLS["Excel"]
            HTML["HTML Report"]
            JIRA["Jira Tickets"]
            DT["Delta Tables"]
            LV["Lakeview Dashboard"]
        end

        A --> AUTH
        AUTH -->|Token| PAT
        AUTH -->|Browser| AAD
        PAT --> API
        AAD --> API

        CHECKS --> SCAN
        API --> SCAN
        SCAN --> MODEL
        SECRET -.->|Optional| MODEL
        MODEL --> SCORE
        SCORE --> REMED
        REMED --> EXP

        EXP --> JSON & CSV & XLS & HTML & JIRA
        COMB --> HTML
        DELTA --> DT
        DASH --> LV

        style Input fill:#f0f9ff,stroke:#0284c7,stroke-width:2px
        style Core fill:#eff6ff,stroke:#1e40af,stroke-width:2px
        style Processing fill:#faf5ff,stroke:#7c3aed,stroke-width:2px
        style Export fill:#f0fdf4,stroke:#16a34a,stroke-width:2px
        style Outputs fill:#fefce8,stroke:#ca8a04,stroke-width:2px
        style SCAN fill:#1e40af,color:#fff
        style SCORE fill:#7c3aed,color:#fff
        style EXP fill:#16a34a,color:#fff
        style SECRET fill:#ea580c,color:#fff
    </pre>'''

        # ── 2. Scan Execution Sequence ──
        arch_panel += '<h3 style="font-size:16px;font-weight:700;color:#0f172a;margin:28px 0 12px;border-left:4px solid #0284c7;padding-left:12px">2. Scan Execution Sequence</h3>'
        arch_panel += '''<pre class="mermaid" style="background:#fff;border:1px solid #e2e8f0;border-radius:8px;padding:16px">
    sequenceDiagram
        actor User
        participant CLI as cli.py
        participant Auth as azure_auth.py
        participant API as api.py
        participant Scan as scanner.py
        participant Checks as checks/*.yaml
        participant Score as scoring.py
        participant Remed as remediation.py
        participant Export as exporters.py

        User->>CLI: sat-scanner --workspace-url ...
        CLI->>Auth: Authenticate (PAT / Azure AD)
        Auth-->>CLI: Token

        CLI->>Checks: Load 331+ check definitions
        Checks-->>CLI: SAT_CHECKS registry

        CLI->>Scan: run_scan(workspace, checks)

        loop For each API endpoint
            Scan->>API: GET /api/2.0/...
            API-->>Scan: JSON response
        end

        loop For each check
            Scan->>Scan: Evaluate check logic
            Scan-->>Scan: Create SATFinding (PASS/FAIL/WARN)
        end

        Scan-->>CLI: List[SATFinding]

        CLI->>Score: compute_score(findings)
        Score-->>CLI: SATScanResult (score, breakdown)

        CLI->>Remed: generate plans + timeline
        Remed-->>CLI: Remediation plans

        CLI->>Export: export_all(result)
        Export-->>User: JSON, CSV, Excel, HTML, Jira
    </pre>'''

        # ── 3. Security Domain Coverage ──
        arch_panel += '<h3 style="font-size:16px;font-weight:700;color:#0f172a;margin:28px 0 12px;border-left:4px solid #0284c7;padding-left:12px">3. Security Domain Coverage</h3>'
        # Build a pie chart from actual category data
        _pie_items = sorted(_cat_counts.items(), key=lambda x: -x[1])[:15]
        _pie_str = "\n".join(f'    "{_esc(c)}" : {n}' for c, n in _pie_items)
        arch_panel += f'''<pre class="mermaid" style="background:#fff;border:1px solid #e2e8f0;border-radius:8px;padding:16px">
    pie title Check Distribution by Category ({_total_checks} total)
    {_pie_str}
    </pre>'''

        # Category detail table
        arch_panel += '<details style="margin:12px 0;font-size:13px"><summary style="cursor:pointer;font-weight:600;color:#1e40af;padding:4px 0">View all categories &amp; check counts</summary>'
        arch_panel += '<table style="margin:8px 0;width:100%"><tr><th style="text-align:left">Category</th><th style="text-align:center">Checks</th><th style="text-align:left">Description</th></tr>'
        _cat_desc = {c: d for c, d in CATEGORY_DEFINITIONS}
        for _cat, _cnt in sorted(_cat_counts.items(), key=lambda x: -x[1]):
            arch_panel += f'<tr><td style="font-weight:600">{_esc(_cat)}</td><td style="text-align:center;font-weight:700;color:#1e40af">{_cnt}</td><td style="font-size:12px;color:#64748b">{_esc(_cat_desc.get(_cat, ""))}</td></tr>'
        arch_panel += '</table></details>'

        # ── 4. Scoring Methodology ──
        arch_panel += '<h3 style="font-size:16px;font-weight:700;color:#0f172a;margin:28px 0 12px;border-left:4px solid #0284c7;padding-left:12px">4. Scoring &amp; Grading</h3>'
        arch_panel += '''<pre class="mermaid" style="background:#fff;border:1px solid #e2e8f0;border-radius:8px;padding:16px">
    flowchart TD
        subgraph Findings["Each Finding"]
            F["Finding"]
            F --> S{"Status"}
            S -->|PASS| P["Penalty: 0"]
            S -->|FAIL| FL["Penalty: Full Weight"]
            S -->|WARN| W["Penalty: 0.5 &times; Weight"]
            S -->|NOT_APPLICABLE| NA["Excluded from score"]
            S -->|API ERROR| AE["Excluded from score"]
        end

        subgraph Weights["Severity Weights"]
            C["Critical = 10"]
            H["High = 7"]
            M["Medium = 4"]
            L["Low = 2"]
        end

        subgraph Calc["Score Calculation"]
            POOL["Total Weight Pool<br/>Sum of all applicable weights"]
            PEN["Total Penalty<br/>Sum of FAIL + WARN penalties"]
            FORMULA["Score = (1 - Penalty / Pool) &times; 100"]
        end

        subgraph Grade["Grade Bands"]
            G1["80 - 100: Good<br/>Strong security posture"]
            G2["60 - 79: Needs Improvement<br/>Gaps weaken posture"]
            G3["0 - 59: Critical<br/>Immediate action required"]
        end

        FL & W & P --> PEN
        Weights --> POOL
        PEN & POOL --> FORMULA
        FORMULA --> Grade

        style Findings fill:#f8fafc,stroke:#94a3b8,stroke-width:2px
        style Weights fill:#fef3c7,stroke:#d97706,stroke-width:2px
        style Calc fill:#faf5ff,stroke:#7c3aed,stroke-width:2px
        style Grade fill:#f0fdf4,stroke:#16a34a,stroke-width:2px
        style FL fill:#dc2626,color:#fff
        style W fill:#ea580c,color:#fff
        style P fill:#16a34a,color:#fff
        style G1 fill:#16a34a,color:#fff
        style G2 fill:#ca8a04,color:#fff
        style G3 fill:#dc2626,color:#fff
    </pre>'''

        # ── 5. Remediation Pipeline ──
        arch_panel += '<h3 style="font-size:16px;font-weight:700;color:#0f172a;margin:28px 0 12px;border-left:4px solid #0284c7;padding-left:12px">5. Remediation Pipeline</h3>'
        arch_panel += '''<pre class="mermaid" style="background:#fff;border:1px solid #e2e8f0;border-radius:8px;padding:16px">
    flowchart LR
        subgraph Input["Finding Metadata"]
            SEV["Severity"]
            EFF["Effort Estimate"]
            CAT["Category"]
            REC["Recommendation Text"]
        end

        subgraph AutoGen["Auto-Generation Engine"]
            PRE["Prerequisites<br/>Permissions, backups,<br/>category-specific"]
            IMP["Impact Assessment<br/>Downtime, blast radius,<br/>affected services"]
            CHK["Step-by-Step Checklist<br/>Pre-checks, steps,<br/>post-validation, rollback"]
            CM["Change Management<br/>Change type, approval,<br/>window, testing, comms"]
            STKH["Stakeholders<br/>Team assignments<br/>per category"]
        end

        subgraph Timeline["Timeline Builder"]
            P1["P1: Immediate<br/>Week 1"]
            P2["P2: Short-term<br/>Weeks 2-3"]
            P3["P3: Medium-term<br/>Weeks 4-8"]
            P4["P4: Long-term<br/>Backlog"]
            RES["Working Days &amp;<br/>Resource Estimation"]
        end

        subgraph Override["Optional YAML Overrides"]
            OV["Per-check overrides<br/>in checks/*.yaml"]
        end

        SEV & EFF & CAT & REC --> AutoGen
        Override -.-> AutoGen
        AutoGen --> Timeline

        style Input fill:#f0f9ff,stroke:#0284c7,stroke-width:2px
        style AutoGen fill:#faf5ff,stroke:#7c3aed,stroke-width:2px
        style Timeline fill:#fef3c7,stroke:#d97706,stroke-width:2px
        style Override fill:#f8fafc,stroke:#94a3b8,stroke-dasharray:5 5
        style P1 fill:#dc2626,color:#fff
        style P2 fill:#ea580c,color:#fff
        style P3 fill:#ca8a04,color:#fff
        style P4 fill:#6b7280,color:#fff
    </pre>'''

        # ── 6. External Integrations ──
        arch_panel += '<h3 style="font-size:16px;font-weight:700;color:#0f172a;margin:28px 0 12px;border-left:4px solid #0284c7;padding-left:12px">6. External Integrations</h3>'
        arch_panel += '''<pre class="mermaid" style="background:#fff;border:1px solid #e2e8f0;border-radius:8px;padding:16px">
    flowchart TB
        subgraph Scanner["SAT Scanner"]
            CORE["Core Engine"]
        end

        subgraph DatabricksAPIs["Databricks REST APIs"]
            direction TB
            WS["Workspace Config<br/>/api/2.0/workspace/*"]
            CL["Clusters &amp; Policies<br/>/api/2.0/clusters/*"]
            JB["Jobs &amp; Runs<br/>/api/2.1/jobs/*"]
            SQL["SQL Warehouses<br/>/api/2.0/sql/*"]
            UC["Unity Catalog<br/>/api/2.1/unity-catalog/*"]
            SEC["Secrets &amp; Tokens<br/>/api/2.0/secrets/*"]
            ML["ML &amp; Serving<br/>/api/2.0/serving-endpoints"]
            SCIM["SCIM / Users<br/>/api/2.0/preview/scim/*"]
        end

        subgraph Azure["Azure Services"]
            AAD["Azure AD<br/>Browser OAuth"]
            ARM["ARM API<br/>Resource Metadata"]
        end

        subgraph OutputSvc["Output Services"]
            JIRA["Jira Cloud<br/>Story/Task Creation"]
            DL["Delta Lake<br/>SQL Warehouse Exec"]
            LV["Lakeview<br/>Dashboard API"]
        end

        subgraph SecScan["Secret Scanning"]
            TH["TruffleHog Engine<br/>800+ Detector Patterns"]
        end

        CORE --> DatabricksAPIs
        CORE --> Azure
        CORE --> OutputSvc
        CORE -.-> SecScan

        style Scanner fill:#0f172a,color:#fff,stroke-width:2px
        style DatabricksAPIs fill:#eff6ff,stroke:#1e40af,stroke-width:2px
        style Azure fill:#f0f9ff,stroke:#0284c7,stroke-width:2px
        style OutputSvc fill:#f0fdf4,stroke:#16a34a,stroke-width:2px
        style SecScan fill:#fff7ed,stroke:#ea580c,stroke-width:2px,stroke-dasharray:5 5
        style CORE fill:#1e40af,color:#fff
    </pre>'''

        # ── 7. Export Pipeline ──
        arch_panel += '<h3 style="font-size:16px;font-weight:700;color:#0f172a;margin:28px 0 12px;border-left:4px solid #0284c7;padding-left:12px">7. Export Pipeline &amp; Report Contents</h3>'
        arch_panel += '''<pre class="mermaid" style="background:#fff;border:1px solid #e2e8f0;border-radius:8px;padding:16px">
    flowchart LR
        subgraph Data["Scan Data"]
            RES["SATScanResult"]
            PRIO["Prioritised<br/>Recommendations"]
            TL["Remediation<br/>Timeline"]
        end

        subgraph JSON_E["JSON Export"]
            J1["Full findings + metadata"]
            J2["Score breakdown"]
            J3["Remediation plans"]
            J4["Timeline roadmap"]
        end

        subgraph Excel_E["Excel Export"]
            E1["Summary sheet"]
            E2["All Findings"]
            E3["P1-P4 sheets"]
            E4["Remediation Plan"]
            E5["Roadmap"]
            E6["Change Management"]
            E7["Checks Reference"]
        end

        subgraph HTML_E["HTML Report"]
            H1["Overview dashboard"]
            H2["P1-P4 panels"]
            H3["Jira Export tools"]
            H4["Roadmap timeline"]
            H5["Change Mgmt cards"]
            H6["Architecture diagrams"]
            H7["Export to Excel button"]
        end

        subgraph Other["Other Exports"]
            CSV1["CSV findings"]
            JIRA1["Jira Story/Task<br/>hierarchy"]
            DELTA1["Delta table<br/>+ Lakeview"]
        end

        Data --> JSON_E & Excel_E & HTML_E & Other

        style Data fill:#f8fafc,stroke:#94a3b8,stroke-width:2px
        style JSON_E fill:#fef3c7,stroke:#d97706,stroke-width:2px
        style Excel_E fill:#dcfce7,stroke:#16a34a,stroke-width:2px
        style HTML_E fill:#eff6ff,stroke:#1e40af,stroke-width:2px
        style Other fill:#faf5ff,stroke:#7c3aed,stroke-width:2px
    </pre>'''

        # ── Module file reference table ──
        arch_panel += '<h3 style="font-size:16px;font-weight:700;color:#0f172a;margin:28px 0 12px;border-left:4px solid #0284c7;padding-left:12px">Module Reference</h3>'
        arch_panel += '<table style="margin:8px 0;width:100%;font-size:13px"><tr><th style="text-align:left">Module</th><th style="text-align:left">Purpose</th><th style="text-align:left">Key Classes / Functions</th></tr>'
        _modules = [
            ("cli.py", "CLI entry point &amp; orchestration", "argparse, run_scan(), export_all()"),
            ("scanner.py", "Async scan engine — executes all checks", "run_checks(), _check_*() functions"),
            ("api.py", "Databricks REST API client", "_get(), _paginate(), _make_finding()"),
            ("checks.py", "YAML check registry &amp; constants", "SAT_CHECKS, CHECK_API_ENDPOINTS, PORTAL_LINKS"),
            ("models.py", "Data models", "SATFinding, SATScanResult"),
            ("scoring.py", "Score computation &amp; priority ranking", "_build_prioritised_recommendations()"),
            ("remediation.py", "Remediation plan auto-generation", "generate_remediation_plan(), build_remediation_timeline()"),
            ("exporters.py", "All export formats", "export_json(), export_excel(), export_recommendation_summary()"),
            ("secret_scan.py", "TruffleHog secret scanning", "scan_notebooks(), scan_clusters()"),
            ("azure_auth.py", "Azure AD browser auth flow", "azure_login_flow(), azure_tenant_flow()"),
            ("combined.py", "Multi-workspace aggregation", "export_combined_html(), export_combined_summary()"),
            ("delta.py", "Delta table persistence", "export_delta(), execute_statement()"),
            ("dashboard.py", "Lakeview dashboard generation", "create_dashboard(), update_dashboard()"),
            ("helpers.py", "Utility functions", "Logging, HTML rendering, evidence extraction"),
            ("validate.py", "Report validation", "Cross-check findings vs API responses"),
        ]
        for _mod, _purpose, _keys in _modules:
            arch_panel += f'<tr><td style="font-family:monospace;font-weight:600;color:#1e40af;white-space:nowrap">{_mod}</td><td>{_purpose}</td><td style="font-size:12px;color:#64748b">{_keys}</td></tr>'
        arch_panel += '</table>'

        arch_panel += '</div>\n'

    # Serialize timeline data for client-side Excel export
    import json as _json2
    timeline_json = _json2.dumps(timeline, ensure_ascii=True).replace("</", "<\\/")

    _mermaid_cdn = '<script src="https://cdn.jsdelivr.net/npm/mermaid@11/dist/mermaid.min.js"></script>' if show_architecture else ''
    _mermaid_init = ("mermaid.initialize({{ startOnLoad: false, theme: 'base', themeVariables: "
                     "{{ primaryColor: '#e0e7ff', primaryBorderColor: '#6366f1', primaryTextColor: '#1e293b', "
                     "lineColor: '#94a3b8', fontSize: '13px' }} }});") if show_architecture else ''

    html = f"""<!DOCTYPE html>
<html lang="en"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>SAT Recommendations Summary</title>
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif; background:#f1f5f9; color:#334155; }}
.header {{ background:linear-gradient(135deg,#0f172a 0%,#1e3a5f 100%); color:#fff; padding:24px 32px; }}
.header h1 {{ font-size:22px; font-weight:700; }}
.header .sub {{ font-size:12px; color:#94a3b8; margin-top:4px; }}
.layout {{ display:flex; min-height:calc(100vh - 72px); }}
.sidebar {{ width:260px; min-width:260px; background:#fff; border-right:1px solid #e2e8f0; padding:16px 0; }}
.tab-btn {{ display:flex; align-items:center; width:100%; padding:10px 20px; border:none; background:none;
  cursor:pointer; font-size:13px; color:#64748b; text-align:left; transition:all .15s; }}
.tab-btn:hover {{ background:#f1f5f9; color:#0f172a; }}
.tab-btn.active {{ background:#eff6ff; color:#1d4ed8; font-weight:600; border-left:3px solid #3b82f6; }}
.main {{ flex:1; padding:28px; overflow-y:auto; }}
.tab-panel {{ display:none; }}
.kpis {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(130px,1fr)); gap:12px; }}
.kpi {{ background:#fff; border:1px solid #e2e8f0; border-radius:10px; padding:16px; text-align:center; }}
.kpi-label {{ font-size:11px; color:#64748b; text-transform:uppercase; letter-spacing:.05em; }}
.kpi-value {{ font-size:28px; font-weight:700; margin-top:4px; }}
table {{ width:100%; border-collapse:collapse; font-size:13px; margin-bottom:8px; }}
th {{ background:#f8fafc; padding:8px 12px; text-align:left; font-size:11px; color:#64748b;
  text-transform:uppercase; letter-spacing:.04em; border-bottom:2px solid #e2e8f0; }}
td {{ padding:8px 12px; border-bottom:1px solid #f1f5f9; }}
tr:hover {{ background:#f8fafc; }}
.jira-story-btn:hover {{ background:#eff6ff !important; }}
.jira-task-btn:hover {{ background:#f0fdf4 !important; }}
details summary {{ padding:4px 0; }}
details p {{ margin:4px 0; line-height:1.5; }}
.footer {{ text-align:center; font-size:11px; color:#94a3b8; margin-top:32px; padding-top:16px; border-top:1px solid #e2e8f0; }}
@media (max-width:768px) {{
  .layout {{ flex-direction:column; }}
  .sidebar {{ width:100%; display:flex; flex-wrap:wrap; padding:8px; gap:4px; }}
  .tab-btn {{ width:auto; padding:6px 12px; font-size:12px; border-radius:6px; }}
  .kpis {{ grid-template-columns:repeat(2,1fr); }}
}}
@media print {{
  .sidebar {{ display:none; }}
  .tab-panel {{ display:block !important; page-break-inside:avoid; }}
  .header {{ padding:12px 16px; }}
}}
</style>
<script src="https://unpkg.com/lucide@latest"></script>
<script src="https://cdn.sheetjs.com/xlsx-0.20.3/package/dist/xlsx.full.min.js"></script>
{_mermaid_cdn}
</head><body>
<div class="header">
  <h1><i data-lucide="shield-check" style="width:24px;height:24px;vertical-align:middle;margin-right:8px"></i>SAT Recommendations Summary</h1>
  <div class="sub">Security Analysis Tool &middot; {_esc(now_str)} &middot; {total} findings across {unique_cats} categories</div>
</div>
<div class="layout">
  <nav class="sidebar">{sidebar}</nav>
  <div class="main">
    <div id="panel-overview" class="tab-panel" style="display:block">
{overview}
</div>
{prio_panels}
{roadmap_panel}
{change_panel}
{arch_panel}
    <div id="panel-jira" class="tab-panel">
      <div style="background:#eff6ff;border-left:4px solid #3b82f6;padding:12px 16px;border-radius:0 8px 8px 0;margin-bottom:20px;font-size:13px;color:#374151">
        Configure your Jira instance, then export findings as CSV for bulk import or generate a shell script for REST API creation. Story/Task hierarchy: Category &rarr; Story, Finding &rarr; Task.</div>
      <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:16px;margin-bottom:16px">
        <div>
          <label style="font-size:12px;font-weight:600;color:#475569">Jira Base URL</label>
          <input id="jira-url" type="text" placeholder="https://your-company.atlassian.net"
            style="width:100%;padding:8px 10px;border:1px solid #cbd5e1;border-radius:6px;font-size:13px;margin-top:4px">
        </div>
        <div>
          <label style="font-size:12px;font-weight:600;color:#475569">Project Key</label>
          <input id="jira-project" type="text" placeholder="SAT"
            style="width:100%;padding:8px 10px;border:1px solid #cbd5e1;border-radius:6px;font-size:13px;margin-top:4px">
        </div>
        <div>
          <label style="font-size:12px;font-weight:600;color:#475569">Project ID <span style="color:#94a3b8;font-weight:400">(numeric)</span></label>
          <div style="display:flex;gap:4px;margin-top:4px">
            <input id="jira-pid" type="text" placeholder="10001"
              style="flex:1;padding:8px 10px;border:1px solid #cbd5e1;border-radius:6px;font-size:13px">
            <button id="jira-find-id" type="button" style="padding:6px 10px;border:1px solid #3b82f6;color:#3b82f6;background:#fff;border-radius:6px;cursor:pointer;font-size:11px;font-weight:600;white-space:nowrap" title="Opens Jira API to show project ID">Find ID</button>
          </div>
          <div style="font-size:11px;color:#94a3b8;margin-top:2px">Click &ldquo;Find ID&rdquo; &rarr; look for <code style="background:#f1f5f9;padding:1px 4px;border-radius:3px">&quot;id&quot;:&quot;12345&quot;</code></div>
        </div>
      </div>
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:16px">
        <div>
          <label style="font-size:12px;font-weight:600;color:#475569">Parent Issue Type <span style="color:#94a3b8;font-weight:400">(default: Story)</span></label>
          <input id="jira-story-type" type="text" placeholder="Story"
            style="width:100%;padding:8px 10px;border:1px solid #cbd5e1;border-radius:6px;font-size:13px;margin-top:4px">
        </div>
        <div>
          <label style="font-size:12px;font-weight:600;color:#475569">Child Issue Type <span style="color:#94a3b8;font-weight:400">(default: Subtask)</span></label>
          <input id="jira-task-type" type="text" placeholder="Subtask"
            style="width:100%;padding:8px 10px;border:1px solid #cbd5e1;border-radius:6px;font-size:13px;margin-top:4px">
        </div>
      </div>
      <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:12px 16px;margin-bottom:16px;font-size:12px;color:#64748b">
        <strong style="color:#475569">How to find IDs:</strong> Click &ldquo;Find ID&rdquo; to open the project API response.
        For issue type IDs, open <code style="background:#e2e8f0;padding:1px 4px;border-radius:3px">/rest/api/2/issue/createmeta?projectKeys=YOUR_KEY</code> in your Jira instance.
        Look for <code style="background:#e2e8f0;padding:1px 4px;border-radius:3px">&quot;id&quot;</code> values next to <code style="background:#e2e8f0;padding:1px 4px;border-radius:3px">&quot;name&quot;:&quot;Story&quot;</code> and <code style="background:#e2e8f0;padding:1px 4px;border-radius:3px">&quot;name&quot;:&quot;Task&quot;</code>.
      </div>
      <div style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:20px">
        <label style="font-size:12px;font-weight:600;color:#475569;width:100%">Priority Filter</label>
        <select id="jira-filter" style="padding:6px 10px;border:1px solid #cbd5e1;border-radius:6px;font-size:13px">
          <option value="ALL">All Priorities</option>
          <option value="P1">P1 Only</option><option value="P2">P2 Only</option>
          <option value="P3">P3 Only</option><option value="P4">P4 Only</option>
        </select>
      </div>
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:16px">
        <div>
          <label style="font-size:12px;font-weight:600;color:#475569">Jira Email</label>
          <input id="jira-email" type="email" placeholder="you@company.com"
            style="width:100%;padding:8px 10px;border:1px solid #cbd5e1;border-radius:6px;font-size:13px;margin-top:4px">
        </div>
        <div>
          <label style="font-size:12px;font-weight:600;color:#475569">API Token <a href="https://id.atlassian.com/manage-profile/security/api-tokens" target="_blank" style="color:#3b82f6;font-size:11px;font-weight:400">(Create token)</a></label>
          <input id="jira-token" type="password" placeholder="Your Jira API token"
            style="width:100%;padding:8px 10px;border:1px solid #cbd5e1;border-radius:6px;font-size:13px;margin-top:4px">
        </div>
      </div>
      <div style="display:flex;gap:12px;flex-wrap:wrap;margin-bottom:12px">
        <button id="jira-push-btn" style="padding:10px 20px;background:#7c3aed;color:#fff;border:none;border-radius:6px;cursor:pointer;font-size:13px;font-weight:600">
          <i data-lucide="upload" style="width:14px;height:14px;vertical-align:middle;margin-right:6px"></i>Export to Jira</button>
        <button id="jira-csv-btn" style="padding:10px 20px;background:#3b82f6;color:#fff;border:none;border-radius:6px;cursor:pointer;font-size:13px;font-weight:600">
          <i data-lucide="download" style="width:14px;height:14px;vertical-align:middle;margin-right:6px"></i>Download CSV</button>
        <button id="jira-sh-btn" style="padding:10px 20px;background:#16a34a;color:#fff;border:none;border-radius:6px;cursor:pointer;font-size:13px;font-weight:600">
          <i data-lucide="terminal" style="width:14px;height:14px;vertical-align:middle;margin-right:6px"></i>Download Shell Script</button>
      </div>
      <div id="jira-progress" style="display:none;background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:16px;margin-bottom:12px">
        <div id="jira-prog-log" style="font-size:13px;color:#374151;line-height:1.7"></div>
      </div>
      <p style="font-size:11px;color:#94a3b8;margin-bottom:16px">Settings saved in localStorage. API token is never stored.</p>

      <details style="margin-bottom:12px">
        <summary style="cursor:pointer;font-size:14px;font-weight:700;color:#1e293b;padding:8px 0">
          <i data-lucide="book-open" style="width:14px;height:14px;vertical-align:middle;margin-right:6px"></i>
          How to Import CSV to Jira (Step-by-Step Guide)</summary>
        <div style="padding:12px 0;font-size:13px;color:#374151;line-height:1.7">

          <h4 style="font-size:13px;font-weight:700;color:#0f172a;margin:12px 0 8px">Option 1: CSV Import (Jira Admin Required)</h4>
          <ol style="margin:0 0 16px 20px;padding:0">
            <li style="margin-bottom:8px"><strong>Download the CSV</strong> &mdash; Click &ldquo;Download CSV&rdquo; above. Select priority filter (P1/P2/P3/P4 or All).</li>
            <li style="margin-bottom:8px"><strong>Go to Jira</strong> &rarr; Click the <strong>gear icon</strong> (top-right) &rarr; <strong>System</strong> &rarr; <strong>External System Import</strong> (left sidebar under &ldquo;Import and Export&rdquo;).</li>
            <li style="margin-bottom:8px"><strong>Select &ldquo;CSV&rdquo;</strong> as the import source.</li>
            <li style="margin-bottom:8px"><strong>Upload the CSV file</strong> and select your target project.</li>
            <li style="margin-bottom:8px"><strong>Map the fields</strong> &mdash; Jira auto-detects most columns:
              <table style="margin:8px 0;font-size:12px;border:1px solid #e2e8f0;border-radius:6px">
                <tr style="background:#f8fafc"><th style="padding:6px 12px;text-align:left">CSV Column</th><th style="padding:6px 12px;text-align:left">Jira Field</th></tr>
                <tr><td style="padding:4px 12px;border-top:1px solid #f1f5f9"><code>Issue Type</code></td><td style="padding:4px 12px;border-top:1px solid #f1f5f9">Issue Type (Story/Task)</td></tr>
                <tr><td style="padding:4px 12px;border-top:1px solid #f1f5f9"><code>Parent Summary</code></td><td style="padding:4px 12px;border-top:1px solid #f1f5f9">Parent &mdash; links Tasks under Stories</td></tr>
                <tr><td style="padding:4px 12px;border-top:1px solid #f1f5f9"><code>Summary</code></td><td style="padding:4px 12px;border-top:1px solid #f1f5f9">Summary</td></tr>
                <tr><td style="padding:4px 12px;border-top:1px solid #f1f5f9"><code>Priority</code></td><td style="padding:4px 12px;border-top:1px solid #f1f5f9">Priority</td></tr>
                <tr><td style="padding:4px 12px;border-top:1px solid #f1f5f9"><code>Description</code></td><td style="padding:4px 12px;border-top:1px solid #f1f5f9">Description</td></tr>
                <tr><td style="padding:4px 12px;border-top:1px solid #f1f5f9"><code>Labels</code></td><td style="padding:4px 12px;border-top:1px solid #f1f5f9">Labels (sat-scanner)</td></tr>
              </table>
            </li>
            <li style="margin-bottom:8px"><strong>Click &ldquo;Begin Import&rdquo;</strong> &mdash; Jira creates Stories first, then Tasks linked under each Story.</li>
            <li style="margin-bottom:8px"><strong>Verify</strong> &mdash; Search <code style="background:#f1f5f9;padding:2px 6px;border-radius:3px">labels = sat-scanner</code> to find all imported issues.</li>
          </ol>
          <div style="background:#fef3c7;border:1px solid #fcd34d;border-radius:6px;padding:10px 14px;margin-bottom:16px;font-size:12px">
            <strong style="color:#92400e">Note:</strong> CSV import requires <strong>Jira Admin</strong> permissions. If you don&rsquo;t have admin access, use Option 2 or 3 below.
          </div>

          <h4 style="font-size:13px;font-weight:700;color:#0f172a;margin:12px 0 8px">Option 2: Export to Jira (Direct API)</h4>
          <ol style="margin:0 0 16px 20px;padding:0">
            <li style="margin-bottom:8px"><strong>Create an API token</strong> &mdash; Go to <a href="https://id.atlassian.com/manage-profile/security/api-tokens" target="_blank" style="color:#3b82f6">Atlassian API Tokens</a> &rarr; <strong>Create API token</strong>.</li>
            <li style="margin-bottom:8px"><strong>Enter credentials above</strong> &mdash; Fill in your Jira email and API token.</li>
            <li style="margin-bottom:8px"><strong>Click &ldquo;Export to Jira&rdquo;</strong> &mdash; If CORS blocks it (common for local files), a <strong>&ldquo;Copy Console Script&rdquo;</strong> button appears.</li>
            <li style="margin-bottom:8px"><strong>Run the console script</strong> &mdash; Open your Jira instance in a browser tab &rarr; Press <strong>F12</strong> &rarr; Go to <strong>Console</strong> tab &rarr; <strong>Paste</strong> the script &rarr; Press <strong>Enter</strong>.</li>
            <li style="margin-bottom:8px">The script runs on the Jira domain (same-origin, no CORS), creating Stories and Tasks with full descriptions.</li>
          </ol>

          <h4 style="font-size:13px;font-weight:700;color:#0f172a;margin:12px 0 8px">Option 3: Shell Script (CLI)</h4>
          <ol style="margin:0 0 16px 20px;padding:0">
            <li style="margin-bottom:8px"><strong>Download the shell script</strong> &mdash; Click &ldquo;Download Shell Script&rdquo; above.</li>
            <li style="margin-bottom:8px"><strong>Create an API token</strong> (same as Option 2).</li>
            <li style="margin-bottom:8px"><strong>Run from terminal:</strong>
              <pre style="background:#0f172a;color:#e2e8f0;padding:12px 16px;border-radius:6px;margin:6px 0;font-size:12px;overflow-x:auto">JIRA_TOKEN=your-api-token JIRA_EMAIL=you@company.com bash sat-jira-create.sh</pre>
            </li>
            <li style="margin-bottom:8px">Requires <code style="background:#f1f5f9;padding:2px 6px;border-radius:3px">curl</code> and <code style="background:#f1f5f9;padding:2px 6px;border-radius:3px">jq</code> installed.</li>
          </ol>

          <h4 style="font-size:13px;font-weight:700;color:#0f172a;margin:12px 0 8px">CSV Structure</h4>
          <p style="margin:0 0 8px">The CSV uses a <strong>Story/Task hierarchy</strong>:</p>
          <ul style="margin:0 0 12px 20px;padding:0">
            <li style="margin-bottom:4px"><strong>Story</strong> = one per security category (e.g., &ldquo;[SAT] Network Security Security Findings&rdquo;)</li>
            <li style="margin-bottom:4px"><strong>Task</strong> = one per finding, linked to its category Story via <code>Parent Summary</code></li>
            <li style="margin-bottom:4px">Each Task description includes: check ID, severity, status, current state, recommendation, why it matters, effort, and reference URL</li>
          </ul>
        </div>
      </details>

      <style>@keyframes spin {{ 0%{{transform:rotate(0deg)}} 100%{{transform:rotate(360deg)}} }}</style>
    </div>
    <div id="panel-ado" class="tab-panel">
      <div style="background:#eff6ff;border-left:4px solid #0078d4;padding:12px 16px;border-radius:0 8px 8px 0;margin-bottom:20px;font-size:13px;color:#374151">
        Configure your Azure DevOps instance, then export findings as CSV for bulk import. Work item hierarchy: Category &rarr; Epic, Finding &rarr; User Story.</div>
      <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:16px;margin-bottom:16px">
        <div>
          <label style="font-size:12px;font-weight:600;color:#475569">Organization URL</label>
          <input id="ado-org" type="text" placeholder="https://dev.azure.com/your-org"
            style="width:100%;padding:8px 10px;border:1px solid #cbd5e1;border-radius:6px;font-size:13px;margin-top:4px">
        </div>
        <div>
          <label style="font-size:12px;font-weight:600;color:#475569">Project Name</label>
          <input id="ado-project" type="text" placeholder="SAT"
            style="width:100%;padding:8px 10px;border:1px solid #cbd5e1;border-radius:6px;font-size:13px;margin-top:4px">
        </div>
        <div>
          <label style="font-size:12px;font-weight:600;color:#475569">Area Path <span style="color:#94a3b8;font-weight:400">(optional)</span></label>
          <input id="ado-area" type="text" placeholder="Project\\Team"
            style="width:100%;padding:8px 10px;border:1px solid #cbd5e1;border-radius:6px;font-size:13px;margin-top:4px">
        </div>
      </div>
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:16px">
        <div>
          <label style="font-size:12px;font-weight:600;color:#475569">Parent Work Item Type <span style="color:#94a3b8;font-weight:400">(default: Epic)</span></label>
          <input id="ado-parent-type" type="text" placeholder="Epic"
            style="width:100%;padding:8px 10px;border:1px solid #cbd5e1;border-radius:6px;font-size:13px;margin-top:4px">
        </div>
        <div>
          <label style="font-size:12px;font-weight:600;color:#475569">Child Work Item Type <span style="color:#94a3b8;font-weight:400">(default: User Story)</span></label>
          <input id="ado-child-type" type="text" placeholder="User Story"
            style="width:100%;padding:8px 10px;border:1px solid #cbd5e1;border-radius:6px;font-size:13px;margin-top:4px">
        </div>
      </div>
      <div style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:20px">
        <label style="font-size:12px;font-weight:600;color:#475569;width:100%">Priority Filter</label>
        <select id="ado-filter" style="padding:6px 10px;border:1px solid #cbd5e1;border-radius:6px;font-size:13px">
          <option value="ALL">All Priorities</option>
          <option value="P1">P1 Only</option><option value="P2">P2 Only</option>
          <option value="P3">P3 Only</option><option value="P4">P4 Only</option>
        </select>
      </div>
      <div style="display:flex;gap:12px;margin-bottom:20px">
        <button id="ado-csv-btn" style="padding:10px 20px;background:#0078d4;color:#fff;border:none;border-radius:6px;cursor:pointer;font-size:13px;font-weight:600">
          <i data-lucide="download" style="width:14px;height:14px;vertical-align:middle;margin-right:4px"></i>Download CSV</button>
      </div>
      <div id="ado-status" style="display:none;padding:10px 16px;border-radius:6px;font-size:13px;margin-bottom:12px"></div>
    </div>
    <div class="footer">Generated by SAT Scanner CLI &middot; {_esc(date_str)}</div>
  </div>
</div>
<script>
lucide.createIcons();
{_mermaid_init}
/* ── Tab navigation ── */
document.querySelectorAll('.tab-btn').forEach(function(btn) {{
  btn.addEventListener('click', function() {{
    document.querySelectorAll('.tab-btn').forEach(function(b) {{ b.classList.remove('active'); }});
    document.querySelectorAll('.tab-panel').forEach(function(p) {{ p.style.display = 'none'; }});
    btn.classList.add('active');
    var panel = document.getElementById('panel-' + btn.getAttribute('data-tab'));
    if (panel) {{
      panel.style.display = 'block';
      if (btn.getAttribute('data-tab') === 'architecture' && !panel.dataset.rendered) {{
        mermaid.run({{ nodes: panel.querySelectorAll('.mermaid') }});
        panel.dataset.rendered = '1';
      }}
    }}
  }});
}});

/* ── Jira helpers ── */
(function() {{
  var urlEl = document.getElementById('jira-url');
  var projEl = document.getElementById('jira-project');
  var pidEl = document.getElementById('jira-pid');
  var storyTypeEl = document.getElementById('jira-story-type');
  var taskTypeEl = document.getElementById('jira-task-type');
  var emailEl = document.getElementById('jira-email');
  var tokenEl = document.getElementById('jira-token');
  if (!urlEl) return;

  /* Restore saved config (token NOT stored) */
  urlEl.value = localStorage.getItem('sat_jira_url') || '';
  projEl.value = localStorage.getItem('sat_jira_project') || '';
  pidEl.value = localStorage.getItem('sat_jira_pid') || '';
  storyTypeEl.value = localStorage.getItem('sat_jira_story_type') || '';
  taskTypeEl.value = localStorage.getItem('sat_jira_task_type') || '';
  emailEl.value = localStorage.getItem('sat_jira_email') || '';
  urlEl.addEventListener('change', function() {{ localStorage.setItem('sat_jira_url', urlEl.value); _toggleBtns(); }});
  projEl.addEventListener('change', function() {{ localStorage.setItem('sat_jira_project', projEl.value); _toggleBtns(); }});
  pidEl.addEventListener('change', function() {{ localStorage.setItem('sat_jira_pid', pidEl.value); _toggleBtns(); }});
  storyTypeEl.addEventListener('change', function() {{ localStorage.setItem('sat_jira_story_type', storyTypeEl.value); }});
  taskTypeEl.addEventListener('change', function() {{ localStorage.setItem('sat_jira_task_type', taskTypeEl.value); }});
  emailEl.addEventListener('change', function() {{ localStorage.setItem('sat_jira_email', emailEl.value); }});

  /* Find ID button — opens Jira REST API to show project details */
  document.getElementById('jira-find-id').addEventListener('click', function() {{
    var base = urlEl.value.trim().replace(/\\/+$/, '');
    var key = projEl.value.trim();
    if (!base || !key) {{ alert('Enter Jira Base URL and Project Key first.'); return; }}
    window.open(base + '/rest/api/2/project/' + encodeURIComponent(key), '_blank');
  }});

  function _toggleBtns() {{
    var show = urlEl.value.trim() && projEl.value.trim();
    document.querySelectorAll('.jira-story-btn,.jira-task-btn').forEach(function(b) {{
      b.style.display = show ? 'inline-block' : 'none';
    }});
  }}
  _toggleBtns();

  /* ── Collect findings from DOM ── */
  function _collectFindings(filter) {{
    var items = [];
    document.querySelectorAll('[data-jira]').forEach(function(el) {{
      var d = JSON.parse(el.getAttribute('data-jira'));
      if (filter === 'ALL' || d.priority_label.indexOf(filter) === 0) items.push(d);
    }});
    return items;
  }}

  function _groupByCategory(items) {{
    var groups = {{}};
    items.forEach(function(d) {{
      if (!groups[d.category]) groups[d.category] = [];
      groups[d.category].push(d);
    }});
    return groups;
  }}

  /* ── CSV cell helper ── */
  function _csvCell(v) {{
    var s = String(v || '').replace(/\u2013/g, '-').replace(/\u2014/g, '--').replace(/\u2018|\u2019/g, "'").replace(/\u201c|\u201d/g, '"');
    return '"' + s.replace(/"/g, '""') + '"';
  }}

  function _csvTaskDesc(d) {{
    var lines = [];
    lines.push('Check: ' + d.check_id + ' | Severity: ' + d.severity.toUpperCase() + ' | Status: ' + d.status);
    if (d.current_state) lines.push('\\nCurrent State:\\n' + d.current_state);
    if (d.recommendation) lines.push('\\nRecommendation:\\n' + d.recommendation);
    if (d.benefits) lines.push('\\nWhy It Matters:\\n' + d.benefits);
    if (d.effort) lines.push('\\nEffort: ' + d.effort);
    if (d.reference_url) lines.push('\\nReference: ' + d.reference_url);
    return lines.join('\\n');
  }}

  function _csvStoryDesc(cat, items) {{
    var lines = ['Security findings for category: ' + cat, '\\nFindings (' + items.length + '):'];
    items.forEach(function(d, i) {{
      lines.push((i + 1) + '. [' + d.check_id + '] ' + d.title + ' (' + d.severity.toUpperCase() + '/' + d.status + ')');
    }});
    return lines.join('\\n');
  }}

  function _jiraPriority(d) {{
    var p = d.priority_label || '';
    if (p.indexOf('P1') === 0) return 'Highest';
    if (p.indexOf('P2') === 0) return 'High';
    if (p.indexOf('P3') === 0) return 'Medium';
    return 'Low';
  }}

  /* ── CSV export ── */
  document.getElementById('jira-csv-btn').addEventListener('click', function() {{
    var filter = document.getElementById('jira-filter').value;
    var items = _collectFindings(filter);
    if (!items.length) {{ alert('No findings for selected filter.'); return; }}
    var groups = _groupByCategory(items);
    var rows = [];
    rows.push(['Issue Type', 'Parent Summary', 'Summary', 'Priority', 'Description', 'Labels'].join(','));

    Object.keys(groups).forEach(function(cat) {{
      var catItems = groups[cat];
      var storySummary = '[SAT] ' + cat + ' Security Findings';
      /* Story row */
      rows.push([
        _csvCell('Story'), _csvCell(''), _csvCell(storySummary),
        _csvCell(_jiraPriority(catItems[0])),
        _csvCell(_csvStoryDesc(cat, catItems)),
        _csvCell('sat-scanner')
      ].join(','));
      /* Task rows */
      catItems.forEach(function(d) {{
        rows.push([
          _csvCell('Task'), _csvCell(storySummary),
          _csvCell('[SAT] ' + d.check_id + ' - ' + d.title),
          _csvCell(_jiraPriority(d)),
          _csvCell(_csvTaskDesc(d)),
          _csvCell('sat-scanner')
        ].join(','));
      }});
    }});

    var csv = '\\uFEFF' + rows.join('\\r\\n');
    var blob = new Blob([csv], {{type: 'text/csv;charset=utf-8'}});
    var a = document.createElement('a');
    a.href = URL.createObjectURL(blob);
    a.download = 'sat-jira-import-' + filter.toLowerCase() + '.csv';
    a.click();
  }});

  /* ── Shell script export ── */
  document.getElementById('jira-sh-btn').addEventListener('click', function() {{
    var base = urlEl.value.trim().replace(/\\/+$/, '');
    var proj = projEl.value.trim();
    if (!base || !proj) {{ alert('Please enter Jira URL and Project Key first.'); return; }}
    var filter = document.getElementById('jira-filter').value;
    var items = _collectFindings(filter);
    if (!items.length) {{ alert('No findings for selected filter.'); return; }}
    var groups = _groupByCategory(items);

    var lines = [];
    lines.push('#!/usr/bin/env bash');
    lines.push('# SAT Scanner - Jira Issue Creator');
    lines.push('# Generated: ' + new Date().toISOString().slice(0, 10));
    lines.push('# Usage: JIRA_TOKEN=your-api-token JIRA_EMAIL=you@company.com bash ' + 'sat-jira-create.sh');
    lines.push('set -euo pipefail');
    lines.push('');
    lines.push('JIRA_URL="' + base + '"');
    lines.push('PROJECT="' + proj + '"');
    lines.push('');
    lines.push('create_issue() ' + '{{');
    lines.push('  local itype="$1" summary="$2" desc="$3" priority="$4" parent_key="$' + '{{5:-}}"');
    lines.push('  local payload');
    lines.push('  if [ -n "$parent_key" ]; then');
    lines.push('    payload=$(jq -n --arg p "$PROJECT" --arg t "$itype" --arg s "$summary" --arg d "$desc" --arg pr "$priority" --arg pk "$parent_key" \\\\');
    lines.push('      ' + "'" + '{{fields:{{project:{{key:$p}},issuetype:{{name:$t}},summary:$s,description:$d,priority:{{name:$pr}},parent:{{key:$pk}}}}}}' + "'" + ')');
    lines.push('  else');
    lines.push('    payload=$(jq -n --arg p "$PROJECT" --arg t "$itype" --arg s "$summary" --arg d "$desc" --arg pr "$priority" \\\\');
    lines.push('      ' + "'" + '{{fields:{{project:{{key:$p}},issuetype:{{name:$t}},summary:$s,description:$d,priority:{{name:$pr}}}}}}' + "'" + ')');
    lines.push('  fi');
    lines.push('  curl -s -X POST "$JIRA_URL/rest/api/2/issue" \\\\');
    lines.push('    -H "Content-Type: application/json" \\\\');
    lines.push('    -u "$JIRA_EMAIL:$JIRA_TOKEN" \\\\');
    lines.push('    -d "$payload" | jq -r ' + "'" + '.key // "ERROR"' + "'" + '');
    lines.push('}}');
    lines.push('');

    Object.keys(groups).forEach(function(cat) {{
      var catItems = groups[cat];
      var storySum = '[SAT] ' + cat + ' Security Findings';
      var storyDesc = 'Security findings for category: ' + cat + '\\\\n\\\\nFindings (' + catItems.length + '):\\\\n';
      catItems.forEach(function(d, i) {{
        storyDesc += (i + 1) + '. [' + d.check_id + '] ' + d.title + '\\\\n';
      }});
      var varName = cat.replace(/[^a-zA-Z0-9]/g, '_').toUpperCase();
      lines.push('echo "Creating Story: ' + cat + '"');
      lines.push(varName + '=$(create_issue "Story" "' + storySum.replace(/"/g, '\\\\"') + '" "' + storyDesc.replace(/"/g, '\\\\"') + '" "' + _jiraPriority(catItems[0]) + '")');
      lines.push('echo "  Story: $' + varName + '"');
      lines.push('');

      catItems.forEach(function(d) {{
        var taskSum = '[SAT] ' + d.check_id + ' - ' + d.title;
        var taskDesc = 'Check: ' + d.check_id + '\\\\nSeverity: ' + d.severity.toUpperCase() + '\\\\nStatus: ' + d.status;
        if (d.current_state) taskDesc += '\\\\n\\\\nCurrent State:\\\\n' + d.current_state.replace(/"/g, '\\\\"').replace(/\\n/g, '\\\\n');
        if (d.recommendation) taskDesc += '\\\\n\\\\nRecommendation:\\\\n' + d.recommendation.replace(/"/g, '\\\\"').replace(/\\n/g, '\\\\n');
        if (d.benefits) taskDesc += '\\\\n\\\\nWhy It Matters:\\\\n' + d.benefits.replace(/"/g, '\\\\"').replace(/\\n/g, '\\\\n');
        if (d.effort) taskDesc += '\\\\n\\\\nEffort: ' + d.effort;
        if (d.reference_url) taskDesc += '\\\\nReference: ' + d.reference_url;
        lines.push('echo "  Creating Task: ' + d.check_id + '"');
        lines.push('create_issue "Task" "' + taskSum.replace(/"/g, '\\\\"') + '" "' + taskDesc + '" "' + _jiraPriority(d) + '" "$' + varName + '"');
      }});
      lines.push('');
    }});

    lines.push('echo "Done! All issues created."');
    var blob = new Blob([lines.join('\\n')], {{type: 'text/x-shellscript'}});
    var a = document.createElement('a');
    a.href = URL.createObjectURL(blob);
    a.download = 'sat-jira-create.sh';
    a.click();
  }});

  /* ── Export to Jira (REST API) ── */
  /* Build a self-contained JS snippet for browser console (same-origin, no CORS).
     Uses JSON.stringify for all data to avoid escaping issues. */
  function _buildConsoleScript(proj, storyType, taskType, groups, email, token) {{
    /* Build a plain data structure, then JSON-encode it into the script */
    var plan = [];
    var catKeys = Object.keys(groups);
    for (var i = 0; i < catKeys.length; i++) {{
      var cat = catKeys[i];
      var catItems = groups[cat];
      var storyDesc = 'Security findings for: ' + cat + '\\n' + catItems.length + ' findings';
      var tasks = [];
      for (var j = 0; j < catItems.length; j++) {{
        var d = catItems[j];
        var tDesc = 'Check: ' + d.check_id + '\\nSeverity: ' + d.severity.toUpperCase() + '\\nStatus: ' + d.status;
        if (d.current_state) tDesc += '\\n\\nCurrent State:\\n' + d.current_state;
        if (d.recommendation) tDesc += '\\n\\nRecommendation:\\n' + d.recommendation;
        if (d.benefits) tDesc += '\\n\\nWhy It Matters:\\n' + d.benefits;
        if (d.effort) tDesc += '\\n\\nEffort: ' + d.effort;
        if (d.reference_url) tDesc += '\\nReference: ' + d.reference_url;
        tasks.push({{ summary: '[SAT] ' + d.check_id + ' - ' + d.title, description: tDesc }});
      }}
      plan.push({{ storySummary: '[SAT] ' + cat + ' Security Findings', storyDesc: storyDesc, tasks: tasks }});
    }}

    var dataJson = JSON.stringify(plan);

    return '(async function() {{\\n'
      + 'var AUTH = "Basic " + btoa(' + JSON.stringify(email + ':' + token) + ');\\n'
      + 'var PROJECT = ' + JSON.stringify(proj) + ';\\n'
      + 'var STORY_TYPE = ' + JSON.stringify(storyType) + ';\\n'
      + 'var TASK_TYPE = ' + JSON.stringify(taskType) + ';\\n'
      + 'var PLAN = ' + dataJson + ';\\n'
      + 'var created = 0, errors = 0;\\n'
      + 'async function mk(type, summary, desc, parent) {{\\n'
      + '  var f = {{project:{{key:PROJECT}},issuetype:{{name:type}},summary:summary,description:desc}};\\n'
      + '  if (parent) f.parent = {{key:parent}};\\n'
      + '  try {{\\n'
      + '    var r = await fetch("/rest/api/2/issue", {{method:"POST",headers:{{"Content-Type":"application/json","Authorization":AUTH}},body:JSON.stringify({{fields:f}})}});\\n'
      + '    var d = await r.json();\\n'
      + '    if (r.ok && d.key) {{ created++; console.log("OK " + d.key + " " + summary); return d.key; }}\\n'
      + '    errors++; console.error("FAIL " + summary, d); return null;\\n'
      + '  }} catch(e) {{ errors++; console.error("ERR " + summary, e); return null; }}\\n'
      + '}}\\n'
      + 'for (var i = 0; i < PLAN.length; i++) {{\\n'
      + '  var g = PLAN[i];\\n'
      + '  console.log("Story: " + g.storySummary);\\n'
      + '  var sk = await mk(STORY_TYPE, g.storySummary, g.storyDesc, null);\\n'
      + '  if (!sk) continue;\\n'
      + '  for (var j = 0; j < g.tasks.length; j++) {{\\n'
      + '    var t = g.tasks[j];\\n'
      + '    await mk(TASK_TYPE, t.summary, t.description, sk);\\n'
      + '  }}\\n'
      + '}}\\n'
      + 'console.log("Done! Created: " + created + ", Errors: " + errors);\\n'
      + 'alert("Done! " + created + " issues created, " + errors + " errors.");\\n'
      + '}})();';
  }}

  document.getElementById('jira-push-btn').addEventListener('click', function() {{
    var base = urlEl.value.trim().replace(/\\/+$/, '');
    var proj = projEl.value.trim();
    var email = emailEl.value.trim();
    var token = tokenEl.value.trim();
    var storyType = storyTypeEl.value.trim() || 'Story';
    var taskType = taskTypeEl.value.trim() || 'Subtask';
    if (!base || !proj) {{ alert('Enter Jira Base URL and Project Key.'); return; }}
    if (!email || !token) {{ alert('Enter your Jira Email and API Token.'); return; }}

    var filter = document.getElementById('jira-filter').value;
    var items = _collectFindings(filter);
    if (!items.length) {{ alert('No findings for selected filter.'); return; }}
    var groups = _groupByCategory(items);
    var cats = Object.keys(groups);
    var totalTasks = 0;
    cats.forEach(function(c) {{ totalTasks += groups[c].length; }});

    /* Build console script */
    var script = _buildConsoleScript(proj, storyType, taskType, groups, email, token);

    /* Copy to clipboard */
    navigator.clipboard.writeText(script).then(function() {{
      /* Open Jira in new tab */
      window.open(base, '_blank');

      /* Show success panel */
      var prog = document.getElementById('jira-progress');
      var progLog = document.getElementById('jira-prog-log');
      prog.style.display = 'block';
      progLog.innerHTML =
        '<div style="background:#f0fdf4;border:1px solid #bbf7d0;border-radius:8px;padding:16px;margin-bottom:12px">'
        + '<div style="font-size:15px;font-weight:700;color:#16a34a;margin-bottom:8px">&#10003; Script copied to clipboard!</div>'
        + '<div style="font-size:13px;color:#374151;line-height:1.8">'
        + '<strong>The script will create ' + cats.length + ' Stories and ' + totalTasks + ' Tasks in project ' + proj + '.</strong><br><br>'
        + 'Complete these steps in the Jira tab that just opened:<br>'
        + '<span style="display:inline-block;background:#7c3aed;color:#fff;border-radius:50%;width:22px;height:22px;text-align:center;line-height:22px;font-size:12px;font-weight:700;margin-right:6px">1</span>'
        + 'Press <kbd style="background:#1e293b;color:#e2e8f0;padding:2px 8px;border-radius:4px;font-size:12px;font-weight:600">F12</kbd> to open Developer Tools<br>'
        + '<span style="display:inline-block;background:#7c3aed;color:#fff;border-radius:50%;width:22px;height:22px;text-align:center;line-height:22px;font-size:12px;font-weight:700;margin-right:6px">2</span>'
        + 'Click the <strong>Console</strong> tab<br>'
        + '<span style="display:inline-block;background:#7c3aed;color:#fff;border-radius:50%;width:22px;height:22px;text-align:center;line-height:22px;font-size:12px;font-weight:700;margin-right:6px">3</span>'
        + 'Press <kbd style="background:#1e293b;color:#e2e8f0;padding:2px 8px;border-radius:4px;font-size:12px;font-weight:600">Ctrl+V</kbd> (or <kbd style="background:#1e293b;color:#e2e8f0;padding:2px 8px;border-radius:4px;font-size:12px;font-weight:600">Cmd+V</kbd>) to paste<br>'
        + '<span style="display:inline-block;background:#7c3aed;color:#fff;border-radius:50%;width:22px;height:22px;text-align:center;line-height:22px;font-size:12px;font-weight:700;margin-right:6px">4</span>'
        + 'Press <kbd style="background:#1e293b;color:#e2e8f0;padding:2px 8px;border-radius:4px;font-size:12px;font-weight:600">Enter</kbd> to run &mdash; issues will be created automatically'
        + '</div></div>'
        + '<button id="jira-recopy" style="padding:8px 16px;background:#7c3aed;color:#fff;border:none;border-radius:6px;cursor:pointer;font-size:12px;font-weight:600;margin-right:8px">Copy Script Again</button>'
        + '<button id="jira-reopen" style="padding:8px 16px;background:#3b82f6;color:#fff;border:none;border-radius:6px;cursor:pointer;font-size:12px;font-weight:600">Open Jira Again</button>';

      /* Wire up re-copy and re-open buttons */
      window._satConsoleScript = script;
      setTimeout(function() {{
        var recopy = document.getElementById('jira-recopy');
        var reopen = document.getElementById('jira-reopen');
        if (recopy) recopy.addEventListener('click', function() {{
          navigator.clipboard.writeText(window._satConsoleScript).then(function() {{
            recopy.textContent = 'Copied!';
            recopy.style.background = '#16a34a';
            setTimeout(function() {{ recopy.textContent = 'Copy Script Again'; recopy.style.background = '#7c3aed'; }}, 2000);
          }});
        }});
        if (reopen) reopen.addEventListener('click', function() {{
          window.open(base, '_blank');
        }});
      }}, 100);
    }}).catch(function() {{
      alert('Failed to copy to clipboard. Your browser may require HTTPS for clipboard access.');
    }});
  }});

  /* ── Individual Copy Story/Task buttons ── */
  function _jiraWikiDesc(d) {{
    var t = 'h3. ' + d.check_id + ' - ' + d.title + '\\n';
    t += '||Property||Value||\\n';
    t += '|Severity|' + d.severity.toUpperCase() + '|\\n';
    t += '|Status|' + d.status + '|\\n';
    if (d.effort) t += '|Effort|' + d.effort + '|\\n';
    if (d.current_state) t += '\\nh4. Current State\\n' + d.current_state + '\\n';
    if (d.recommendation) t += '\\nh4. Recommendation\\n' + d.recommendation + '\\n';
    if (d.benefits) t += '\\nh4. Why It Matters\\n' + d.benefits + '\\n';
    if (d.reference_url) t += '\\n[Reference|' + d.reference_url + ']';
    return t;
  }}

  /* Copy to clipboard + show toast notification */
  function _copyAndOpen(summary, description, openCreate) {{
    var text = 'Summary:\\n' + summary + '\\n\\nDescription:\\n' + description;
    navigator.clipboard.writeText(text).then(function() {{
      _showToast('Copied to clipboard! Paste into Jira.');
    }}).catch(function() {{
      /* Fallback for older browsers */
      var ta = document.createElement('textarea');
      ta.value = text; ta.style.position = 'fixed'; ta.style.opacity = '0';
      document.body.appendChild(ta); ta.select(); document.execCommand('copy');
      document.body.removeChild(ta);
      _showToast('Copied to clipboard! Paste into Jira.');
    }});
    if (openCreate) {{
      var base = urlEl.value.trim().replace(/\\/+$/, '');
      var proj = projEl.value.trim();
      window.open(base + '/jira/software/projects/' + encodeURIComponent(proj) + '/board', '_blank');
    }}
  }}

  function _showToast(msg) {{
    var t = document.getElementById('jira-toast');
    if (!t) {{
      t = document.createElement('div');
      t.id = 'jira-toast';
      t.style.cssText = 'position:fixed;bottom:24px;right:24px;background:#16a34a;color:#fff;padding:12px 20px;border-radius:8px;font-size:13px;font-weight:600;z-index:9999;box-shadow:0 4px 16px rgba(0,0,0,.2);transition:opacity .3s;';
      document.body.appendChild(t);
    }}
    t.textContent = msg;
    t.style.opacity = '1';
    setTimeout(function() {{ t.style.opacity = '0'; }}, 3000);
  }}

  document.querySelectorAll('.jira-task-btn').forEach(function(btn) {{
    btn.addEventListener('click', function(e) {{
      e.preventDefault(); e.stopPropagation();
      if (!urlEl.value.trim() || !projEl.value.trim()) {{ alert('Configure Jira URL and Project Key in the Jira Export tab first.'); return; }}
      var detail = btn.closest('[data-jira]');
      if (!detail) return;
      var d = JSON.parse(detail.getAttribute('data-jira'));
      var summary = '[SAT] ' + d.check_id + ' - ' + d.title;
      _copyAndOpen(summary, _jiraWikiDesc(d), false);
    }});
  }});

  document.querySelectorAll('.jira-story-btn').forEach(function(btn) {{
    btn.addEventListener('click', function(e) {{
      e.preventDefault();
      if (!urlEl.value.trim() || !projEl.value.trim()) {{ alert('Configure Jira URL and Project Key in the Jira Export tab first.'); return; }}
      var cat = btn.getAttribute('data-category');
      var items = _collectFindings('ALL').filter(function(d) {{ return d.category === cat; }});
      var summary = '[SAT] ' + cat + ' Security Findings';
      var desc = cat + ' Security Findings\\n\\n' + items.length + ' findings:\\n';
      items.forEach(function(d, i) {{
        desc += (i + 1) + '. [' + d.check_id + '] ' + d.title + ' (' + d.severity.toUpperCase() + '/' + d.status + ')\\n';
      }});
      _copyAndOpen(summary, desc, false);
    }});
  }});
}})();

/* ── ADO helpers ── */
(function() {{
  var orgEl = document.getElementById('ado-org');
  if (!orgEl) return;
  var projEl = document.getElementById('ado-project');
  var areaEl = document.getElementById('ado-area');
  var parentTypeEl = document.getElementById('ado-parent-type');
  var childTypeEl = document.getElementById('ado-child-type');
  var filterEl = document.getElementById('ado-filter');
  var statusEl = document.getElementById('ado-status');

  /* Restore saved config */
  orgEl.value = localStorage.getItem('sat_ado_org') || '';
  projEl.value = localStorage.getItem('sat_ado_project') || '';
  areaEl.value = localStorage.getItem('sat_ado_area') || '';
  parentTypeEl.value = localStorage.getItem('sat_ado_parent_type') || '';
  childTypeEl.value = localStorage.getItem('sat_ado_child_type') || '';
  orgEl.addEventListener('change', function() {{ localStorage.setItem('sat_ado_org', orgEl.value); }});
  projEl.addEventListener('change', function() {{ localStorage.setItem('sat_ado_project', projEl.value); }});
  areaEl.addEventListener('change', function() {{ localStorage.setItem('sat_ado_area', areaEl.value); }});
  parentTypeEl.addEventListener('change', function() {{ localStorage.setItem('sat_ado_parent_type', parentTypeEl.value); }});
  childTypeEl.addEventListener('change', function() {{ localStorage.setItem('sat_ado_child_type', childTypeEl.value); }});

  function _adoCsvCell(v) {{
    var s = String(v || '').replace(/\u2013/g, '-').replace(/\u2014/g, '--').replace(/\u2018|\u2019/g, "'").replace(/\u201c|\u201d/g, '"');
    return '"' + s.replace(/"/g, '""') + '"';
  }}

  function _adoSeverity(sev) {{
    var m = {{critical: '1 - Critical', high: '2 - High', medium: '3 - Medium', low: '4 - Low'}};
    return m[(sev || '').toLowerCase()] || '3 - Medium';
  }}

  function _adoDesc(d) {{
    var h = '<h3>' + (d.check_id || '') + ' - ' + (d.title || '') + '</h3>';
    h += '<table><tr><td><b>Severity</b></td><td>' + (d.severity || '').toUpperCase() + '</td></tr>';
    h += '<tr><td><b>Status</b></td><td>' + (d.status || '') + '</td></tr>';
    if (d.effort) h += '<tr><td><b>Effort</b></td><td>' + d.effort + '</td></tr>';
    h += '</table>';
    if (d.current_state) h += '<h4>Current State</h4><p>' + d.current_state + '</p>';
    if (d.recommendation) h += '<h4>Recommendation</h4><p>' + d.recommendation + '</p>';
    if (d.benefits) h += '<h4>Why It Matters</h4><p>' + d.benefits + '</p>';
    if (d.reference_url) h += '<p><a href="' + d.reference_url + '">Reference</a></p>';
    return h;
  }}

  /* CSV download for ADO bulk import */
  document.getElementById('ado-csv-btn').addEventListener('click', function() {{
    var filter = filterEl.value;
    var items = [];
    document.querySelectorAll('[data-jira]').forEach(function(el) {{
      var d = JSON.parse(el.getAttribute('data-jira'));
      if (filter === 'ALL' || d.priority_label.indexOf(filter) === 0) items.push(d);
    }});
    if (!items.length) {{ alert('No findings for selected filter.'); return; }}

    var parentType = parentTypeEl.value.trim() || 'Epic';
    var childType = childTypeEl.value.trim() || 'User Story';
    var area = areaEl.value.trim();

    /* Group by category */
    var groups = {{}};
    items.forEach(function(d) {{
      if (!groups[d.category]) groups[d.category] = [];
      groups[d.category].push(d);
    }});

    var rows = [];
    rows.push(['Work Item Type', 'Title 1', 'Title 2', 'State', 'Assigned To', 'Area Path', 'Priority', 'Severity', 'Description', 'Tags'].join(','));

    Object.keys(groups).forEach(function(cat) {{
      var catItems = groups[cat];
      /* Parent (Epic) row */
      var epicDesc = '<h3>' + cat + ' Security Findings</h3><p>' + catItems.length + ' findings in this category.</p>';
      rows.push([
        _adoCsvCell(parentType), _adoCsvCell('[SAT] ' + cat + ' Security Findings'),
        _adoCsvCell(''), _adoCsvCell('New'), _adoCsvCell(''),
        _adoCsvCell(area), _adoCsvCell('2'), _adoCsvCell(_adoSeverity(catItems[0].severity)),
        _adoCsvCell(epicDesc), _adoCsvCell('sat-scanner')
      ].join(','));
      /* Child (User Story) rows */
      catItems.forEach(function(d) {{
        rows.push([
          _adoCsvCell(childType), _adoCsvCell('[SAT] ' + cat + ' Security Findings'),
          _adoCsvCell('[SAT] ' + d.check_id + ' - ' + d.title),
          _adoCsvCell('New'), _adoCsvCell(''),
          _adoCsvCell(area), _adoCsvCell(d.priority_label.indexOf('P1') === 0 ? '1' : d.priority_label.indexOf('P2') === 0 ? '2' : d.priority_label.indexOf('P3') === 0 ? '3' : '4'),
          _adoCsvCell(_adoSeverity(d.severity)),
          _adoCsvCell(_adoDesc(d)), _adoCsvCell('sat-scanner')
        ].join(','));
      }});
    }});

    var csv = '\\uFEFF' + rows.join('\\r\\n');
    var blob = new Blob([csv], {{type: 'text/csv;charset=utf-8'}});
    var a = document.createElement('a');
    a.href = URL.createObjectURL(blob);
    a.download = 'sat-ado-import-' + filter.toLowerCase() + '.csv';
    a.click();

    statusEl.style.display = 'block';
    statusEl.style.background = '#f0fdf4';
    statusEl.style.border = '1px solid #bbf7d0';
    statusEl.style.color = '#16a34a';
    statusEl.innerHTML = '<strong>&#10003; CSV downloaded!</strong> Import it into Azure DevOps via <em>Boards &rarr; Work Items &rarr; Import Work Items</em>.';
  }});
}})();

/* ── Excel Export ── */
document.getElementById('excel-export-btn').addEventListener('click', function() {{
  if (typeof XLSX === 'undefined') {{ alert('SheetJS library not loaded. Check your internet connection.'); return; }}

  /* Collect all findings from DOM */
  var items = [];
  document.querySelectorAll('[data-jira]').forEach(function(el) {{
    items.push(JSON.parse(el.getAttribute('data-jira')));
  }});
  if (!items.length) {{ alert('No findings to export.'); return; }}

  var wb = XLSX.utils.book_new();

  /* ── Summary sheet ── */
  var summaryData = [
    ['SAT Recommendations Summary'],
    ['Generated', '{_esc(now_str)}'],
    ['Total Findings', items.length],
    ['Failures', items.filter(function(d) {{ return d.status === 'FAIL'; }}).length],
    ['Warnings', items.filter(function(d) {{ return d.status === 'WARN'; }}).length],
    [],
    ['Priority', 'Count'],
    ['P1 - Fix Immediately', items.filter(function(d) {{ return d.priority_label.indexOf('P1') === 0; }}).length],
    ['P2 - Fix This Sprint', items.filter(function(d) {{ return d.priority_label.indexOf('P2') === 0; }}).length],
    ['P3 - Plan Next Sprint', items.filter(function(d) {{ return d.priority_label.indexOf('P3') === 0; }}).length],
    ['P4 - Backlog', items.filter(function(d) {{ return d.priority_label.indexOf('P4') === 0; }}).length],
    [],
    ['Category', 'Fail', 'Warn', 'Total']
  ];
  var catCounts = {{}};
  items.forEach(function(d) {{
    if (!catCounts[d.category]) catCounts[d.category] = {{fail: 0, warn: 0}};
    if (d.status === 'FAIL') catCounts[d.category].fail++;
    else catCounts[d.category].warn++;
  }});
  Object.keys(catCounts).sort(function(a, b) {{
    return (catCounts[b].fail + catCounts[b].warn) - (catCounts[a].fail + catCounts[a].warn);
  }}).forEach(function(cat) {{
    var c = catCounts[cat];
    summaryData.push([cat, c.fail, c.warn, c.fail + c.warn]);
  }});
  var ws0 = XLSX.utils.aoa_to_sheet(summaryData);
  ws0['!cols'] = [{{wch: 30}}, {{wch: 12}}, {{wch: 12}}, {{wch: 12}}];
  XLSX.utils.book_append_sheet(wb, ws0, 'Summary');

  /* ── All Findings sheet ── */
  var header = ['Priority', 'Category', 'Check ID', 'Severity', 'Status', 'Effort', 'Title', 'Current State', 'Recommendation', 'Why It Matters', 'Reference URL'];
  var allRows = [header];
  items.forEach(function(d) {{
    allRows.push([
      d.priority_label || '', d.category || '', d.check_id || '',
      (d.severity || '').toUpperCase(), d.status || '', d.effort || '',
      d.title || '', d.current_state || '', d.recommendation || '',
      d.benefits || '', d.reference_url || ''
    ]);
  }});
  var ws1 = XLSX.utils.aoa_to_sheet(allRows);
  ws1['!cols'] = [{{wch:12}},{{wch:20}},{{wch:18}},{{wch:10}},{{wch:8}},{{wch:20}},{{wch:40}},{{wch:40}},{{wch:50}},{{wch:40}},{{wch:30}}];
  XLSX.utils.book_append_sheet(wb, ws1, 'All Findings');

  /* ── Per-priority sheets ── */
  ['P1', 'P2', 'P3', 'P4'].forEach(function(px) {{
    var pItems = items.filter(function(d) {{ return d.priority_label.indexOf(px) === 0; }});
    if (!pItems.length) return;
    var rows = [header];
    pItems.forEach(function(d) {{
      rows.push([
        d.priority_label || '', d.category || '', d.check_id || '',
        (d.severity || '').toUpperCase(), d.status || '', d.effort || '',
        d.title || '', d.current_state || '', d.recommendation || '',
        d.benefits || '', d.reference_url || ''
      ]);
    }});
    var ws = XLSX.utils.aoa_to_sheet(rows);
    ws['!cols'] = [{{wch:12}},{{wch:20}},{{wch:18}},{{wch:10}},{{wch:8}},{{wch:20}},{{wch:40}},{{wch:40}},{{wch:50}},{{wch:40}},{{wch:30}}];
    XLSX.utils.book_append_sheet(wb, ws, px + ' Findings');
  }});

  /* ── Remediation Plan sheet ── */
  var rpHeader = ['Check ID', 'Title', 'Category', 'Severity', 'Priority', 'Prerequisites', 'Pre-Checks', 'Steps', 'Post-Validation', 'Rollback', 'Downtime', 'Blast Radius', 'Est. Hours', 'Working Days', 'Stakeholders', 'Change Type', 'Approval'];
  var rpRows = [rpHeader];
  items.forEach(function(d) {{
    var plan = d.remediation_plan || {{}};
    var cl = plan.checklist || {{}};
    var ia = plan.impact_assessment || {{}};
    var cm = plan.change_management || {{}};
    var hrs = plan.estimated_duration_hours || 0;
    rpRows.push([
      d.check_id || '', d.title || '', d.category || '',
      (d.severity || '').toUpperCase(), d.priority_label || '',
      (plan.prerequisites || []).join('; '),
      (cl.pre_checks || []).join('; '),
      (cl.steps || []).join('; '),
      (cl.post_validation || []).join('; '),
      (cl.rollback || []).join('; '),
      ia.downtime || '', ia.blast_radius || '',
      hrs, hrs ? Math.round(hrs / 8 * 10) / 10 : '',
      (plan.stakeholders || []).join('; '),
      cm.change_type || '',
      cm.approval_required ? 'Yes' : 'No'
    ]);
  }});
  var wsRp = XLSX.utils.aoa_to_sheet(rpRows);
  wsRp['!cols'] = [{{wch:18}},{{wch:40}},{{wch:20}},{{wch:10}},{{wch:10}},{{wch:40}},{{wch:40}},{{wch:50}},{{wch:40}},{{wch:40}},{{wch:10}},{{wch:14}},{{wch:10}},{{wch:12}},{{wch:30}},{{wch:12}},{{wch:10}}];
  XLSX.utils.book_append_sheet(wb, wsRp, 'Remediation Plan');

  /* ── Roadmap sheet ── */
  var tl = {timeline_json};
  var ts = tl.summary || {{}};
  var roadmapRows = [
    ['Remediation Roadmap'],
    ['Total Findings', ts.total_findings || 0],
    ['Total Effort (hours)', ts.total_effort_hours || 0],
    ['Total Working Days', ts.total_working_days || 0],
    ['Estimated Resources', ts.estimated_resources || 0],
    ['Categories', ts.categories || 0],
    ['Note', 'Resources = parallel team members needed per phase (8h/day). P1: 1wk, P2: 2wks, P3: 5wks, P4: flexible.'],
    [],
    ['Phase', 'Hours', 'Working Days', 'Est. Resources']
  ];
  (tl.phases || []).forEach(function(phase) {{
    roadmapRows.push([phase.phase, phase.total_hours, phase.total_working_days, phase.estimated_resources]);
  }});
  roadmapRows.push([]);
  roadmapRows.push(['Phase', 'Category', 'Check ID', 'Title', 'Severity', 'Effort', 'Est. Hours', 'Working Days']);
  (tl.phases || []).forEach(function(phase) {{
    var cats = phase.categories || {{}};
    Object.keys(cats).forEach(function(cat) {{
      cats[cat].findings.forEach(function(f) {{
        roadmapRows.push([phase.phase, cat, f.check_id || '', f.title || '', (f.severity || '').toUpperCase(), f.effort || '', f.effort_hours || '', f.working_days || '']);
      }});
    }});
    if (phase.total_hours) {{
      roadmapRows.push([phase.phase + ' — Total', '', '', '', '', '', phase.total_hours, phase.total_working_days]);
    }}
  }});
  var wsRoadmap = XLSX.utils.aoa_to_sheet(roadmapRows);
  wsRoadmap['!cols'] = [{{wch:30}},{{wch:22}},{{wch:18}},{{wch:40}},{{wch:10}},{{wch:22}},{{wch:12}},{{wch:14}}];
  XLSX.utils.book_append_sheet(wb, wsRoadmap, 'Roadmap');

  /* ── Change Management sheet ── */
  var cmHeader = ['Check ID', 'Title', 'Severity', 'Priority', 'Change Type', 'Approval Required', 'Change Window', 'Testing Plan', 'Communication Plan', 'Downtime', 'Blast Radius', 'Rollback Guidance', 'Stakeholders', 'Est. Hours', 'Working Days'];
  var cmRows = [cmHeader];
  items.forEach(function(d) {{
    var plan = d.remediation_plan || {{}};
    var ia = plan.impact_assessment || {{}};
    var cm = plan.change_management || {{}};
    var hrs = plan.estimated_duration_hours || 0;
    cmRows.push([
      d.check_id || '', d.title || '',
      (d.severity || '').toUpperCase(), d.priority_label || '',
      cm.change_type || '', cm.approval_required ? 'Yes' : 'No',
      cm.suggested_change_window || '', cm.testing_plan || '',
      cm.communication_plan || '', ia.downtime || '', ia.blast_radius || '',
      plan.rollback_guidance || '',
      (plan.stakeholders || []).join('; '),
      hrs, hrs ? Math.round(hrs / 8 * 10) / 10 : ''
    ]);
  }});
  var wsCm = XLSX.utils.aoa_to_sheet(cmRows);
  wsCm['!cols'] = [{{wch:18}},{{wch:40}},{{wch:10}},{{wch:10}},{{wch:14}},{{wch:14}},{{wch:22}},{{wch:40}},{{wch:40}},{{wch:10}},{{wch:14}},{{wch:40}},{{wch:30}},{{wch:10}},{{wch:12}}];
  XLSX.utils.book_append_sheet(wb, wsCm, 'Change Management');

  XLSX.writeFile(wb, 'SAT_Recommendations_{date_str}.xlsx');
}});
</script>
</body></html>"""

    path = output_dir / f"Recommendation_Summary.html"
    path.write_text(html, encoding="utf-8")
    return str(path)


# ---------------------------------------------------------------------------
# Jira export via httpx
# ---------------------------------------------------------------------------

_JIRA_PRIO_MAP = {
    "P1": "Highest",
    "P2": "High",
    "P3": "Medium",
    "P4": "Low",
}


def export_jira(
    result: SATScanResult | None,
    *,
    jira_url: str,
    jira_email: str,
    jira_token: str,
    project_key: str,
    parent_type: str = "Story",
    child_type: str = "Subtask",
    findings: list[SATFinding] | None = None,
    dry_run: bool = False,
) -> dict:
    """Create Jira issues from scan findings using httpx.

    Creates one parent issue (Story) per category, then one child issue
    (Subtask) per finding linked to the parent.

    Returns dict with 'created', 'errors', and 'issues' list.
    """
    import httpx

    src = findings if findings is not None else (result.findings if result else [])
    prio_items = _build_prioritised_recommendations(src)
    if not prio_items:
        return {"created": 0, "errors": 0, "issues": []}

    # Group by category
    groups: dict[str, list[dict]] = {}
    for item in prio_items:
        groups.setdefault(item["category"], []).append(item)

    base = jira_url.rstrip("/")
    auth = (jira_email, jira_token)
    headers = {"Content-Type": "application/json"}
    created = 0
    errors = 0
    issues: list[dict] = []

    def _create_issue(client: httpx.Client, itype: str, summary: str,
                      description: str, priority: str,
                      parent_key: str | None = None) -> str | None:
        nonlocal created, errors
        fields: dict[str, Any] = {
            "project": {"key": project_key},
            "issuetype": {"name": itype},
            "summary": summary[:255],
            "description": description,
            "priority": {"name": priority},
        }
        if parent_key:
            fields["parent"] = {"key": parent_key}

        if dry_run:
            created += 1
            key = f"DRY-{created}"
            print(f"  [dry-run] {itype}: {summary[:80]}  (parent={parent_key or 'none'})")
            issues.append({"key": key, "type": itype, "summary": summary})
            return key

        resp = client.post(
            f"{base}/rest/api/2/issue",
            json={"fields": fields},
            headers=headers,
        )
        if resp.status_code in (200, 201):
            key = resp.json().get("key", "???")
            created += 1
            issues.append({"key": key, "type": itype, "summary": summary})
            return key
        else:
            errors += 1
            detail = resp.text[:300]
            print(f"  ERROR creating {itype} '{summary[:60]}': {resp.status_code} {detail}")
            return None

    with httpx.Client(auth=auth, timeout=30.0) as client:
        for cat, items in groups.items():
            # Build story description
            finding_list = "\n".join(
                f"  {i}. [{d['check_id']}] {d['title']}" for i, d in enumerate(items, 1)
            )
            story_desc = (
                f"Security findings for category: {cat}\n\n"
                f"Findings ({len(items)}):\n{finding_list}"
            )
            prio_label = items[0]["priority_label"][:2]
            story_prio = _JIRA_PRIO_MAP.get(prio_label, "Medium")

            print(f"  Creating {parent_type}: {cat} ({len(items)} findings)")
            story_key = _create_issue(
                client, parent_type,
                f"[SAT] {cat} Security Findings",
                story_desc, story_prio,
            )
            if story_key:
                print(f"    -> {story_key}")
            else:
                print(f"    -> FAILED, skipping children")
                continue

            # Create child issues
            for d in items:
                desc_parts = [f"Check: {d['check_id']}", f"Severity: {d['severity'].upper()}", f"Status: {d['status']}"]
                if d.get("current_state"):
                    desc_parts.append(f"\nCurrent State:\n{d['current_state']}")
                if d.get("recommendation"):
                    desc_parts.append(f"\nRecommendation:\n{d['recommendation']}")
                if d.get("benefits"):
                    desc_parts.append(f"\nWhy It Matters:\n{d['benefits']}")
                if d.get("effort"):
                    desc_parts.append(f"\nEffort: {d['effort']}")
                if d.get("reference_url"):
                    desc_parts.append(f"\nReference: {d['reference_url']}")
                task_desc = "\n".join(desc_parts)
                task_prio = _JIRA_PRIO_MAP.get(d["priority_label"][:2], "Medium")

                task_key = _create_issue(
                    client, child_type,
                    f"[SAT] {d['check_id']} - {d['title']}",
                    task_desc, task_prio, story_key,
                )
                if task_key:
                    print(f"      {task_key}: {d['check_id']}")

    print(f"\n  Jira export complete: {created} created, {errors} errors")
    return {"created": created, "errors": errors, "issues": issues}


# ---------------------------------------------------------------------------
# Azure DevOps export via httpx
# ---------------------------------------------------------------------------

_ADO_SEVERITY_MAP = {
    "P1": "1 - Critical",
    "P2": "2 - High",
    "P3": "3 - Medium",
    "P4": "4 - Low",
}


def export_ado(
    result: SATScanResult | None,
    *,
    ado_org: str,
    ado_project: str,
    ado_token: str,
    parent_type: str = "Epic",
    child_type: str = "User Story",
    findings: list[SATFinding] | None = None,
    dry_run: bool = False,
    area_path: str = "",
    iteration_path: str = "",
) -> dict:
    """Create Azure DevOps work items from scan findings using httpx.

    Creates one parent work item (Epic) per category, then one child work item
    (User Story) per finding linked to the parent.

    Returns dict with 'created', 'errors', and 'work_items' list.
    """
    import httpx
    import base64

    src = findings if findings is not None else (result.findings if result else [])
    prio_items = _build_prioritised_recommendations(src)
    if not prio_items:
        return {"created": 0, "errors": 0, "work_items": []}

    # Group by category
    groups: dict[str, list[dict]] = {}
    for item in prio_items:
        groups.setdefault(item["category"], []).append(item)

    org = ado_org.rstrip("/")
    # ADO uses Basic auth with empty username and PAT as password
    b64_token = base64.b64encode(f":{ado_token}".encode()).decode()
    headers = {
        "Content-Type": "application/json-patch+json",
        "Authorization": f"Basic {b64_token}",
    }
    created = 0
    errors = 0
    work_items: list[dict] = []

    def _create_work_item(client: httpx.Client, wi_type: str, title: str,
                          description: str, severity: str,
                          parent_id: int | None = None) -> int | None:
        nonlocal created, errors
        ops: list[dict] = [
            {"op": "add", "path": "/fields/System.Title", "value": title[:255]},
            {"op": "add", "path": "/fields/System.Description", "value": description},
            {"op": "add", "path": "/fields/Microsoft.VSTS.Common.Severity", "value": severity},
        ]
        if area_path:
            ops.append({"op": "add", "path": "/fields/System.AreaPath", "value": area_path})
        if iteration_path:
            ops.append({"op": "add", "path": "/fields/System.IterationPath", "value": iteration_path})
        if parent_id is not None:
            ops.append({
                "op": "add",
                "path": "/relations/-",
                "value": {
                    "rel": "System.LinkTypes.Hierarchy-Reverse",
                    "url": f"{org}/{ado_project}/_apis/wit/workItems/{parent_id}",
                },
            })

        if dry_run:
            created += 1
            fake_id = created
            print(f"  [dry-run] {wi_type}: {title[:80]}  (parent={parent_id or 'none'})")
            work_items.append({"id": fake_id, "type": wi_type, "title": title})
            return fake_id

        url_type = wi_type.replace(" ", "%20")
        resp = client.post(
            f"{org}/{ado_project}/_apis/wit/workitems/${url_type}?api-version=7.1",
            json=ops,
            headers=headers,
        )
        if resp.status_code in (200, 201):
            wi_id = resp.json().get("id", 0)
            created += 1
            work_items.append({"id": wi_id, "type": wi_type, "title": title})
            return wi_id
        else:
            errors += 1
            detail = resp.text[:300]
            print(f"  ERROR creating {wi_type} '{title[:60]}': {resp.status_code} {detail}")
            return None

    with httpx.Client(timeout=30.0) as client:
        for cat, items in groups.items():
            # Build Epic description
            finding_list = "\n".join(
                f"  {i}. [{d['check_id']}] {d['title']}" for i, d in enumerate(items, 1)
            )
            epic_desc = (
                f"<p>Security findings for category: <b>{_esc(cat)}</b></p>"
                f"<p>Findings ({len(items)}):</p><pre>{_esc(finding_list)}</pre>"
            )
            prio_label = items[0]["priority_label"][:2]
            epic_severity = _ADO_SEVERITY_MAP.get(prio_label, "3 - Medium")

            print(f"  Creating {parent_type}: {cat} ({len(items)} findings)")
            epic_id = _create_work_item(
                client, parent_type,
                f"[SAT] {cat} Security Findings",
                epic_desc, epic_severity,
            )
            if epic_id is not None:
                print(f"    -> #{epic_id}")
            else:
                print(f"    -> FAILED, skipping children")
                continue

            # Create child work items
            for d in items:
                desc_parts = [
                    f"<b>Check:</b> {_esc(d['check_id'])}",
                    f"<b>Severity:</b> {_esc(d['severity'].upper())}",
                    f"<b>Status:</b> {_esc(d['status'])}",
                ]
                if d.get("current_state"):
                    desc_parts.append(f"<br><b>Current State:</b><br>{_esc(d['current_state'])}")
                if d.get("recommendation"):
                    desc_parts.append(f"<br><b>Recommendation:</b><br>{_esc(d['recommendation'])}")
                if d.get("benefits"):
                    desc_parts.append(f"<br><b>Why It Matters:</b><br>{_esc(d['benefits'])}")
                if d.get("effort"):
                    desc_parts.append(f"<br><b>Effort:</b> {_esc(d['effort'])}")
                if d.get("reference_url"):
                    desc_parts.append(f'<br><b>Reference:</b> <a href="{_esc(d["reference_url"])}">{_esc(d["reference_url"])}</a>')
                task_desc = "<br>".join(desc_parts)
                task_severity = _ADO_SEVERITY_MAP.get(d["priority_label"][:2], "3 - Medium")

                wi_id = _create_work_item(
                    client, child_type,
                    f"[SAT] {d['check_id']} - {d['title']}",
                    task_desc, task_severity, epic_id,
                )
                if wi_id is not None:
                    print(f"      #{wi_id}: {d['check_id']}")

    print(f"\n  Azure DevOps export complete: {created} created, {errors} errors")
    return {"created": created, "errors": errors, "work_items": work_items}


def export_webhook(result: SATScanResult, webhook_url: str, quiet: bool = False) -> bool:
    """POST a scan summary to a webhook URL (Slack, Teams, Discord, or generic HTTP)."""
    import urllib.request
    import urllib.error

    fails = [f for f in result.findings if f.status in ("FAIL", "WARN")]
    high = sum(1 for f in fails if f.severity == "high")
    medium = sum(1 for f in fails if f.severity == "medium")
    low = sum(1 for f in fails if f.severity == "low")
    score = getattr(result, "overall_score", "N/A")

    summary = (
        f"SAT Scanner — {result.workspace_name or result.workspace_url}\n"
        f"Score: {score}  |  {len(fails)} issue(s): {high} high, {medium} medium, {low} low\n"
        f"Total checks: {len(result.findings)}"
    )

    # Auto-detect Slack vs Teams vs generic format
    if "hooks.slack.com" in webhook_url or "slack" in webhook_url.lower():
        payload = {"text": summary}
    elif "webhook.office.com" in webhook_url or "microsoft" in webhook_url.lower():
        payload = {"text": summary}
    else:
        payload = {
            "text": summary,
            "workspace": result.workspace_name or result.workspace_url,
            "score": score,
            "findings": {"total": len(result.findings), "high": high, "medium": medium, "low": low},
        }

    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(webhook_url, data=data, headers={"Content-Type": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            if not quiet:
                print(f"  Webhook notification sent ({resp.status})")
            return True
    except urllib.error.URLError as e:
        print(f"  Webhook notification failed: {e}")
        return False


# ─────────────────────────────────────────────────────────────────────────────
# SARIF 2.1.0 export (for CI/CD integration — GitHub Code Scanning, Azure DevOps, etc.)
# ─────────────────────────────────────────────────────────────────────────────

def _sarif_severity_level(severity: str) -> str:
    """Map SAT severity to SARIF defaultConfiguration level."""
    s = (severity or "").lower()
    if s in ("critical", "high"):
        return "error"
    if s == "medium":
        return "warning"
    return "note"


def _sarif_status_level(status: str) -> str:
    """Map SAT check status to SARIF result level."""
    s = (status or "").upper()
    if s == "FAIL":
        return "error"
    if s == "WARN":
        return "warning"
    if s == "PASS":
        return "none"
    return "note"  # NOT_APPLICABLE or unknown


def export_sarif(result: SATScanResult, output_path: str, quiet: bool = False) -> str:
    """Export scan results in SARIF 2.1.0 format for CI/CD integration."""

    # Build rules list (one per unique check_id)
    seen_rules: dict[str, int] = {}
    rules: list[dict] = []
    for f in result.findings:
        if f.check_id not in seen_rules:
            seen_rules[f.check_id] = len(rules)
            rules.append({
                "id": f.check_id,
                "name": f.check_id,
                "shortDescription": {"text": f.title},
                "fullDescription": {"text": f.description or f.title},
                "helpUri": f.reference_url or "",
                "defaultConfiguration": {
                    "level": _sarif_severity_level(f.severity),
                },
                "properties": {
                    "severity": f.severity,
                    "category": f.category,
                },
            })

    # Build results list
    sarif_results: list[dict] = []
    for f in result.findings:
        sarif_result: dict = {
            "ruleId": f.check_id,
            "ruleIndex": seen_rules[f.check_id],
            "level": _sarif_status_level(f.status),
            "message": {"text": f.current_state or f.title},
            "locations": [
                {
                    "physicalLocation": {
                        "artifactLocation": {
                            "uri": result.workspace_url,
                            "uriBaseId": "DATABRICKS",
                        }
                    }
                }
            ],
            "properties": {
                "status": f.status,
                "recommendation": f.recommendation,
            },
        }
        if f.effort:
            sarif_result["properties"]["effort"] = f.effort
        if f.benefits:
            sarif_result["properties"]["benefits"] = f.benefits
        sarif_results.append(sarif_result)

    # Assemble the SARIF envelope
    sarif: dict = {
        "$schema": "https://raw.githubusercontent.com/oasis-tcs/sarif-spec/main/sarif-2.1/schema/sarif-schema-2.1.0.json",
        "version": "2.1.0",
        "runs": [
            {
                "tool": {
                    "driver": {
                        "name": "SAT Scanner",
                        "version": "0.1.0",
                        "informationUri": "https://github.com/databricks-industry-solutions/security-analysis-tool",
                        "rules": rules,
                    }
                },
                "results": sarif_results,
                "invocations": [
                    {
                        "executionSuccessful": True,
                        "startTimeUtc": result.scanned_at,
                        "endTimeUtc": result.scanned_at,
                    }
                ],
            }
        ],
    }

    path = Path(output_path)
    path.write_text(json.dumps(sarif, indent=2, default=str), encoding="utf-8")
    if not quiet:
        print(f"  SARIF report: {path}")


# ─────────────────────────────────────────────────────────────────────────────
# Unity Catalog Inventory exports (JSON + Excel)
# ─────────────────────────────────────────────────────────────────────────────

def _inventory_prefix(inv) -> str:
    prefix = "sat-uc-inventory"
    if inv.workspace_name:
        prefix += f"-{_sanitize_name(inv.workspace_name)}"
    prefix += f"-{datetime.now().strftime('%Y-%m-%d')}"
    return prefix


def export_inventory_json(inv, output_dir: Path) -> str:
    """Write the full UC + Azure inventory tree as JSON."""
    path = output_dir / f"{_inventory_prefix(inv)}.json"
    path.write_text(json.dumps(inv.to_dict(), indent=2, default=str))
    return str(path)


def export_inventory_hierarchy_html(inv, output_dir: Path) -> list[str]:
    """Write the offline, interactive views of the UC inventory: an overview
    dashboard, a collapsible tree, a zoomable sunburst, a hub-and-spoke drill-down
    (all over Metastore → Catalog → Schema → Table/View/Volume/Function/Model with
    data-volume weighting), and an infrastructure topology (storage accounts ←
    external locations, credentials, connections). All cross-link via a nav bar."""
    from .uc_hierarchy import (
        build_uc_tree, build_uc_topology, build_overview, build_nav,
        render_tree_html, render_star_html, render_hub_html,
        render_topology_html, render_overview_html,
    )
    label = inv.workspace_name or "Unity Catalog"
    prefix = _inventory_prefix(inv)
    tree = build_uc_tree(inv)
    written: list[str] = []

    def _write(suffix: str, content: str) -> None:
        path = output_dir / f"{prefix}-{suffix}.html"
        path.write_text(content, encoding="utf-8")
        written.append(str(path))

    def nav(active: str) -> str:
        return build_nav(prefix, active)

    _write("overview", render_overview_html(
        build_overview(inv), f"Unity Catalog Inventory — {label} — Overview", nav("Overview")))
    _write("tree", render_tree_html(
        tree, f"Unity Catalog Inventory — {label} — Tree", nav("Tree")))
    _write("star", render_star_html(
        tree, f"Unity Catalog Inventory — {label} — Sunburst", nav("Sunburst")))
    _write("hubspoke", render_hub_html(
        tree, f"Unity Catalog Inventory — {label} — Hub & Spoke", nav("Hub & Spoke")))
    _write("topology", render_topology_html(
        build_uc_topology(inv), f"Unity Catalog Inventory — {label} — Infrastructure", nav("Infrastructure")))
    return written


def export_inventory_fleet_html(results, output_dir: Path) -> list[str]:
    """Write the five **fleet-combined** interactive diagrams at the run-directory
    root: a single set spanning every workspace, grouped by metastore
    (Fleet → Metastore → Catalog → Schema → object), plus a unioned infrastructure
    topology and a fleet-wide overview. Filenames share the combined-summary prefix."""
    from .uc_hierarchy import (
        build_fleet_tree, build_fleet_topology, build_fleet_overview, build_nav,
        render_tree_html, render_star_html, render_hub_html,
        render_topology_html, render_overview_html,
    )
    inventories = [inv for (_n, _h, _t, inv) in results if inv is not None]
    if not inventories:
        return []
    prefix = _combined_prefix()
    tree = build_fleet_tree(inventories)
    written: list[str] = []

    def _write(suffix: str, content: str) -> None:
        path = output_dir / f"{prefix}-{suffix}.html"
        path.write_text(content, encoding="utf-8")
        written.append(str(path))

    def nav(active: str) -> str:
        return build_nav(prefix, active)

    _write("overview", render_overview_html(
        build_fleet_overview(inventories), "Unity Catalog Inventory — Fleet — Overview", nav("Overview")))
    _write("tree", render_tree_html(
        tree, "Unity Catalog Inventory — Fleet — Tree", nav("Tree")))
    _write("star", render_star_html(
        tree, "Unity Catalog Inventory — Fleet — Sunburst", nav("Sunburst")))
    _write("hubspoke", render_hub_html(
        tree, "Unity Catalog Inventory — Fleet — Hub & Spoke", nav("Hub & Spoke")))
    _write("topology", render_topology_html(
        build_fleet_topology(inventories), "Unity Catalog Inventory — Fleet — Infrastructure", nav("Infrastructure")))
    return written


# Control characters openpyxl refuses to write to a worksheet (XML 1.0 forbids
# them). Tab (\x09), newline (\x0a) and carriage return (\x0d) are allowed and
# deliberately excluded. AI-generated text often smuggles these in — e.g. LaTeX
# like "\bigl" / "\frac" whose backslash escapes decoded to \x08 / \x0c.
_ILLEGAL_XL_CHARS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _xl_cell(val: Any) -> Any:
    """Coerce a value to an Excel-safe cell (serialise dict/list, strip illegal
    control chars, truncate long text)."""
    if isinstance(val, (dict, list)):
        val = json.dumps(val, default=str, ensure_ascii=False)
    if isinstance(val, str):
        val = _ILLEGAL_XL_CHARS_RE.sub("", val)
        if len(val) > EXCEL_CELL_LIMIT:
            return val[: EXCEL_CELL_LIMIT - 30] + "... [truncated]"
    return val


# ─────────────────────────────────────────────────────────────────────────────
# Azure resource hierarchy exporters
#   (Management group → Subscription → Resource group → Resource)
# ─────────────────────────────────────────────────────────────────────────────

def _azure_hierarchy_prefix() -> str:
    return f"azure-hierarchy-{datetime.now().strftime('%Y-%m-%d_%H%M%S')}"


def export_azure_hierarchy_json(tree: dict, output_dir: Path) -> str:
    """Write the assembled Azure resource hierarchy tree as JSON."""
    path = output_dir / f"{_azure_hierarchy_prefix()}.json"
    path.write_text(json.dumps(tree, indent=2, default=str), encoding="utf-8")
    return str(path)


def export_azure_hierarchy_html(tree: dict, output_dir: Path) -> list[str]:
    """Write the three self-contained, offline HTML views (tree, sunburst, hub-and-spoke)."""
    from .azure_hierarchy import render_tree_html, render_star_html, render_hub_html
    prefix = _azure_hierarchy_prefix()
    written: list[str] = []
    for suffix, render in (("tree", render_tree_html),
                           ("star", render_star_html),
                           ("hubspoke", render_hub_html)):
        path = output_dir / f"{prefix}-{suffix}.html"
        path.write_text(render(tree), encoding="utf-8")
        written.append(str(path))
    return written


def export_azure_hierarchy_excel(tree: dict, output_dir: Path) -> str:
    """Write the Azure resource hierarchy as a multi-sheet Excel workbook.

    Sheets: Summary, Hierarchy (indented full tree), Management Groups,
    Subscriptions, Resource Groups, Resources (with SKU + Tags).
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: 'openpyxl' is required for Excel export.  Install with:  pip install sat-scanner[excel]")
        sys.exit(1)
    from .azure_hierarchy import (
        count_by_type, walk_tree, MGMT_GROUP, SUBSCRIPTION, RESOURCE_GROUP, RESOURCE,
    )

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))

    def _sheet(title: str, cols: list[str], rows: list[list]):
        ws = wb.create_sheet(title[:31])
        for c, val in enumerate(cols, 1):
            cell = ws.cell(row=1, column=c, value=val)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin
        for r, row in enumerate(rows, 2):
            for c, val in enumerate(row, 1):
                ws.cell(row=r, column=c, value=_xl_cell(val)).border = thin
        for c in range(1, len(cols) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 26
        return ws

    counts = count_by_type(tree)

    # ── Summary ──
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.cell(row=1, column=1, value="Azure Resource Hierarchy").font = title_font
    summary_rows = [
        ("Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Root", tree.get("name", "")),
        ("Management Groups", counts.get(MGMT_GROUP, 0)),
        ("Subscriptions", counts.get(SUBSCRIPTION, 0)),
        ("Resource Groups", counts.get(RESOURCE_GROUP, 0)),
        ("Resources", counts.get(RESOURCE, 0)),
    ]
    for r, (k, v) in enumerate(summary_rows, 3):
        ws_sum.cell(row=r, column=1, value=k).border = thin
        ws_sum.cell(row=r, column=2, value=_xl_cell(v)).border = thin
    ws_sum.column_dimensions["A"].width = 24
    ws_sum.column_dimensions["B"].width = 64

    # ── Walk once; build the indented Hierarchy sheet + the flat per-level sheets ──
    hier_rows, mg_rows, sub_rows, rg_rows, res_rows = [], [], [], [], []
    for depth, node in walk_tree(tree):
        t = node.get("type", "")
        hier_rows.append([
            depth, ("    " * depth) + node.get("name", ""), t,
            node.get("rtype_short", "") or node.get("resource_type", ""),
            node.get("location", ""),
            node.get("subscription_id") or node.get("subscription", "") or node.get("id", ""),
        ])
        if t == MGMT_GROUP:
            mg_rows.append([node.get("name", ""), node.get("subs", ""),
                            node.get("access", ""), node.get("id", "")])
        elif t == SUBSCRIPTION:
            sub_rows.append([node.get("name", ""),
                             node.get("subscription_id") or node.get("id", ""),
                             node.get("access", "")])
        elif t == RESOURCE_GROUP:
            rg_rows.append([node.get("name", ""), node.get("location", ""),
                            node.get("subscription", ""), node.get("id", "")])
        elif t == RESOURCE:
            res_rows.append([
                node.get("name", ""), node.get("resource_type", ""), node.get("kind", ""),
                node.get("location", ""), node.get("rg", ""), node.get("subscription", ""),
                node.get("sku", ""), node.get("tags", {}), node.get("id", ""),
            ])

    _sheet("Hierarchy", ["Level", "Name", "Type", "Resource Type", "Location", "Subscription / Id"], hier_rows)
    _sheet("Management Groups", ["Name", "Subscriptions", "Access", "Id"], mg_rows)
    _sheet("Subscriptions", ["Name", "Subscription Id", "Access"], sub_rows)
    _sheet("Resource Groups", ["Name", "Location", "Subscription Id", "Id"], rg_rows)
    _sheet("Resources", ["Name", "Type", "Kind", "Location", "Resource Group",
                         "Subscription Id", "SKU", "Tags", "Id"], res_rows)

    path = output_dir / f"{_azure_hierarchy_prefix()}.xlsx"
    wb.save(str(path))
    return str(path)


def export_inventory_excel(inv, output_dir: Path) -> str:
    """Write the UC + Azure inventory as a standalone multi-sheet Excel workbook."""
    try:
        from openpyxl import Workbook
    except ImportError:
        print("ERROR: 'openpyxl' is required for Excel export.  Install with:  pip install sat-scanner[excel]")
        sys.exit(1)
    wb = Workbook()
    _build_inventory_sheets(wb, inv)
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]
    path = output_dir / f"{_inventory_prefix(inv)}.xlsx"
    wb.save(str(path))
    return str(path)


def _build_inventory_sheets(wb, inv, prefix: str = "") -> None:
    """Add UC + Azure inventory sheets to an existing workbook.

    Used both for the standalone inventory workbook and to enrich the SAT scan
    workbook (with prefix='UC ' so sheet names don't collide).
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))

    def _sheet(title: str, cols: list[str], rows: list[list]):
        ws = wb.create_sheet((prefix + title)[:31])
        for c, val in enumerate(cols, 1):
            cell = ws.cell(row=1, column=c, value=val)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin
        for r, row in enumerate(rows, 2):
            for c, val in enumerate(row, 1):
                ws.cell(row=r, column=c, value=_xl_cell(val)).border = thin
        # rough column widths
        for c in range(1, len(cols) + 1):
            ws.column_dimensions[chr(64 + c) if c <= 26 else "AA"].width = 24
        return ws

    # ── Summary ──
    ws_sum = wb.create_sheet((prefix + "Summary")[:31])
    ws_sum.cell(row=1, column=1, value="Databricks Unity Catalog Inventory").font = title_font
    az = inv.azure or {}
    az_ws = az.get("workspace") or {}
    rows = [
        ("Databricks Workspace Name", inv.workspace_name or "N/A"),
        ("Workspace URL", inv.workspace_url),
        ("Scanned At", inv.scanned_at),
        ("Metastore", (inv.metastore.get("current_assignment", {}) or {}).get("metastore_id", "")),
        ("Azure Subscription", az_ws.get("subscription_id", "")),
        ("Azure Resource Group", az_ws.get("resource_group", "")),
        ("Managed Resource Group", az_ws.get("managed_resource_group_id", "")),
        ("Azure Region / Geo", f"{az_ws.get('location','')} / {az_ws.get('geo','')}" if az_ws else ""),
        ("Workspace SKU", az_ws.get("sku", "")),
        ("Azure Discovery", "available" if az.get("available") else f"skipped — {az.get('reason','')}"),
    ]
    for k, v in (inv.stats or {}).items():
        rows.append((k.replace("_", " ").title(), v))
    for r, (k, v) in enumerate(rows, 3):
        ws_sum.cell(row=r, column=1, value=k).border = thin
        ws_sum.cell(row=r, column=2, value=_xl_cell(v)).border = thin
    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 90

    # ── Catalogs / Schemas / Tables / Views / Columns / Volumes / Functions / Models ──
    cat_rows, sch_rows, tab_rows, view_rows, col_rows = [], [], [], [], []
    vol_rows, fn_rows, mdl_rows, grant_rows, tag_rows = [], [], [], [], []
    cons_rows, rfilter_rows, mask_rows, mver_rows, bind_rows, mon_rows = [], [], [], [], [], []

    def _add_grants(grants):
        for g in grants:
            grant_rows.append([g.securable_type, g.full_name, g.principal,
                               ", ".join(g.privileges), g.inherited_from])

    for g in inv.metastore_grants:
        grant_rows.append([g.securable_type, g.full_name, g.principal,
                           ", ".join(g.privileges), g.inherited_from])

    for c in inv.catalogs:
        cat_rows.append([c.name, c.catalog_type, c.owner, c.comment, c.storage_root,
                         c.isolation_mode, len(c.schemas), c.properties, c.tags])
        _add_grants(c.grants)
        for b in c.bindings:
            bind_rows.append([c.name, b.get("workspace_id", ""), b.get("binding_type", "")])
        for tk, tv in (c.tags or {}).items():
            tag_rows.append(["catalog", c.name, tk, tv])
        for s in c.schemas:
            sch_rows.append([s.full_name, s.catalog, s.name, s.owner, s.comment,
                             len(s.tables), len(s.volumes), len(s.functions), len(s.models),
                             s.properties])
            _add_grants(s.grants)
            for t in s.tables:
                row = [t.full_name, t.catalog, t.schema, t.name, t.table_type,
                       t.data_source_format, t.storage_location, t.owner, t.comment,
                       len(t.columns), t.created_at, t.updated_at]
                if t.table_type == "VIEW":
                    view_rows.append([t.full_name, t.owner, t.comment, t.view_definition])
                else:
                    tab_rows.append(row)
                _add_grants(t.grants)
                for con in t.constraints:
                    cons_rows.append([t.full_name, con.get("name", ""), con.get("type", ""),
                                      ", ".join(con.get("columns", []))])
                if t.row_filter:
                    rfilter_rows.append([t.full_name, (t.row_filter or {}).get("function_name", ""),
                                         ", ".join((t.row_filter or {}).get("input_column_names", []) or [])])
                if t.monitor:
                    mon_rows.append([t.full_name, t.monitor.get("status", ""),
                                     t.monitor.get("output_schema_name", ""),
                                     t.monitor.get("assets_dir", "")])
                for col in t.columns:
                    mask_fn = (col.mask or {}).get("function_name", "") if col.mask else ""
                    col_rows.append([t.full_name, col.position, col.name, col.type_text,
                                     col.nullable, col.comment, col.tags, mask_fn])
                    if col.mask:
                        mask_rows.append([t.full_name, col.name, mask_fn])
                for tk, tv in (t.tags or {}).items():
                    tag_rows.append(["table", t.full_name, tk, tv])
            for v in s.volumes:
                vol_rows.append([v.full_name, v.volume_type, v.storage_location, v.owner, v.comment])
                _add_grants(v.grants)
            for fn in s.functions:
                fn_rows.append([fn.full_name, fn.data_type, fn.routine_body, fn.owner, fn.comment])
                _add_grants(fn.grants)
            for md in s.models:
                mdl_rows.append([md.full_name, md.owner, md.comment, len(md.versions)])
                _add_grants(md.grants)
                for mv in md.versions:
                    mver_rows.append([md.full_name, mv.get("version", ""), mv.get("status", ""),
                                      mv.get("created_at", ""), mv.get("run_id", "")])

    _sheet("Catalogs", ["Name", "Type", "Owner", "Comment", "Storage Root", "Isolation",
                        "# Schemas", "Properties", "Tags"], cat_rows)
    _sheet("Schemas", ["Full Name", "Catalog", "Name", "Owner", "Comment", "# Tables",
                       "# Volumes", "# Functions", "# Models", "Properties"], sch_rows)
    _sheet("Tables", ["Full Name", "Catalog", "Schema", "Name", "Type", "Format",
                      "Location", "Owner", "Comment", "# Columns", "Created", "Updated"], tab_rows)
    _sheet("Views", ["Full Name", "Owner", "Comment", "Definition"], view_rows)
    _sheet("Columns", ["Table", "Position", "Name", "Type", "Nullable", "Comment", "Tags", "Mask"], col_rows)
    _sheet("Constraints", ["Table", "Constraint", "Type", "Columns"], cons_rows)
    _sheet("Row Filters", ["Table", "Filter Function", "Input Columns"], rfilter_rows)
    _sheet("Column Masks", ["Table", "Column", "Mask Function"], mask_rows)
    _sheet("Volumes", ["Full Name", "Type", "Location", "Owner", "Comment"], vol_rows)
    _sheet("Functions", ["Full Name", "Data Type", "Routine Body", "Owner", "Comment"], fn_rows)
    _sheet("Registered Models", ["Full Name", "Owner", "Comment", "# Versions"], mdl_rows)
    _sheet("Model Versions", ["Model", "Version", "Status", "Created", "Run Id"], mver_rows)
    _sheet("Catalog Bindings", ["Catalog", "Workspace Id", "Binding Type"], bind_rows)
    _sheet("Monitors", ["Table", "Status", "Output Schema", "Assets Dir"], mon_rows)
    _sheet("Grants", ["Securable", "Full Name", "Principal", "Privileges", "Inherited From"], grant_rows)
    _sheet("Tags", ["Object Type", "Full Name", "Tag Key", "Tag Value"], tag_rows)

    # Storage account → {resource_group, subscription_id} lookup for joining onto UC rows
    acct_meta = {a.get("name", ""): a for a in (az.get("storage_accounts") or [])}
    ws_name = inv.workspace_name or ""

    def _rg_for(account: str) -> str:
        return (acct_meta.get(account, {}) or {}).get("resource_group", "")

    # ── Metastore-level securables ──
    _sheet("External Locations",
           ["Workspace", "Name", "URL", "Credential", "Read Only",
            "Azure Storage Account", "Container", "Resource Group"],
           [[ws_name, l.get("name", ""), l.get("url", ""), l.get("credential_name", ""),
             l.get("read_only", ""),
             (l.get("azure", {}) or {}).get("storage_account", ""),
             (l.get("azure", {}) or {}).get("container", ""),
             _rg_for((l.get("azure", {}) or {}).get("storage_account", ""))]
            for l in inv.external_locations])
    _sheet("Storage Credentials",
           ["Name", "Owner", "Comment", "Azure Managed Identity"],
           [[sc.get("name", ""), sc.get("owner", ""), sc.get("comment", ""),
             sc.get("azure_managed_identity", "")] for sc in inv.storage_credentials])
    _sheet("Service Credentials",
           ["Name", "Purpose", "Owner", "Comment"],
           [[sc.get("name", ""), sc.get("purpose", ""), sc.get("owner", ""), sc.get("comment", "")]
            for sc in inv.service_credentials])
    _sheet("Connections",
           ["Name", "Type", "Owner", "Comment"],
           [[cn.get("name", ""), cn.get("connection_type", ""), cn.get("owner", ""), cn.get("comment", "")]
            for cn in inv.connections])
    _sheet("Sharing Providers",
           ["Name", "Auth Type", "Owner", "Comment"],
           [[p.get("name", ""), p.get("authentication_type", ""), p.get("owner", ""), p.get("comment", "")]
            for p in inv.providers])

    # ── Azure sheets ──
    ws = az.get("workspace") or {}
    if ws:
        _sheet("Azure Workspace", ["Field", "Value"], [[k, v] for k, v in ws.items()])
    _sheet("Azure Storage Accounts",
           ["Name", "Resource Group", "Subscription", "Location", "HNS Enabled",
            "Public Network", "Default Action", "Min TLS", "Resolved"],
           [[a.get("name", ""), a.get("resource_group", ""), a.get("subscription_id", ""),
             a.get("location", ""), a.get("hns_enabled", ""), a.get("public_network_access", ""),
             a.get("network_default_action", ""), a.get("min_tls_version", ""), a.get("resolved", "")]
            for a in az.get("storage_accounts", [])])
    role_rows = []
    for a in az.get("storage_accounts", []):
        for ra in a.get("role_assignments", []):
            role_rows.append([a.get("name", ""), a.get("resource_group", ""), ra.get("role_name", ""),
                              ra.get("principal_id", ""), ra.get("principal_type", "")])
    _sheet("Azure Role Assignments",
           ["Storage Account", "Resource Group", "Role", "Principal Id", "Principal Type"], role_rows)
    _sheet("Azure UC Mapping",
           ["Workspace", "UC Type", "UC Name", "URL", "Storage Account", "Container",
            "Resource Group", "Resolved", "Credential", "Identity", "Granting Roles", "Notes"],
           [[ws_name, m.get("uc_object_type", ""), m.get("uc_name", ""), m.get("url", ""),
             m.get("storage_account", ""), m.get("container", ""), _rg_for(m.get("storage_account", "")),
             m.get("storage_account_resolved", ""), m.get("credential_name", ""),
             (m.get("identity") or {}).get("name", "") if m.get("identity") else "",
             ", ".join(m.get("granting_roles", [])), "; ".join(m.get("notes", []))]
            for m in az.get("mappings", [])])

    # ── Errors ──
    _sheet("Errors", ["Level", "Full Name", "HTTP Status", "Error"],
           [[e.get("level", ""), e.get("full_name", ""), e.get("http_status", ""), e.get("error", "")]
            for e in inv.errors])


def export_inventory_html(inv, output_dir: Path) -> str:
    """Render the UC + Azure inventory as a self-contained styled HTML report."""
    import html as _html
    esc = _html.escape
    az = inv.azure or {}
    az_ws = az.get("workspace") or {}
    acct_meta = {a.get("name", ""): a for a in (az.get("storage_accounts") or [])}
    s = inv.stats or {}

    def _table(headers: list[str], rows: list[list], empty: str = "None found") -> str:
        if not rows:
            return f'<p class="empty">{esc(empty)}</p>'
        head = "".join(f"<th>{esc(str(h))}</th>" for h in headers)
        body = []
        for r in rows:
            cells = "".join(f"<td>{esc('' if c is None else str(c))}</td>" for c in r)
            body.append(f"<tr>{cells}</tr>")
        return (f'<div class="tablewrap"><table><thead><tr>{head}</tr></thead>'
                f'<tbody>{"".join(body)}</tbody></table></div>')

    def _card(label: str, value) -> str:
        return f'<div class="stat"><div class="stat-v">{esc(str(value))}</div><div class="stat-l">{esc(label)}</div></div>'

    # Stat cards
    card_keys = [("catalogs", "Catalogs"), ("schemas", "Schemas"), ("tables", "Tables"),
                 ("views", "Views"), ("columns", "Columns"), ("volumes", "Volumes"),
                 ("functions", "Functions"), ("registered_models", "Models"),
                 ("grants", "Grants"), ("constraints", "Constraints"),
                 ("masked_columns", "Masked Cols"), ("row_filters", "Row Filters"),
                 ("external_locations", "External Locations")]
    cards = "".join(_card(lbl, s.get(k, 0)) for k, lbl in card_keys)

    # Azure overview
    if az.get("available"):
        az_overview = _table(["Field", "Value"], [
            ["Subscription", az_ws.get("subscription_id", "")],
            ["Resource Group", az_ws.get("resource_group", "")],
            ["Managed Resource Group", az_ws.get("managed_resource_group_id", "")],
            ["Region / Geo", f"{az_ws.get('location','')} / {az_ws.get('geo','')}"],
            ["SKU", az_ws.get("sku", "")],
            ["VNet Injected", az_ws.get("vnet_injected", "")],
            ["No Public IP (SCC)", az_ws.get("no_public_ip", "")],
            ["Storage Accounts", len(az.get("storage_accounts", []))],
            ["UC→Azure Mappings", len(az.get("mappings", []))],
        ])
        storage_tbl = _table(
            ["Name", "Resource Group", "Subscription", "Location", "HNS", "Public Network", "Default Action"],
            [[a.get("name", ""), a.get("resource_group", ""), a.get("subscription_id", ""),
              a.get("location", ""), a.get("hns_enabled", ""), a.get("public_network_access", ""),
              a.get("network_default_action", "")] for a in az.get("storage_accounts", [])])
        map_tbl = _table(
            ["UC Name", "Type", "Storage Account", "Container", "Resource Group", "Resolved", "Granting Roles", "Notes"],
            [[m.get("uc_name", ""), m.get("uc_object_type", ""), m.get("storage_account", ""),
              m.get("container", ""), (acct_meta.get(m.get("storage_account", ""), {}) or {}).get("resource_group", ""),
              "✓" if m.get("storage_account_resolved") else "✗",
              ", ".join(m.get("granting_roles", [])), "; ".join(m.get("notes", []))]
             for m in az.get("mappings", [])])
        azure_section = (f'<h2>Azure Infrastructure</h2>{az_overview}'
                         f'<h3>Storage Accounts</h3>{storage_tbl}'
                         f'<h3>UC → Azure Mapping</h3>{map_tbl}')
    else:
        azure_section = (f'<h2>Azure Infrastructure</h2>'
                         f'<p class="empty">Skipped — {esc(az.get("reason", ""))}</p>')

    # Catalogs + tables
    cat_rows, tab_rows = [], []
    for c in inv.catalogs:
        cat_rows.append([c.name, c.catalog_type, c.owner, len(c.schemas),
                         sum(len(sc.tables) for sc in c.schemas)])
        for sc in c.schemas:
            for t in sc.tables:
                tab_rows.append([t.full_name, t.table_type, t.data_source_format,
                                 len(t.columns), t.owner])
    catalogs_tbl = _table(["Catalog", "Type", "Owner", "# Schemas", "# Tables"], cat_rows)
    tables_tbl = _table(["Full Name", "Type", "Format", "# Columns", "Owner"], tab_rows)

    grant_rows = []
    for c in inv.catalogs:
        for g in c.grants:
            grant_rows.append([g.securable_type, g.full_name, g.principal, ", ".join(g.privileges)])
        for sc in c.schemas:
            for g in sc.grants:
                grant_rows.append([g.securable_type, g.full_name, g.principal, ", ".join(g.privileges)])
    grants_tbl = _table(["Securable", "Full Name", "Principal", "Privileges"], grant_rows)

    errors_tbl = ""
    if inv.errors:
        errors_tbl = ('<h2>Warnings &amp; Errors</h2>' + _table(
            ["Level", "Object", "HTTP", "Error"],
            [[e.get("level", ""), e.get("full_name", ""), e.get("http_status", ""),
              str(e.get("error", ""))[:300]] for e in inv.errors]))

    # Links to the interactive diagrams (written alongside this report by
    # export_inventory_hierarchy_html; same prefix, suffixed).
    _diag_prefix = _inventory_prefix(inv)
    diagrams_nav = (
        '<div class="diagrams">'
        '<span class="diagrams-l">Interactive diagrams:</span>'
        f'<a href="{esc(_diag_prefix)}-overview.html">Overview</a>'
        f'<a href="{esc(_diag_prefix)}-tree.html">Tree</a>'
        f'<a href="{esc(_diag_prefix)}-star.html">Sunburst</a>'
        f'<a href="{esc(_diag_prefix)}-hubspoke.html">Hub &amp; Spoke</a>'
        f'<a href="{esc(_diag_prefix)}-topology.html">Infrastructure</a>'
        '</div>'
    )

    title = f"Unity Catalog Inventory — {inv.workspace_name or inv.workspace_url}"
    html_doc = f"""<!DOCTYPE html><html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{esc(title)}</title>
<style>
  * {{ box-sizing: border-box; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
         margin: 0; background: #f8fafc; color: #0f172a; font-size: 14px; }}
  header {{ background: linear-gradient(135deg, #0c4a6e, #2563eb); color: #fff; padding: 28px 40px; }}
  header h1 {{ margin: 0 0 6px; font-size: 22px; }}
  header .meta {{ font-size: 13px; opacity: .9; }}
  header .meta code {{ background: rgba(255,255,255,.15); padding: 2px 6px; border-radius: 4px; }}
  main {{ max-width: 1200px; margin: 0 auto; padding: 24px 40px 60px; }}
  h2 {{ color: #0c4a6e; border-bottom: 2px solid #bae6fd; padding-bottom: 6px; margin-top: 36px; }}
  h3 {{ color: #0369a1; margin-top: 22px; }}
  .stats {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(120px, 1fr)); gap: 12px; margin-top: 20px; }}
  .stat {{ background: #fff; border: 1px solid #e0f2fe; border-radius: 10px; padding: 16px; text-align: center;
          box-shadow: 0 1px 2px rgba(0,0,0,.04); }}
  .stat-v {{ font-size: 26px; font-weight: 700; color: #0c4a6e; }}
  .stat-l {{ font-size: 12px; color: #6b7280; margin-top: 4px; text-transform: uppercase; letter-spacing: .03em; }}
  .tablewrap {{ overflow-x: auto; margin-top: 12px; border: 1px solid #e2e8f0; border-radius: 10px; }}
  table {{ border-collapse: collapse; width: 100%; background: #fff; font-size: 13px; }}
  thead th {{ position: sticky; top: 0; background: #0c4a6e; color: #fff; text-align: left;
             padding: 8px 12px; font-weight: 600; }}
  td {{ padding: 7px 12px; border-top: 1px solid #eef2f7; vertical-align: top; word-break: break-word; }}
  tbody tr:nth-child(even) {{ background: #f0f9ff; }}
  .empty {{ color: #94a3b8; font-style: italic; margin-top: 10px; }}
  .diagrams {{ display: flex; align-items: center; gap: 10px; flex-wrap: wrap; margin-top: 20px;
              background: #fff; border: 1px solid #e0f2fe; border-radius: 10px; padding: 12px 16px; }}
  .diagrams-l {{ font-size: 13px; color: #6b7280; font-weight: 600; }}
  .diagrams a {{ font-size: 13px; color: #0c4a6e; text-decoration: none; border: 1px solid #bae6fd;
                background: #f0f9ff; border-radius: 7px; padding: 6px 14px; }}
  .diagrams a:hover {{ background: #e0f2fe; }}
  footer {{ text-align: center; color: #94a3b8; font-size: 12px; padding: 24px; }}
</style></head><body>
<header>
  <h1>{esc(title)}</h1>
  <div class="meta">Workspace <code>{esc(inv.workspace_url)}</code> &nbsp;·&nbsp;
    Scanned {esc(inv.scanned_at)} &nbsp;·&nbsp;
    Azure: {"available" if az.get("available") else "not discovered"}</div>
</header>
<main>
  <div class="stats">{cards}</div>
  {diagrams_nav}
  {azure_section}
  <h2>Catalogs</h2>{catalogs_tbl}
  <h2>Tables &amp; Views</h2>{tables_tbl}
  <h2>Grants (catalog &amp; schema)</h2>{grants_tbl}
  {errors_tbl}
</main>
<footer>Generated by SAT Scanner — Unity Catalog Inventory</footer>
</body></html>"""

    path = output_dir / f"{_inventory_prefix(inv)}.html"
    path.write_text(html_doc, encoding="utf-8")
    return str(path)


def export_source_diff_json(diff: dict, output_dir: Path, workspace_name: str = "") -> str:
    """Write an api-vs-sql inventory comparison report as JSON."""
    name = _sanitize_name(workspace_name) if workspace_name else ""
    fn = "sat-uc-inventory" + (f"-{name}" if name else "") \
        + f"-source-diff-{datetime.now().strftime('%Y-%m-%d')}.json"
    path = output_dir / fn
    path.write_text(json.dumps(diff, indent=2, default=str))
    return str(path)


# ─────────────────────────────────────────────────────────────────────────────
# Combined cross-workspace inventory exports
# ─────────────────────────────────────────────────────────────────────────────

def _combined_prefix() -> str:
    return f"sat-uc-inventory-combined-{datetime.now().strftime('%Y-%m-%d')}"

_WS_COLS = [
    ("workspace", "Workspace"), ("url", "URL"), ("status", "Status"),
    ("resource_group", "Resource Group"), ("region", "Region"), ("geo", "Geo"),
    ("azure_available", "Azure"), ("catalogs", "Catalogs"), ("schemas", "Schemas"),
    ("tables", "Tables"), ("views", "Views"), ("columns", "Columns"),
    ("volumes", "Volumes"), ("functions", "Functions"), ("grants", "Grants"),
    ("external_locations", "Ext Locations"), ("storage_accounts", "Storage Accts"),
    ("errors", "Errors"),
]


# Deduped per-metastore count columns (order shared by the Excel + HTML reports).
_DEDUP_DT_KEYS = ["catalogs", "schemas", "tables", "views", "columns", "grants",
                  "volumes", "functions", "registered_models", "external_locations"]


def export_combined_inventory_json(agg: dict, output_dir: Path) -> str:
    """Write the fleet-wide inventory roll-up as JSON."""
    path = output_dir / f"{_combined_prefix()}.json"
    path.write_text(json.dumps(agg, indent=2, default=str))
    return str(path)


def export_combined_inventory_excel(agg: dict, output_dir: Path) -> str:
    """Write the fleet roll-up as a multi-sheet Excel workbook."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("ERROR: 'openpyxl' is required for Excel export.  Install with:  pip install sat-scanner[excel]")
        sys.exit(1)

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))

    def _sheet(title, cols, rows, active=None):
        ws = active if active is not None else wb.create_sheet(title[:31])
        if active is not None:
            ws.title = title[:31]
        for c, val in enumerate(cols, 1):
            cell = ws.cell(row=1, column=c, value=val)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin
        for r, row in enumerate(rows, 2):
            for c, val in enumerate(row, 1):
                ws.cell(row=r, column=c, value=_xl_cell(val)).border = thin
        return ws

    totals = agg.get("totals", {})
    ws_t = _sheet("Totals", ["Metric", "Value"],
                  [["Workspaces", agg.get("workspace_count", 0)],
                   ["Reachable", agg.get("reachable", 0)]] +
                  [[k.replace("_", " ").title(), v] for k, v in totals.items()],
                  active=wb.active)
    ws_t.cell(row=1, column=4, value="Databricks UC Inventory — Fleet Summary").font = title_font

    _sheet("Workspaces", [lbl for _, lbl in _WS_COLS],
           [[w.get(k, "") for k, _ in _WS_COLS] for w in agg.get("workspaces", [])])
    _sheet("Metastores",
           ["Metastore Id", "Workspaces"] + [k.replace("_", " ").title() for k in _DEDUP_DT_KEYS],
           [[m.get("metastore_id", ""), ", ".join(m.get("workspaces", []))]
            + [m.get("deduped_totals", {}).get(k, 0) for k in _DEDUP_DT_KEYS]
            for m in agg.get("metastores", [])])
    _sheet("Azure Footprint",
           ["Workspace", "Storage Account", "Resource Group", "Subscription", "Location",
            "HNS", "Public Network"],
           [[a.get("workspace", ""), a.get("name", ""), a.get("resource_group", ""),
             a.get("subscription_id", ""), a.get("location", ""), a.get("hns_enabled", ""),
             a.get("public_network_access", "")] for a in agg.get("azure_storage_footprint", [])])

    path = output_dir / f"{_combined_prefix()}.xlsx"
    wb.save(str(path))
    return str(path)


def export_combined_inventory_html(agg: dict, output_dir: Path) -> str:
    """Write the fleet roll-up as a self-contained HTML report."""
    import html as _html
    esc = _html.escape
    totals = agg.get("totals", {})

    def _table(headers, rows):
        if not rows:
            return '<p class="empty">None</p>'
        head = "".join(f"<th>{esc(str(h))}</th>" for h in headers)
        body = "".join("<tr>" + "".join(f"<td>{esc('' if c is None else str(c))}</td>" for c in r) + "</tr>"
                       for r in rows)
        return f'<div class="tablewrap"><table><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table></div>'

    cards = "".join(
        f'<div class="stat"><div class="stat-v">{esc(str(v))}</div>'
        f'<div class="stat-l">{esc(k.replace("_"," "))}</div></div>'
        for k, v in [("workspaces", agg.get("workspace_count", 0))] + list(totals.items()))

    deduped = agg.get("deduped_totals", {})
    dcards = "".join(
        f'<div class="stat"><div class="stat-v">{esc(str(deduped.get(k, 0)))}</div>'
        f'<div class="stat-l">{esc(k.replace("_"," "))}</div></div>'
        for k in _DEDUP_DT_KEYS)
    ms_cols = ["Metastore Id", "Workspaces"] + [k.replace("_", " ").title() for k in _DEDUP_DT_KEYS]
    ms_rows = [[m.get("metastore_id", ""), ", ".join(m.get("workspaces", []))]
               + [m.get("deduped_totals", {}).get(k, 0) for k in _DEDUP_DT_KEYS]
               for m in agg.get("metastores", [])]

    ws_rows = [[w.get(k, "") for k, _ in _WS_COLS] for w in agg.get("workspaces", [])]
    fp_rows = [[a.get("workspace", ""), a.get("name", ""), a.get("resource_group", ""),
                a.get("location", ""), a.get("hns_enabled", ""), a.get("public_network_access", "")]
               for a in agg.get("azure_storage_footprint", [])]

    doc = f"""<!DOCTYPE html><html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Unity Catalog Inventory — Fleet Summary</title>
<style>
  * {{ box-sizing: border-box; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
         margin: 0; background: #f8fafc; color: #0f172a; font-size: 14px; }}
  header {{ background: linear-gradient(135deg, #0c4a6e, #2563eb); color: #fff; padding: 28px 40px; }}
  header h1 {{ margin: 0; font-size: 22px; }}
  main {{ max-width: 1300px; margin: 0 auto; padding: 24px 40px 60px; }}
  h2 {{ color: #0c4a6e; border-bottom: 2px solid #bae6fd; padding-bottom: 6px; margin-top: 32px; }}
  .stats {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(120px, 1fr)); gap: 12px; margin-top: 20px; }}
  .stat {{ background: #fff; border: 1px solid #e0f2fe; border-radius: 10px; padding: 16px; text-align: center; }}
  .stat-v {{ font-size: 24px; font-weight: 700; color: #0c4a6e; }}
  .stat-l {{ font-size: 12px; color: #6b7280; margin-top: 4px; text-transform: uppercase; }}
  .tablewrap {{ overflow-x: auto; margin-top: 12px; border: 1px solid #e2e8f0; border-radius: 10px; }}
  table {{ border-collapse: collapse; width: 100%; background: #fff; font-size: 13px; }}
  thead th {{ position: sticky; top: 0; background: #0c4a6e; color: #fff; text-align: left; padding: 8px 12px; }}
  td {{ padding: 7px 12px; border-top: 1px solid #eef2f7; word-break: break-word; }}
  tbody tr:nth-child(even) {{ background: #f0f9ff; }}
  .empty {{ color: #94a3b8; font-style: italic; }}
  footer {{ text-align: center; color: #94a3b8; font-size: 12px; padding: 24px; }}
</style></head><body>
<header><h1>Unity Catalog Inventory — Fleet Summary</h1>
  <div>{esc(str(agg.get("reachable", 0)))} of {esc(str(agg.get("workspace_count", 0)))} workspaces reachable</div>
</header>
<main>
  <h2>Fleet totals <small style="font-weight:400;color:#64748b">(summed across workspaces)</small></h2>
  <div class="stats">{cards}</div>
  <h2>Deduped totals <small style="font-weight:400;color:#64748b">(each shared UC object counted once per metastore)</small></h2>
  <div class="stats">{dcards}</div>
  <h2>Metastores</h2>{_table(ms_cols, ms_rows)}
  <h2>Workspaces</h2>{_table([lbl for _, lbl in _WS_COLS], ws_rows)}
  <h2>Azure Storage Footprint</h2>{_table(["Workspace","Storage Account","Resource Group","Location","HNS","Public Network"], fp_rows)}
</main>
<footer>Generated by SAT Scanner — Unity Catalog Inventory (fleet)</footer>
</body></html>"""

    path = output_dir / f"{_combined_prefix()}.html"
    path.write_text(doc, encoding="utf-8")
    return str(path)
